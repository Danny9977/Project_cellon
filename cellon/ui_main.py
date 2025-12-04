# ui_main.py
# ================== 크롬 크롤링 + 구글시트 + 쿠팡 OpenAPI (UI로 조회기간/헬스체크/주문현황) ==================
import sys
import os
import re
import json
import platform
import socket
import subprocess
import time
import io
import base64

from urllib.parse import urlparse, urlencode, quote  # canonical query 생성을 위해 quote 사용
from datetime import datetime, timedelta, timezone, date
from pathlib import Path

# ==== PyQt6 ====
from PyQt6.QtGui import QKeySequence, QShortcut
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton, QLabel, QTextEdit, QHBoxLayout, QSpinBox
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal

# ==== UI/OS/입력 ====
import pygetwindow as gw
import pyautogui
import pyperclip
from pynput import mouse
from pynput.mouse import Listener as MouseListener

# ==== Selenium ====
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ==== HTTP/HMAC ====
import requests
import hmac, hashlib

# ==== Excel ====
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ==== costco 크롤링 관련 ====
from PIL import Image

# config에서 필요한 값 import
from .config import *  # 가능하면 * 대신 필요한 것만 가져오는 쪽으로 나중에 정리

import gspread
from google.oauth2.service_account import Credentials

# 카테고리 매칭 모듈
from .core.category_matcher import CategoryMatcher


# 시트/쿠팡 API: 분리된 모듈
from .sheets_client import (
    SheetsClient,
    _cp_request,
    extract_paid_price_from_item,
)

# ui_main.py – config 및 category_ai
from .config import (
    today_fmt,
    label_for_domain,
    _a1_col,
    digits_only,
    is_macos,
    CATEGORY_EXCEL_DIR,
)

from .category_ai.category_worker import CategoryBuildWorker


# =========================
# 유틸 함수
# =========================
def safe_str(v) -> str:
    try:
        if callable(v):
            v = v()
    except Exception:
        pass
    try:
        return "" if v is None else str(v)
    except Exception:
        return ""


def today_iso() -> str:
    """YYYY-MM-DD 형식 오늘 날짜"""
    return datetime.now().strftime("%Y-%m-%d")


def is_port_open(host: str, port: int, timeout=0.3) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except OSError:
        return False


def selectors_for_url(url: str):
    host = urlparse(url).netloc if url else ""
    site_specific = []
    for key, sels in SITE_SELECTORS.items():
        if key in host:
            site_specific += sels
    seen, ordered = set(), []
    for sel in site_specific + DEFAULT_SELECTORS:
        if sel not in seen:
            seen.add(sel)
            ordered.append(sel)
    return ordered


def price_selectors_for_url(url: str):
    host = urlparse(url).netloc if url else ""
    site_specific = []
    for key, sels in SITE_PRICE_SELECTORS.items():
        if key in host:
            site_specific += sels
    general = [
        "#lItemPrice", ".lItemPrice", ".price .num", ".price-value", ".final_price",
        ".sale_price", ".price", "[data-testid='price']"
    ]
    seen, ordered = set(), []
    for sel in site_specific + general:
        if sel not in seen:
            seen.add(sel)
            ordered.append(sel)
    return ordered


def is_costco_url(url: str) -> bool:
    host = urlparse(url or "").netloc.lower()
    return "costco.co.kr" in host


def is_domeme_url(url: str) -> bool:
    host = urlparse(url or "").netloc.lower()
    return "domeme.domeggook.com" in host


def _mask(s: str, left: int = 4, right: int = 3) -> str:
    """키 마스킹: 앞/뒤 일부만 보이고 나머지는 * 처리"""
    s = str(s or "")
    if len(s) <= left + right:
        return "*" * len(s)
    return s[:left] + "*" * (len(s) - left - right) + s[-right:]

# =========================
# 코스트코 카테고리(브레드크럼) 추출
# =========================
COSTCO_CATEGORY_SELECTOR = (
    "div.container.bottom-header.BottomHeader.has-components "
    "ol.breadcrumb li a"
)

def extract_costco_category(driver) -> str | None:
    """
    코스트코 상품페이지에서 상단 breadcrumb 카테고리 텍스트를 추출.
    예) '메인 / 홈/키친 / 조리용품 / 쿡웨어' -> '홈/키친 / 조리용품 / 쿡웨어'
    """
    try:
        # breadcrumb 영역이 뜰 때까지 잠깐 대기
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, COSTCO_CATEGORY_SELECTOR)
            )
        )
    except TimeoutException:
        return None
    except Exception:
        return None

    try:
        els = driver.find_elements(By.CSS_SELECTOR, COSTCO_CATEGORY_SELECTOR)
        crumbs = [e.text.strip() for e in els if e.text.strip()]

        if not crumbs:
            return None

        # 맨 앞 '메인'은 보통 버리는 게 보기 좋음
        if crumbs[0] == "메인":
            crumbs = crumbs[1:]

        return " / ".join(crumbs) if crumbs else None
    except Exception as e:
        # 혹시 모를 예외는 로그에만 남기고 None 리턴
        print("코스트코 카테고리 추출 에러:", e)
        return None

# =========================
# 쿠팡 OpenAPI HMAC 서명 (성공 예제 기준)
# =========================
def _cp_build_query(params: dict | None) -> str:
    if not params:
        return ""
    return urlencode(params, doseq=True)  # quote_plus 방식 (공백→+)


def _cp_signed_headers_v2(
    method: str,
    path: str,
    sign_query: str,
    access_key: str,
    secret_key: str,
    *,
    signed_date: str | None = None,
    vendor_id: str | None = None
) -> dict:
    if signed_date is None:
        signed_date = datetime.now(timezone.utc).strftime("%y%m%dT%H%M%SZ")  # YYMMDDTHHMMSSZ
    message = f"{signed_date}{method.upper()}{path}{sign_query}"
    signature = hmac.new(
        secret_key.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    authorization = (
        f"CEA algorithm=HmacSHA256, access-key={access_key}, "
        f"signed-date={signed_date}, signature={signature}"
    )
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Authorization": authorization,
    }
    if vendor_id:
        headers["X-Requested-By"] = vendor_id
    return headers


def _build_ordersheets_params(
    date_from_utc: datetime,
    date_to_utc: datetime,
    status: str,
    max_per_page: int = 50
):
    d_from = date_from_utc.strftime("%Y-%m-%d")
    d_to = date_to_utc.strftime("%Y-%m-%d")
    primary = {
        "createdAtFrom": d_from,
        "createdAtTo": d_to,
        "status": status,
        "maxPerPage": max_per_page,
    }
    fallback = {
        "startTime": d_from,
        "endTime": d_to,
        "status": status,
        "maxPerPage": max_per_page,
    }
    return [primary, fallback]


def _try_ordersheets_with_variants(path: str, param_variants: list[dict]) -> dict:
    last_err = None
    for params in param_variants:
        try:
            return _cp_request("GET", path, params)
        except requests.HTTPError as e:
            resp = getattr(e, "response", None)
            status = getattr(resp, "status_code", None)
            body = ""
            try:
                body = (resp.text or "")[:500]
            except Exception:
                pass
            if status == 400 and "yyyy-MM-dd" in body:
                last_err = e
                continue
            raise
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise RuntimeError("ordersheets 호출 시도 실패: 유효한 파라미터 조합이 없습니다.")

# =========================
# 카테고리 마스터 생성 (QThread)
# =========================
def start_category_build(self):
    from cellon.category_ai.category_worker import CategoryBuildWorker
    from pathlib import Path

    # ★ 경로는 Danny님 환경에 맞게 직접 입력 (또는 config에서 불러와도 됨)
    #category_dir = Path("/Users/jeehoonkim/Desktop/category_excels")
    #category_dir = Path.home() / "Desktop" / "category_excels"  # 예시: 바탕화면의 category_excels 폴더 mac / win 공통
    category_dir = CATEGORY_EXCEL_DIR
    
    self._log("📂 카테고리 분석을 시작합니다...")

    # 스레드 생성
    self.cat_worker = CategoryBuildWorker(category_dir)
    self.cat_worker.progress.connect(self._on_cat_progress)
    self.cat_worker.finished.connect(self._on_cat_finished)

    self.cat_worker.start()

# ---- 콜백: 카테고리 빌드 완료 ----
def _on_cat_progress(self, percent: int, msg: str):
    self._log(f"{percent}% | {msg}")

# ---- 콜백: 카테고리 빌드 완료 ----
def _on_cat_finished(self, df):
    if df is None:
        self._log("❌ 카테고리 분석 실패")
        return

    self._log(f"✅ 카테고리 분석 완료 — 총 {len(df)}개 카테고리")
    # 여기서 df를 멤버 변수에 저장하거나 UI에 반영할 수 있음
    

# =========================
# 메인 앱 (UI + 로직)
# =========================
class ChromeCrawler(QWidget):
    clickDetected = pyqtSignal(int, int)

    # 테스트용 플래그: True로 두면 무조건 "다운로드 건너뛰고 캡처" 경로로 테스트
    FORCE_CAPTURE_TEST = False  # 테스트 미진행. 다운로드 우선순위 진행

    def __init__(self):
        super().__init__()
        self.setWindowTitle("크롬 크롤링 도구 (gspread + Coupang OpenAPI)")
        self.setGeometry(0, 0, 460, 580)

        # 카테고리 매칭용 매처 (kitchen 그룹 기준)
        self.cat_matcher = CategoryMatcher(group="kitchen", logger=self._log)
        
        # 등록상품명 캐시 (sellerProductId -> 등록상품명)
        self._cp_seller_name_cache: dict[str, str] = {}

        # 상태값
        self.target_title = None
        self.target_window = None
        self.driver = None
        self._listener = None
        self._waiting_click = False
        self._sheet_click_wait = False
        self._click_timer = None

        # 크롤 결과
        self.crawled_title = ""
        self.crawled_price = ""
        self.crawled_url = ""

        # 카테고리 관련 (원본/쿠팡)
        self.crawled_category = ""          # 코스트코/도매매 등 원본 카테고리 path
        self.coupang_category_id = ""       # 매칭된 쿠팡 category_id
        self.coupang_category_path = ""     # 매칭된 쿠팡 category_path
        
        # Google Sheets
        self.sheets = SheetsClient(
            SERVICE_ACCOUNT_JSON,
            SHEET_ID,
            WORKSHEET_NAME,
            self._log
        )
        self.row_index_cache = None

        # =========================
        # UI
        # =========================
        layout = QVBoxLayout()
        layout.setSpacing(6)
        layout.setContentsMargins(8, 8, 8, 8)

        self.label = QLabel("🖱 대상 윈도우: 없음")
        layout.addWidget(self.label)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        # 0) 카테고리 자동화 버튼 추가
        self.btn_cat_build = QPushButton("카테고리 분석 시작")
        self.btn_cat_build.clicked.connect(self.start_category_build)
        layout.addWidget(self.btn_cat_build)

        # 1) clear + Sheets 연결
        row_a = QHBoxLayout()
        self.btn_clear = QPushButton("Txt clear")
        self.btn_clear.clicked.connect(self.log.clear)
        row_a.addWidget(self.btn_clear)

        self.btn_sheets = new_btn = QPushButton("Sheets 연결")
        new_btn.clicked.connect(self.connect_sheets)
        row_a.addWidget(new_btn)
        layout.addLayout(row_a)

        # 2) 크롬(디버그) + 기존 창 연결 테스트
        row_b = QHBoxLayout()
        self.btn_launch = QPushButton("크롬(디버그) 실행")
        self.btn_launch.clicked.connect(self.launch_debug_chrome)
        row_b.addWidget(self.btn_launch)

        self.btn_test = QPushButton("기존 창 연결 테스트")
        self.btn_test.clicked.connect(self.test_attach_existing)
        row_b.addWidget(self.btn_test)
        layout.addLayout(row_b)

        # 3) 대상윈도우 + 기록
        row_c = QHBoxLayout()
        self.btn_select = QPushButton("대상윈도우 (Shift+Z)")
        self.btn_select.clicked.connect(self.select_target_window)
        row_c.addWidget(self.btn_select)

        self.btn_record = QPushButton("기록 (Shift+X)")
        self.btn_record.clicked.connect(self.record_data)
        row_c.addWidget(self.btn_record)
        layout.addLayout(row_c)

        # 4) STOP + 네이버(최저가)
        row_d = QHBoxLayout()
        self.btn_stop = QPushButton("STOP (프로그램 off)")
        self.btn_stop.clicked.connect(self.close)
        row_d.addWidget(self.btn_stop)

        self.btn_health = QPushButton("네이버 (최저가))")
        self.btn_health.clicked.connect(self.naver_check)
        row_d.addWidget(self.btn_health)
        layout.addLayout(row_d)

        # 5) 금일 상품 갯수 + (통합) 쿠팡 키 확인 + 헬스체크
        row_e = QHBoxLayout()

        self.lbl_today_count = QLabel("금일 상품 갯수 : 0")
        row_e.addWidget(self.lbl_today_count)

        self.btn_calc_today = QPushButton("상품개수계산")
        self.btn_calc_today.clicked.connect(self.update_today_product_count)
        row_e.addWidget(self.btn_calc_today)

        self.btn_cp_keyhealth = QPushButton("쿠팡 키+헬스체크")
        self.btn_cp_keyhealth.clicked.connect(self.coupang_key_and_health)
        row_e.addWidget(self.btn_cp_keyhealth)

        layout.addLayout(row_e)

        # 6) 하단: 확인기간 + 스핀박스 + (우측) 쿠팡 주문현황
        row_z = QHBoxLayout()
        self.lbl_days = QLabel("확인기간 :")
        row_z.addWidget(self.lbl_days)

        self.spin_days = QSpinBox()
        self.spin_days.setRange(1, 365)
        self.spin_days.setValue(DEFAULT_LOOKBACK_DAYS)
        self.spin_days.setSuffix(" 일")
        self.spin_days.setSingleStep(1)
        row_z.addWidget(self.spin_days)

        row_z.addStretch(1)

        self.btn_coupang = QPushButton("쿠팡 주문현황")
        self.btn_coupang.clicked.connect(self.coupang_orders)
        row_z.addWidget(self.btn_coupang)

        self.btn_order_settle = QPushButton("주문정리")
        self.btn_order_settle.clicked.connect(self.settle_orders)
        row_z.addWidget(self.btn_order_settle)

        self.btn_google_underline = QPushButton("구글시트 밑줄")
        self.btn_google_underline.clicked.connect(self.google_underline)
        row_z.addWidget(self.btn_google_underline)

        layout.addLayout(row_z)

        for btn in (
            self.btn_clear, self.btn_sheets, self.btn_launch, self.btn_test,
            self.btn_select, self.btn_record, self.btn_stop, self.btn_health,
            self.btn_cp_keyhealth, self.btn_coupang, self.btn_order_settle, self.btn_google_underline
        ):
            btn.setMinimumHeight(28)
            btn.setStyleSheet("QPushButton { padding: 4px 8px; }")

        self._log(
            "ℹ️ 사용법:\n"
            "1) [Sheets 연결] → [크롬(디버그) 실행] 후 대상 페이지를 엽니다.\n"
            "2) [대상윈도우] 클릭 → 안내에 따라 '본문'을 클릭(5초 내).\n"
            "3) 하단 [확인기간] 일수를 설정 후 [쿠팡 주문현황]으로 조회합니다.\n"
            "4) [쿠팡 키+헬스체크] 버튼으로 키/서명/경로 정상 여부를 점검합니다.\n"
        )

        self.setLayout(layout)

        # 단축키
        QShortcut(QKeySequence("Shift+Z"), self, activated=self.select_target_window)
        QShortcut(QKeySequence("Shift+X"), self, activated=self.record_data)

        # 전역 클릭 시그널
        self.clickDetected.connect(self._handle_click_on_main)

        # 자동 초기화
        QTimer.singleShot(300, self._startup_sequence)
        
        # 카테고리 마스터 관련 상태
        self.category_worker: CategoryBuildWorker | None = None
        self.category_master_df = None  # 필요하면 나중에 다른 곳에서 참조

    # --------------------------
    # 이하 메서드는 기존 main_app.py 의 ChromeCrawler 메서드들을
    # 1:1 그대로 가져온 것입니다.
    # (connect_sheets, launch_debug_chrome, crawl_data, coupang_orders 등등)
    # --------------------------

    def _log(self, msg: str):
        self.log.append(msg)
        print(msg)

    # ... (여기부터는 기존 ChromeCrawler 의 모든 메서드들을
    #      main_app.py 에서 그대로 복사해 오시면 됩니다.
    #      이미 위에서 전체 코드를 보여드렸으니, 그대로 붙여넣으셔도 됩니다.)

    # ---------- 자동 시작 시퀀스 ----------
    def _startup_sequence(self):
        """프로그램 시작 시 자동으로 Sheets 연결 / 크롬 연결을 시도."""
        self._log("🚀 시작: 자동 초기화 시퀀스 실행")
        try:
            self.connect_sheets()
        except Exception as e:
            self._log(f"⚠️ 자동 시트 연결 실패: {e}")

        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결 → '기존 창 연결 테스트' 수행")
            ok = self._attach_existing_ok()
            if ok:
                self.test_attach_existing()
            else:
                self._log("ℹ️ 기존 창 연결 실패 → '크롬(디버그) 실행' 수행")
                self.launch_debug_chrome()
 
        # === 카테고리 마스터 생성 시작 ===
    
    def start_category_build(self):
        """
        [카테고리 분석 시작] 버튼 클릭 시 호출될 메서드.
        - QThread 워커를 띄워서 엑셀들을 분석
        - 진행 상황을 log 창에 5% 단위로 표시 (category_loader가 이미 5% 단위로 콜백 호출)
        - UI는 동안에도 다른 버튼/기능 사용 가능
        """
        # 이미 돌고 있으면 중복 실행 방지
        if self.category_worker is not None and self.category_worker.isRunning():
            self._log("ℹ️ 카테고리 분석이 이미 진행 중입니다.")
            return

        # config.py 에 정의된 CATEGORY_EXCEL_DIR 사용
        category_dir = CATEGORY_EXCEL_DIR

        self._log(f"📂 카테고리 엑셀 분석 시작: {category_dir}")
        self._log("⏳ 엑셀 파일을 분석하며 캐시를 갱신합니다. (진행률은 5% 단위로 표시)")

        # 워커 생성
        self.category_worker = CategoryBuildWorker(category_dir, parent=self)
        self.category_worker.progress.connect(self._on_category_progress)
        self.category_worker.finished.connect(self._on_category_finished)
        self.category_worker.error.connect(self._on_category_error)

        # 백그라운드에서 실행
        self.category_worker.start()

    def _on_category_progress(self, percent: int, message: str):
        """
        워커에서 progress_cb로 호출한 진행 상황을 받아서 log 창에 출력.
        """
        # percent 를 앞에 붙여서 로그 표시
        self._log(f"[카테고리] {percent}% - {message}")

    def _on_category_finished(self, df):
        """
        워커가 정상 완료되었을 때 호출.
        df 는 category_master DataFrame.
        """
        self.category_worker = None
        self.category_master_df = df
        try:
            n = len(df) if df is not None else 0
        except Exception:
            n = 0
        self._log(f"✅ 카테고리 마스터 생성 완료 (총 {n}개 카테고리)")

    def _on_category_error(self, msg: str):
        """
        워커 내부에서 예외 발생 시 호출.
        """
        self.category_worker = None
        self._log(f"❌ 카테고리 마스터 생성 중 오류: {msg}")

    
    # ---------- 구글시트 연결 ----------
    def connect_sheets(self):
        """구글시트 연결 버튼 동작용 메서드"""
        try:
            self.sheets.connect()
            self._log("✅ Sheets 연결 완료")
        except Exception as e:
            self._log(f"❌ Sheets 연결 실패: {e}")

    # ---------- 디버그 크롬 실행 ----------
    def launch_debug_chrome(self):
        try:
            # 이미 디버그 포트가 열려 있으면 새로 띄우지 않음
            if is_port_open("127.0.0.1", DEBUGGER_PORT):
                self._log(f"ℹ️ 디버그 포트 {DEBUGGER_PORT} 이미 열림. 기존 창에 연결하세요.")
                return

            chrome_bin = None
            for p in CHROME_PATHS:
                if os.path.exists(p):
                    chrome_bin = p
                    break

            if chrome_bin is None:
                self._log("⚠️ Chrome 실행 파일을 찾지 못했습니다.")
                return

            Path(USER_DATA_DIR).mkdir(parents=True, exist_ok=True)

            cmd = [
                chrome_bin,
                f"--remote-debugging-port={DEBUGGER_PORT}",
                f"--user-data-dir={USER_DATA_DIR}",
                "--no-first-run",
                "--no-default-browser-check",
            ]

            subprocess.Popen(
                cmd,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )

            # 디버그 포트가 열릴 때까지 최대 5초 정도 대기
            for _ in range(25):
                if is_port_open("127.0.0.1", DEBUGGER_PORT):
                    self._log(f"✅ 디버깅 모드 Chrome 실행됨 (포트 {DEBUGGER_PORT}).")
                    return
                time.sleep(0.2)

            self._log("⚠️ 디버그 포트 연결 확인 실패")
        except Exception as e:
            self._log(f"[오류] 크롬(디버그) 실행 실패: {e}")


            self._log(f"크롬 디버그 실행 실패: {e}")

    # ---------- 기존 디버그 크롬 연결 테스트 ----------            
    def test_attach_existing(self):
        try:
            driver = self._attach_driver()
            tabs_info = []
            for h in driver.window_handles:
                driver.switch_to.window(h)
                tabs_info.append(f"- {safe_str(driver.title).strip()} | {safe_str(driver.current_url).strip()}")
            msg = "🔗 디버그 세션 탭 목록:\n" + ("\n".join(tabs_info) if tabs_info else "(없음)")
            self._log(msg)
        except Exception as e:
            self._log(f"[오류] 기존 창 연결 테스트 실패: {e}")

    # ---------- 기존 디버그 크롬 연결 테스트 ----------
    def _attach_existing_ok(self) -> bool:
        """이미 떠 있는 디버그 크롬에 정상 연결 가능한지 간단 체크."""
        try:
            if not is_port_open("127.0.0.1", DEBUGGER_PORT):
                self._log("ℹ️ 디버그 포트가 열려 있지 않음")
                return False
            driver = self._attach_driver()
            _ = driver.window_handles
            self._log("✅ 기존 창 연결 OK")
            return True
        except Exception as e:
            self._log(f"ℹ️ 기존 창 연결 실패: {e}")
            return False

    # ---------- 디버그 크롬 연결 ----------
    def _attach_driver(self):
        """
        이미 디버그 모드로 떠 있는 Chrome 에 Selenium 을 붙이는 함수.
        - 디버그 포트가 안 떠 있으면 RuntimeError 발생.
        """
        if not is_port_open("127.0.0.1", DEBUGGER_PORT):
            raise RuntimeError("디버그 포트가 열려 있지 않습니다. 먼저 '크롬(디버그) 실행'을 눌러주세요.")

        if self.driver:
            return self.driver

        options = webdriver.ChromeOptions()
        options.debugger_address = f"127.0.0.1:{DEBUGGER_PORT}"
        self.driver = webdriver.Chrome(options=options)
        return self.driver

    # ---------- 네이버 최저가 체크 ----------    
    def naver_check(self):
        self._open_naver_shopping_with_title(sort_low_price=True)

    # ---------- 네이버 최저가 열기 ----------
    def _open_naver_shopping_with_title(self, sort_low_price: bool = True):
        try:
            title = (self.crawled_title or "").strip()
            if not title:
                self._log("ℹ️ 제목이 없어 네이버 쇼핑 검색을 생략합니다.")
                return

            driver = self._attach_driver()
            from urllib.parse import quote_plus

            base_url = "https://search.shopping.naver.com/search/all"
            q = f"query={quote_plus(title)}"
            sort = "sort=price_asc" if sort_low_price else "sort=rel"
            search_url = f"{base_url}?{q}&{sort}"

            # 새 탭으로 네이버 쇼핑 열기
            driver.execute_script("window.open(arguments[0], '_blank');", search_url)
            driver.switch_to.window(driver.window_handles[-1])
            self._log(f"🟢 네이버 쇼핑 검색 탭 오픈(낮은가격순 시도): {search_url}")

            if not sort_low_price:
                return

            # 페이지 로딩 대기
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
            except Exception:
                pass

            # URL에 sort=price_asc 가 이미 붙어 있으면 그대로 사용
            if "sort=price_asc" in (driver.current_url or ""):
                return

            # 정렬 UI에서 ‘낮은가격순/가격낮은순’ 버튼 찾아서 클릭 시도
            click_js = r"""
            const keywords = ['낮은가격순', '가격낮은순'];
            function clickByText(nodes) {
            for (const el of nodes) {
                try {
                const t = (el.innerText || el.textContent || '').trim();
                if (!t) continue;
                for (const k of keywords) {
                    if (t.includes(k)) { el.click(); return true; }
                }
                } catch (e) {}
            }
            return false;
            }
            const order = ['button','a','span','div','li'];
            for (const tag of order) {
            const list = document.querySelectorAll(tag);
            if (clickByText(list)) return true;
            }
            return false;
            """

            clicked = driver.execute_script(click_js)
            if clicked:
                self._log("✅ 정렬 UI 클릭으로 '낮은 가격순' 적용 시도")
                try:
                    WebDriverWait(driver, 5).until(
                        lambda d: "price_asc" in (d.current_url or "")
                    )
                except Exception:
                    pass
            else:
                self._log("⚠️ 정렬 UI 요소를 찾지 못했습니다. (페이지 UI 변경 가능)")
        except Exception as e:
            self._log(f"⚠️ 네이버 쇼핑 검색/정렬 처리 실패: {e}")
        
            
    def select_target_window(self):
        # 대상윈도우 버튼을 누를 때 금일 상품 갯수 자동 계산
        self.update_today_product_count()
        
        self._log("🖱 **크롤링할 크롬 탭의 본문**을 클릭해 주세요. (5초 내)")
        self.label.setText("🔍 본문을 클릭하세요 (주소창 X). 5초 내 미클릭 시 경고.")

        self.showMinimized()
        self._waiting_click = True
        self._sheet_click_wait = False

        if self._click_timer is None:
            self._click_timer = QTimer(self)
            self._click_timer.setSingleShot(True)
            self._click_timer.timeout.connect(self._on_click_timeout_select)
        self._click_timer.start(CLICK_TIMEOUT_MS_SELECT)

        def on_click(x, y, button, pressed):
            if pressed and self._waiting_click:
                self.clickDetected.emit(int(x), int(y))
        self._listener = MouseListener(on_click=on_click)
        self._listener.start()

    def _on_click_timeout_select(self):
        if not self._waiting_click:
            return
        self._waiting_click = False
        try:
            if self._listener: self._listener.stop()
        except Exception:
            pass
        finally:
            self._listener = None
        self._log("⏰ 5초 내 클릭이 감지되지 않았습니다. 다시 [대상윈도우]를 눌러 본문을 클릭하세요.")

    def _handle_click_on_main(self, x: int, y: int):
        if not self._waiting_click:
            return
        self._waiting_click = False
        if self._click_timer and self._click_timer.isActive():
            self._click_timer.stop()
        try:
            if self._listener: self._listener.stop()
        except Exception:
            pass
        finally:
            self._listener = None

        wins_at = self._gw_get_windows_at(x, y)
        win = wins_at[0] if wins_at else None
        picked_title = safe_str(getattr(win, "title", "")) if win else ""
        if not picked_title:
            self._log("❌ 클릭 지점에서 활성 창 제목을 찾지 못했습니다. 본문 클릭/권한 확인.")
            return

        self.target_window = win
        self.target_title = picked_title
        self.label.setText(f"🎯 대상 윈도우: {self.target_title}")

        self.showNormal(); self.raise_(); self.activateWindow()
        self.crawl_data()

    def _gw_get_windows_at(self, x: int, y: int):
        try:
            fn = getattr(gw, "getWindowsAt", None)
            if callable(fn):
                return fn(x, y)
        except Exception:
            pass
        try:
            active = getattr(gw, "getActiveWindow", lambda: None)()
            return [active] if active else []
        except Exception:
            return []

    def crawl_data(self):
        if not self.target_title:
            self._log("⚠️ 대상 탭이 선택되지 않았습니다.")
            return
        try:
            if self.target_window:
                try:
                    self.target_window.activate(); time.sleep(0.2)
                except Exception:
                    pass

            driver = self._attach_driver()

            self._log("🧭 탭 매칭: URL패턴 → 제목 포함")
            end_time = time.time() + 5.0
            target_handle = None

            if URL_PATTERNS:
                while time.time() < end_time and not target_handle:
                    for h in driver.window_handles:
                        driver.switch_to.window(h)
                        if any(p in (driver.current_url or "") for p in URL_PATTERNS):
                            target_handle = h; break
                    if not target_handle:
                        time.sleep(0.2)

            if not target_handle:
                end_time2 = time.time() + 5.0
                raw_want = safe_str(self.target_title).strip()

                # 윈도우 제목에서 ' - ' 뒤에 붙는 브라우저 이름 제거 (예: " - Google Chrome")
                want_base = raw_want.split(" - ")[0].strip() if raw_want else ""

                while time.time() < end_time2 and not target_handle:
                    for h in driver.window_handles:
                        driver.switch_to.window(h)
                        page_title = safe_str(driver.title).strip()
                        page_base = page_title.split(" - ")[0].strip() if page_title else ""

                        # 1) 전체 제목 포함 여부
                        if raw_want and raw_want in page_title:
                            target_handle = h
                            break

                        # 2) '앞부분만' 비교 (상품명 부분만 비교)
                        if want_base and want_base in page_base:
                            target_handle = h
                            break

                    if not target_handle:
                        time.sleep(0.2)


            if not target_handle:
                self._log("❌ 5초 내 '대상 탭'을 찾지 못했습니다.")
                return

            driver.switch_to.window(target_handle)

            current_url = safe_str(driver.current_url).strip()
            self.crawled_url = current_url
            self._log(f"🔗 URL: {current_url}")

            blocked = ("chrome://", "chrome-extension://", "edge://", "about:", "data:")
            if any(current_url.startswith(s) for s in blocked) or current_url.lower().endswith(".pdf"):
                self._log("❌ 이 페이지는 DOM 접근이 제한됩니다.")
                return

            try:
                WebDriverWait(driver, 3).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
            except Exception:
                pass
            
            # === 코스트코 카테고리(breadcrumb) 추출 ===
            self.crawled_category = ""
            self.coupang_category_id = ""
            self.coupang_category_path = ""
            
            if is_costco_url(current_url):
                try:
                    cat = extract_costco_category(driver)
                    if cat:
                        self.crawled_category = cat
                        self._log(f"📂 원본 카테고리(코스트코): {self.crawled_category}")
                    else:
                        self._log("📂 원본 카테고리(코스트코): (없음 또는 추출 실패)")
                except Exception as e:
                    self._log(f"⚠️ 코스트코 카테고리 추출 중 오류: {e}")
            

            title_value = ""
            wait = WebDriverWait(driver, 5)
            for sel in selectors_for_url(current_url):
                try:
                    el = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, sel)))
                    title_value = (el.text or "").strip()
                    if title_value:
                        break
                except Exception:
                    continue
            if not title_value:
                try:
                    el = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.TAG_NAME, "h1")))
                    title_value = (el.text or "").strip()
                except Exception:
                    title_value = ""
            self.crawled_title = title_value
            self._log(f"🟢 제목: {self.crawled_title or '(없음)'}")

            price_digits = ""
            wait_p = WebDriverWait(driver, 3)
            for sel in price_selectors_for_url(current_url):
                try:
                    el = wait_p.until(EC.visibility_of_element_located((By.CSS_SELECTOR, sel)))
                    txt = (el.text or "").strip()
                    if not txt:
                        txt = (driver.execute_script(
                            "const e=document.querySelector(arguments[0]); return e?(e.innerText||e.textContent||''):'';", sel
                        ) or "").strip()
                    if txt:
                        price_digits = re.sub(r"[^0-9]", "", txt)
                        if price_digits:
                            break
                except Exception:
                    continue
            if not price_digits:
                try:
                    body = (driver.execute_script(
                        "return (document.body && document.body.innerText) ? document.body.innerText : '';"
                    ) or "")
                    m = re.search(r'([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)\s*원', body)
                    if not m:
                        m = re.search(r'₩\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)', body)
                    if m:
                        price_digits = re.sub(r"[^0-9]", "", m.group(1))
                except Exception:
                    pass
            self.crawled_price = price_digits
            self._log(f"💰 가격(숫자만): {self.crawled_price or '(없음)'}")

            self._log("—" * 40)
            self._log(f"제목: {self.crawled_title or '(없음)'}")
            self._log(f"카테고리(원본): {self.crawled_category or '(없음)'}")   # ← 추가
            self._log(f"가격(숫자만): {self.crawled_price or '(없음)'}")
            self._log(f"URL: {self.crawled_url or '(없음)'}")
            self._log("—" * 40)

            # === 쿠팡 카테고리 매칭 ===
            try:
                # 1) source 판단 (현재는 costco/domemae만 사용)
                source = ""
                if is_costco_url(current_url):
                    source = "costco"
                elif is_domeme_url(current_url):
                    source = "domemae"
                # TODO: owner 클랜 붙이면 elif "owner" 추가

                self._log("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
                self._log("[UI] 카테고리 매칭 진입")
                self._log(f"  - source={source or '(빈 값)'}")
                self._log(f"  - 원본 카테고리(path)='{self.crawled_category or ''}'")
                self._log(f"  - 상품명='{self.crawled_title or ''}'")

                if source:
                    match = self.cat_matcher.match_category(
                        source=source,
                        source_category_path=self.crawled_category or "",
                        product_name=self.crawled_title or "",
                        brand=None,
                        extra_text=None,
                    )

                    if not match:
                        self._log("  ❌ CategoryMatcher가 None 또는 빈 dict를 반환했습니다.")
                        self.coupang_category_id = ""
                        self.coupang_category_path = ""
                    else:
                        self.coupang_category_id = match.get("category_id") or ""
                        self.coupang_category_path = match.get("category_path") or ""

                        used_llm = match.get("used_llm")
                        meta_key = match.get("meta_key")
                        num_candidates = match.get("num_candidates")
                        reason = match.get("reason")

                        self._log("  🔎 [카테고리 매칭 결과 요약]")
                        self._log(f"    - category_id={self.coupang_category_id or '(없음)'}")
                        self._log(f"    - category_path={self.coupang_category_path or '(없음)'}")
                        self._log(f"    - meta_key={meta_key}")
                        self._log(f"    - num_candidates={num_candidates}")
                        self._log(f"    - used_llm={used_llm}")
                        if reason:
                            self._log(f"    - reason={reason}")
                else:
                    self._log("  ℹ️ 현재 URL은 costco/domemae가 아니라서 카테고리 매칭을 건너뜁니다.")
                self._log("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            except Exception as e:
                self._log(f"⚠️ [UI] 카테고리 매칭 중 예외 발생: {e}")


            self._log("📝 크롤 완료: 시트에 바로 기록합니다.")
            self.record_data()

        except Exception as e:
            self._log(f"[오류] 크롤링 실패: {e}")

    # ---------- 구글시트 창 앞으로 가져오기 ----------
    def _bring_sheet_to_front(self):
        """
        현재 사용 중인 스프레드시트(SHEET_ID)를 브라우저에서 앞으로 띄워준다.
        - macOS: AppleScript 로 Chrome 탭을 찾아서 활성화
        - 기타 OS: 타이틀로 대충 찾고, 없으면 새 창/탭으로 open
        """
        try:
            sheet_url_prefix = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}"

            import platform
            if platform.system().lower() == "darwin":
                # macOS: AppleScript 로 크롬 탭 포커싱
                osa = f'''
                tell application "Google Chrome"
                    set thePrefix to "{sheet_url_prefix}"
                    set foundWin to missing value
                    set foundIdx to -1
                    repeat with w in windows
                        set i to 0
                        repeat with t in tabs of w
                            set i to i + 1
                            if (URL of t) starts with thePrefix then
                                set foundWin to w
                                set active tab index of w to i
                                set index of w to 1
                                activate
                                return
                            end if
                        end repeat
                    end repeat
                    open location thePrefix & "/edit"
                    activate
                end tell
                '''
                subprocess.run(["osascript", "-e", osa], check=False)
            else:
                # 윈도우/리눅스: 제목에 "Google Sheets" 들어가는 창을 대충 찾아봄
                titles = []
                try:
                    titles = gw.getAllTitles()
                except Exception:
                    pass

                cand = [
                    t for t in titles
                    if isinstance(t, str) and ("Google Sheets" in t or "스프레드시트" in t)
                ]
                if cand:
                    wlist = gw.getWindowsWithTitle(cand[0])
                    if wlist:
                        try:
                            wlist[0].activate()
                        except Exception:
                            pass

                # 그래도 없으면 그냥 새 탭으로 open
                try:
                    import webbrowser
                    webbrowser.open(sheet_url_prefix + "/edit", new=0, autoraise=True)
                except Exception:
                    pass

        except Exception as e:
            self._log(f"⚠️ 시트 창 활성화 실패: {e}")

    # ---------- 시트 기록(핵심) ----------
    def _write_row_to_first_empty_a(self):
        if self.sheets.ws is None:
            self._log("⚠️ 먼저 [Sheets 연결]을 눌러 구글시트에 연결해 주세요.")
            return

        target_row = self.sheets.find_first_empty_row_in_col_a_from_top()

        COLS = {c:i for i,c in enumerate(
            ["A","B","C","D","E","F","G","H","I","J",
             "K","L","M","N","O","P","Q","R","S","T",
             "U","V","W","X","Y"], start=1)}

        a_index = str(self.sheets.get_next_index())

        row_buffer = [""] * 25
        row_buffer[COLS["A"]-1] = a_index
        row_buffer[COLS["B"]-1] = today_fmt()
        row_buffer[COLS["C"]-1] = label_for_domain(self.crawled_url)
        row_buffer[COLS["F"]-1] = self.crawled_title or ""
        row_buffer[COLS["H"]-1] = self.crawled_price or ""
        row_buffer[COLS["I"]-1] = FIXED_CONST_FEE
        row_buffer[COLS["J"]-1] = f"=H{target_row}+I{target_row}"
        row_buffer[COLS["K"]-1] = "10.8%"
        row_buffer[COLS["M"]-1] = f"=J{target_row}+(R{target_row}*(K{target_row}*1.1))"
        row_buffer[COLS["N"]-1] = f"=O{target_row}/R{target_row}"
        row_buffer[COLS["O"]-1] = f"=R{target_row}-M{target_row}+K{target_row}-P{target_row}+L{target_row}"
        row_buffer[COLS["R"]-1] = f"=Q{target_row}"
        row_buffer[COLS["S"]-1] = f"=R{target_row}/1.1"
        row_buffer[COLS["T"]-1] = f"=S{target_row}*1.1-S{target_row}"
        row_buffer[COLS["V"]-1] = self.crawled_url or ""

        rng = f"A{target_row}:Y{target_row}"
        self.sheets.ws.update(values=[row_buffer], range_name=rng, value_input_option="USER_ENTERED")
        self._log(f"✅ 행 {target_row} (A..Y)에 기록 완료")

        try:
            if self.crawled_url:
                pyperclip.copy(self.crawled_url)
                self._log("📋 현재 상품 URL을 클립보드에 복사했습니다.")
        except Exception as e:
            self._log(f"⚠️ 클립보드 복사 실패: {e}")

        self._bring_sheet_to_front()

    # ---------- 코스트코 → sellertool_upload.xlsm 기록 ----------
    def _write_costco_to_seller_excel(self):
        """
        코스트코 상품(현재 self.crawled_title / self.crawled_price / self.crawled_url)을
        sellertool_upload.xlsm 에 다음 규칙으로 기록한다.

        - A열 : 카테고리 명 (지금은 공란)
        - B열 : 등록상품명 (크롤링한 상품명)
        - C열 : 오늘 날짜 'YYYY-MM-DD'
        - D열 : 공란
        - E열 : '새상품'
        - F열 : 공란
        - G열 : 등록상품명의 첫 단어
        - H열 : 등록상품명의 첫 단어
        - I열 ~ Z열 : 공란

        - BJ : 5만원 이하 -> 코스트코 가격 * 1.3
               5만원 초과 ~ 10만원 이하 -> 코스트코 가격 * 1.2
               10만원 초과 -> 원가 그대로
        - BL : BJ * 1.05
        - BM : 999
        - BN : 2
        - BX : '상세정보별도표기'
        - CK : '기타재화'
        - CZ : 행번호.png (예: 5행이면 '5.png')
        """

        if not self.crawled_title:
            self._log("⚠️ 코스트코 엑셀 기록: 상품명이 없습니다.")
            return None

        if not os.path.exists(SELLERTOOL_XLSM_PATH):
            self._log(f"❌ 코스트코 엑셀 기록: 파일을 찾지 못했습니다 → {SELLERTOOL_XLSM_PATH}")
            return None

        try:
            self._log(f"📂 엑셀 열기: {SELLERTOOL_XLSM_PATH}")
            wb = load_workbook(SELLERTOOL_XLSM_PATH, keep_vba=True)
        except Exception as e:
            self._log(f"❌ 엑셀 로드 실패: {e}")
            return None

        # 시트 선택
        try:
            if SELLERTOOL_SHEET_NAME in wb.sheetnames:
                ws = wb[SELLERTOOL_SHEET_NAME]
            else:
                ws = wb[wb.sheetnames[0]]
                self._log(f"⚠️ 시트 '{SELLERTOOL_SHEET_NAME}'를 찾지 못해 첫 번째 시트('{ws.title}')를 사용합니다.")
        except Exception as e:
            self._log(f"❌ 시트 선택 실패: {e}")
            return None

        # ==== 1) 입력할 행 찾기 (3행부터) ====
        start_row = 3
        row_idx = start_row
        while True:
            cell_val = ws.cell(row=row_idx, column=2).value  # B열
            if cell_val is None or str(cell_val).strip() == "":
                break
            row_idx += 1

        # ==== 2) 공통 데이터 준비 ====
        full_name = self.crawled_title.strip()
        words = full_name.split()
        first_word = words[0] if words else ""

        # 가격(숫자)
        try:
            base_price = int(digits_only(self.crawled_price))
        except Exception:
            base_price = 0

        # BJ 계산
        bj_price = 0
        if base_price > 0:
            if base_price <= 50000:
                bj_price = int(round(base_price * 1.3))
            elif base_price <= 100000:
                bj_price = int(round(base_price * 1.2))
            else:
                bj_price = base_price

        # BL = BJ * 1.05
        bl_price = int(round(bj_price * 1.05)) if bj_price > 0 else 0

        today_str = today_iso()

        # ==== 3) A~Z 채우기 ====
        ws.cell(row=row_idx, column=1).value  = ""         # A
        ws.cell(row=row_idx, column=2).value  = full_name  # B
        ws.cell(row=row_idx, column=3).value  = today_str  # C
        ws.cell(row=row_idx, column=4).value  = ""         # D
        ws.cell(row=row_idx, column=5).value  = "새상품"   # E
        ws.cell(row=row_idx, column=6).value  = ""         # F
        ws.cell(row=row_idx, column=7).value  = first_word # G
        ws.cell(row=row_idx, column=8).value  = first_word # H
        ws.cell(row=row_idx, column=9).value  = ""         # I

        for col in range(10, 27):                          # J~Z
            ws.cell(row=row_idx, column=col).value = ""

        # ==== 4) 확장 열 채우기 ====
        col_BJ = column_index_from_string("BJ")
        col_BL = column_index_from_string("BL")
        col_BM = column_index_from_string("BM")
        col_BN = column_index_from_string("BN")
        col_BX = column_index_from_string("BX")
        col_CK = column_index_from_string("CK")
        col_CZ = column_index_from_string("CZ")

        ws.cell(row=row_idx, column=col_BJ).value = bj_price
        ws.cell(row=row_idx, column=col_BL).value = bl_price
        ws.cell(row=row_idx, column=col_BM).value = 999
        ws.cell(row=row_idx, column=col_BN).value = 2
        ws.cell(row=row_idx, column=col_BX).value = "상세정보별도표기"
        ws.cell(row=row_idx, column=col_CK).value = "기타재화"
        ws.cell(row=row_idx, column=col_CZ).value = f"{row_idx}.png"

        try:
            wb.save(SELLERTOOL_XLSM_PATH)
            self._log(f"✅ 코스트코 상품 기록 완료 → 행 {row_idx}")
        except Exception as e:
            self._log(f"❌ 엑셀 저장 실패: {e}")
            return None

        # URL 클립보드 (선택)
        try:
            if self.crawled_url:
                pyperclip.copy(self.crawled_url)
                self._log("📋 코스트코 상품 URL을 클립보드에 복사했습니다.")
        except Exception as e:
            self._log(f"⚠️ 클립보드 복사 실패: {e}")

        return row_idx

    def _capture_costco_image(self, row_idx: int):
        """
        코스트코 상품 이미지 여러 장 저장 (다운로드 우선, 실패 시 캡처 백업)
        - 메인(가장 큰) 이미지는 건너뛰고
        - 그 아래에 있는 썸네일들만 저장
        - 첫 번째 저장 이미지는 row_idx.png,
        이후는 row_idx-1.png, row_idx-2.png ...
        """
        try:
            driver = self._attach_driver()
        except Exception as e:
            self._log(f"❌ 코스트코 이미지 처리: 드라이버 연결 실패: {e}")
            return

        try:
            # 코스트코 상품 영역의 이미지들(메인 + 썸네일)
            raw_imgs = driver.find_elements(By.CSS_SELECTOR, "picture img")
        except Exception as e:
            self._log(f"❌ 이미지 요소 검색 실패: {e}")
            return

        if not raw_imgs:
            self._log("⚠️ 처리할 picture img 요소를 찾지 못했습니다. 셀렉터를 점검해 주세요.")
            return

        # ====== 1) 화면상 크기 기준으로 '메인(히어로)' 이미지 추정 ======
        sized_imgs = []
        hero_el = None
        max_area = 0.0

        for el in raw_imgs:
            try:
                w, h = driver.execute_script(
                    "const r = arguments[0].getBoundingClientRect();"
                    "return [r.width, r.height];",
                    el,
                ) or (0, 0)
            except Exception:
                w, h = 0, 0

            area = float(w) * float(h)
            sized_imgs.append((el, w, h, area))

            if area > max_area:
                max_area = area
                hero_el = el

        if hero_el is not None:
            self._log(
                f"🧩 코스트코 메인 이미지(가장 큰 img)를 area={max_area:.1f} 로 추정 → 다운로드에서 제외"
            )

        # ====== 2) 날짜별 폴더 생성 (YYYYMMDD) ======
        base_dir = Path(SELLERTOOL_XLSM_PATH).parent
        date_folder = datetime.now().strftime("%Y%m%d")
        save_dir = base_dir / date_folder
        save_dir.mkdir(parents=True, exist_ok=True)

        saved_count = 0

        # 필터 임계값 (코스트코 전용 완화 값)
        # - natural 160x160 썸네일은 통과
        NAT_MIN_W, NAT_MIN_H = 120, 120      # 원본 크기가 이보다 작으면 진짜 작은 아이콘으로 봄
        VIEW_MIN_W, VIEW_MIN_H = 120, 120    # 화면 표시 크기가 이보다 작으면 건너뜀

        for el, view_w, view_h, area in sized_imgs:
            # 1) 메인(가장 큰) 이미지는 건너뜀
            if el is hero_el:
                self._log("↩️ 메인 상품 이미지는 건너뜁니다.")
                continue

            # 2) 화면 표시 크기가 너무 작은 아이콘은 건너뜀
            if view_w < VIEW_MIN_W or view_h < VIEW_MIN_H:
                self._log(
                    f"↩️ 너무 작은 화면 이미지(view {view_w:.0f}x{view_h:.0f}) → 건너뜀"
                )
                continue

            # 3) 원본 크기 기준으로도 너무 작은 것은 건너뜀
            try:
                nat_w, nat_h = driver.execute_script(
                    "return [arguments[0].naturalWidth, arguments[0].naturalHeight];",
                    el,
                ) or (0, 0)
            except Exception:
                nat_w, nat_h = 0, 0

            if nat_w < NAT_MIN_W or nat_h < NAT_MIN_H:
                self._log(
                    f"↩️ 너무 작은 원본 이미지(natural {nat_w}x{nat_h}) → 건너뜀"
                )
                continue

            # ===== 파일명 구성 =====
            if saved_count == 0:
                final_name = f"{row_idx}.png"
            else:
                final_name = f"{row_idx}-{saved_count}.png"

            temp_path = save_dir / f"{row_idx}_raw_{saved_count}.png"
            final_path = save_dir / final_name

            # ===== URL 뽑기 (중복 제거는 하지 않음) =====
            image_url = self._pick_image_url(el) if hasattr(self, "_pick_image_url") else ""

            downloaded = False

            # ✅ 테스트 플래그가 꺼져 있으면 다운로드 수행
            if not self.FORCE_CAPTURE_TEST and image_url:
                try:
                    self._log(f"🌐 [브라우저 fetch] 이미지 다운로드 시도: {image_url}")
                    start = time.time()

                    # 브라우저 세션 그대로 활용해서 fetch
                    img_bytes = self._fetch_image_via_browser(
                        driver, image_url, timeout=15.0
                    )

                    elapsed = time.time() - start
                    self._log(
                        f"⏱ 다운로드 소요시간(fetch+base64): {elapsed:.2f}초 | {image_url}"
                    )

                    if img_bytes:
                        # temp 파일로 한 번 저장 후, 1000x1000 후처리
                        with open(temp_path, "wb") as f:
                            f.write(img_bytes)

                        self._process_and_save_image_1000x1000(
                            temp_path, final_path
                        )

                        try:
                            temp_path.unlink()
                        except Exception:
                            pass

                        self._log(f"📥 브라우저 fetch 다운로드 성공 → {final_path.name}")
                        saved_count += 1
                        downloaded = True
                    else:
                        self._log("⚠️ 브라우저 fetch로 이미지를 가져오지 못했습니다.")
                except Exception as e:
                    self._log(f"⚠️ 브라우저 fetch 이미지 다운로드 중 예외 발생: {e}")

            elif self.FORCE_CAPTURE_TEST:
                self._log(
                    "🧪 [TEST] FORCE_CAPTURE_TEST=True → 다운로드 단계 건너뛰고 캡처 경로로 이동"
                )

            # ✅ 첫 이미지가 다운로드에 실패했을 때만 캡처 백업
            if saved_count == 0 and not downloaded:
                try:
                    self._log(f"🌐 브라우저에서 이미지 직접 저장 시도")
                    self._save_image_from_browser(driver, el, final_path)
                    self._log(f"📥 브라우저 저장 성공 → {final_path.name}")
                    saved_count += 1
                    downloaded = True
                except Exception as e:
                    self._log(f"⚠️ 브라우저 이미지 저장 실패: {e}")

        if saved_count == 0:
            self._log("⚠️ 어떤 이미지도 저장하지 못했습니다.")
        else:
            self._log(f"✅ 총 {saved_count}장의 코스트코 이미지를 저장했습니다.")

    def _process_and_save_image_1000x1000(self, src_path: Path, dst_path: Path):
        """
        - 배경 제거(흰색을 투명으로 만드는 작업)를 하지 않는다.
        - 원본 비율을 유지하면서 긴 변 기준 1000 이하로 축소하고
        - 1000x1000 흰색 배경 캔버스에 중앙 정렬해서 저장한다.
        """
        try:
            img = Image.open(src_path).convert("RGB")

            # 긴 변 기준 1000으로 축소
            img.thumbnail((1000, 1000), Image.Resampling.LANCZOS)

            # 1000x1000 흰색 캔버스 위에 중앙 배치
            canvas = Image.new("RGB", (1000, 1000), (255, 255, 255))
            x = (1000 - img.width) // 2
            y = (1000 - img.height) // 2
            canvas.paste(img, (x, y))

            canvas.save(dst_path, format="PNG")
        except Exception as e:
            self._log(f"❌ 이미지 후처리 실패: {e}")

    def record_data(self):
        if not self.crawled_url:
            self._log("⚠️ 먼저 [대상윈도우]로 제목/가격/URL을 크롤링해 주세요.")
            return

        # === 1) 도메인에 따라 분기 ===
        host = urlparse(self.crawled_url or "").netloc.lower()

        # 코스트코: sellertool_upload.xlsm 에 기록 + 이미지 캡처
        if "costco.co.kr" in host:
            self._log("🧾 코스트코 상품으로 인식 → 엑셀 기록 + 이미지/스펙 캡처")
            row_idx = None
            try:
                row_idx = self._write_costco_to_seller_excel()
            except Exception as e:
                self._log(f"[오류] 코스트코 엑셀 기록 실패: {e}")

            if row_idx:
                # 상품 이미지 여러 장
                try:
                    self._capture_costco_image(row_idx)
                except Exception as e:
                    self._log(f"[오류] 코스트코 이미지 캡처 실패: {e}")

                # 스펙 영역 캡처
                try:
                    self._capture_costco_spec(row_idx)
                except Exception as e:
                    self._log(f"[오류] 코스트코 스펙 캡처 실패: {e}")

            return




        # (선택) 도매매 전용 분기도 가능하지만,
        # 현재는 '도매매 외 사이트도 구글시트 소싱목록에 기록' 구조이므로 그대로 둠.
        # if "domeme.domeggook.com" in host:
        #     ...  # 필요시 별도 처리

        # === 2) 나머지(도매매/네이버/기타)는 기존 로직 그대로 ===
        try:
            # ✅ 1) 먼저 구글 밑줄 실행 (에러가 나도 기록은 계속 진행)
            try:
                self.google_underline()
            except Exception as e:
                self._log(f"⚠️ 구글 밑줄 실행 중 오류(기록은 계속 진행): {e}")

            # ✅ 2) 그 다음 실제 데이터 기록 (소싱상품목록 시트)
            self._write_row_to_first_empty_a()

        except Exception as e:
            self._log(f"[오류] 시트 기록 실패: {e}")

    # ---------- 시트 클릭 대기 → 기록 ----------
    def _wait_for_sheet_click_then_write(self):
        if self._sheet_click_wait:
            return
        self._sheet_click_wait = True
        start_ts = time.time()
        self._log("⌛ 시트 클릭 대기 시작 (10초)")

        def wait_click():
            nonlocal start_ts
            with mouse.Events() as events:
                for event in events:
                    if (time.time() - start_ts) * 1000 >= CLICK_TIMEOUT_MS_RECORD:
                        break
                    if isinstance(event, mouse.Events.Click) and event.pressed:
                        self._sheet_click_wait = False
                        self.record_data()
                        return
            self._sheet_click_wait = False
            self._log("⏰ 10초 내 시트 클릭이 감지되지 않았습니다. [기록] 버튼으로 입력하세요.")

        import threading
        t = threading.Thread(target=wait_click, daemon=True)
        t.start()

    # ---------- 등록상품명 URL + 숫자 분리 ----------
    def _split_registered_name(self, text: str) -> tuple[str, str, str, str]:
        """
        '등록상품명' 문자열을 다음 4개로 분리:
          - orig : 원문 전체
          - num_part : 맨 앞의 숫자 (숫자 + 공백 패턴, 없으면 "")
          - mid_part : 숫자와 URL 사이의 중간 텍스트
          - url_part : https:// 로 시작하는 URL (여러 개면 첫 번째만)
        예)
          '10 샴푸 세트 https://example.com/abc'
            → num_part='10', mid_part='샴푸 세트', url_part='https://example.com/abc'
          '샴푸 세트 https://example.com/abc'
            → num_part='', mid_part='샴푸 세트', url_part='https://example.com/abc'
          '10 샴푸 세트'
            → num_part='10', mid_part='샴푸 세트', url_part=''
        """
        t = (text or "").strip()
        if not t:
            return "", "", "", ""

        # 1) URL 먼저 분리
        m_url = re.search(r'(https?://\S+)', t)
        url_part = ""
        before = t
        if m_url:
            url_part = m_url.group(1).rstrip('),].;\'"')  # 흔한 꼬리표 제거
            before = t[:m_url.start()].strip()
        else:
            before = t

        # 2) 맨 앞 숫자 분리 (숫자 + 공백 + 나머지)
        num_part = ""
        mid_part = ""
        m_num = re.match(r'^\s*(\d+)\s+(.*)$', before)
        if m_num:
            num_part = m_num.group(1)
            mid_part = m_num.group(2).strip()
        else:
            # 맨 앞에 숫자가 없으면, 전체를 중간 텍스트로 사용
            mid_part = before

        return t, num_part, mid_part, url_part

    # ==== 등록상품명(셀러상품 상세) 조회 유틸 ====
    def _cp_get_registered_product_name(self, seller_product_id: str) -> str | None:
        if not seller_product_id:
            return None
        if seller_product_id in self._cp_seller_name_cache:
            return self._cp_seller_name_cache[seller_product_id]
        paths = [
            f"/v2/providers/openapi/apis/api/v1/marketplace/seller-products/{seller_product_id}",
            f"/v2/providers/openapi/apis/api/v2/vendors/{COUPANG_VENDOR_ID}/seller-products/{seller_product_id}",
        ]
        for path in paths:
            try:
                data = _cp_request("GET", path, None)
                info = (data or {}).get("data") or {}
                name = info.get("sellerProductName") or info.get("name")
                if name:
                    self._cp_seller_name_cache[seller_product_id] = name
                    return name
            except Exception:
                continue
        return None

    # ==== 쿠팡 주문조회 + 시트기록 ====
    def _fetch_coupang_orders(self) -> list[dict]:
        if not (COUPANG_VENDOR_ID and COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY):
            self._log("❌ 쿠팡 API 키/벤더ID 설정이 비어 있습니다.")
            return []

        days = int(self.spin_days.value() if hasattr(self, "spin_days") else DEFAULT_LOOKBACK_DAYS)
        to_dt = datetime.now(timezone.utc)
        from_dt = to_dt - timedelta(days=days)
        created_from = from_dt.strftime("%Y-%m-%d")
        created_to = to_dt.strftime("%Y-%m-%d")
        self._log(f"🔍 조회기간: 최근 {days}일 (UTC {created_from} ~ {created_to})")

        path = f"/v2/providers/openapi/apis/api/v4/vendors/{COUPANG_VENDOR_ID}/ordersheets"
        all_rows: list[dict] = []

        for st in CP_QUERY_STATUSES:
            api_status_candidates = ORDER_STATUS_ALIASES.get(st, [st])
            status_succeeded = False

            for api_status in api_status_candidates:
                next_token = None
                while True:
                    params = {
                        "createdAtFrom": created_from,
                        "createdAtTo": created_to,
                        "status": api_status,
                        "maxPerPage": 50,
                    }
                    if next_token:
                        params["nextToken"] = next_token

                    try:
                        data = _cp_request("GET", path, params)
                    except requests.HTTPError as e:
                        resp = getattr(e, "response", None)
                        body = ""
                        try:
                            body = resp.text or ""
                        except Exception:
                            pass
                        if getattr(resp, "status_code", None) == 400 and "Invalid Status" in body:
                            self._log(f"ℹ️ 상태 '{api_status}' 미허용 → 다음 후보로 폴백 시도")
                            break
                        self._log_http_error(e, context=f"쿠팡 API 호출 실패(status={st}, api_status={api_status})")
                        break
                    except Exception as e:
                        self._log(f"⚠️ 쿠팡 API 호출 실패(status={st}, api_status={api_status}): {repr(e)}")
                        break

                    result_code = str(data.get("code", "")).upper()
                    if result_code and result_code not in ("SUCCESS", "OK", "200"):
                        msg = safe_str(data.get("message"))
                        if "Invalid Status" in msg:
                            self._log(f"ℹ️ 상태 '{api_status}' 미허용(code={result_code}) → 다음 후보로 폴백")
                            break
                        self._log(f"⚠️ 응답 코드 이상(status={st}, api_status={api_status}): {msg}")
                        break

                    datas = data.get("data")
                    if isinstance(datas, list):
                        sheets = datas
                    elif isinstance(datas, dict):
                        sheets = (
                            datas.get("orderSheets")
                            or datas.get("shipmentBoxInfos")
                            or datas.get("items")
                            or []
                        )
                    else:
                        sheets = []
                    if isinstance(sheets, dict):
                        sheets = [sheets]
                    if not isinstance(sheets, list):
                        sheets = []

                    for sheet in sheets:
                        if not isinstance(sheet, dict):
                            continue
                        order_id = (sheet.get("orderId") or sheet.get("orderIdMask") or sheet.get("orderNo") or "")
                        order_date = (sheet.get("orderedAt") or sheet.get("orderDate") or sheet.get("orderTime") or "")
                        receiver = sheet.get("receiver") or {}
                        if not isinstance(receiver, dict):
                            receiver = {}
                        recv_name = (receiver.get("name") or receiver.get("receiverName") or "")
                        recv_addr = (receiver.get("addr1") or receiver.get("address1") or receiver.get("address") or "")
                        recv_phone = (receiver.get("contact1") or receiver.get("contact2") or receiver.get("phone") or "")
                        items = (sheet.get("orderItems") or sheet.get("orderSheetItems") or sheet.get("items") or [])
                        if isinstance(items, dict):
                            items = [items]
                        if not isinstance(items, list):
                            items = []

                        for it in items:
                            if not isinstance(it, dict):
                                continue
                            item_name = (
                                it.get("sellerProductName")
                                or it.get("vendorItemName")
                                or it.get("productName")
                                or ""
                            )
                            order_item_id = it.get("orderItemId") or it.get("vendorItemId") or ""
                            
                            # 수량: ordersheets 문서 기준 shippingCount 가 정식 필드
                            qty = it.get("shippingCount") or it.get("quantity") or 1

                            # 결제금액: orderPrice / salesPrice × 수량 등 복합 로직
                            paid_price = extract_paid_price_from_item(it)

                            
                            tracking_no = it.get("invoiceNumber") or it.get("trackingNumber") or ""
                            carrier = it.get("deliveryCompanyName") or it.get("deliveryCompanyCode") or ""
                            status_text = CP_STATUS_MAP.get(st, st)

                            seller_product_id = (
                                it.get("sellerProductId")
                                or sheet.get("sellerProductId")
                                or ""
                            )
                            registered_name = (
                                it.get("sellerProductName")
                                or (self._cp_get_registered_product_name(str(seller_product_id)) if seller_product_id else None)
                                or ""
                            )

                            orig_reg, reg_num, reg_mid, reg_url = self._split_registered_name(registered_name)


                            all_rows.append({
                                "주문일시": order_date,
                                "상태": status_text,
                                "주문번호": order_id,
                                "주문아이템ID": order_item_id,

                                # 등록상품명 관련 4분할
                                "등록상품명":   orig_reg,   # 원문 전체
                                "등록상품명-1": reg_num,    # 맨 앞 숫자 (없으면 "")
                                "등록상품명-2": reg_mid,    # 숫자와 URL 사이 텍스트
                                "등록상품명-3": reg_url,    # URL

                                "수량": qty,
                                "결제금액": paid_price,

                                # ★ 새 컬럼: 최종 수익 (초기에는 빈 값, 나중에 '주문정리' 버튼에서 채움)
                                "최종 수익": "",

                                "수취인": recv_name,
                                "연락처": recv_phone,
                                "주소": recv_addr,
                                "송장번호": tracking_no,
                                "택배사": carrier,

                                "셀러상품ID": str(seller_product_id or ""),
                            })



                    next_token = None
                    # 1) data["data"] 는 리스트이므로, 여기서 nextToken 을 찾지 말고
                    # 2) 응답 최상위에서 nextToken 을 읽어야 함
                    if isinstance(data, dict):
                        nt = data.get("nextToken")
                        if nt:
                            next_token = nt
                    if not next_token:
                        status_succeeded = True
                        break  # while
                # 다음 상태 별칭으로 폴백
            if not status_succeeded:
                self._log(f"ℹ️ 상태 '{st}'는 제공 계정/엔드포인트 조합에서 미허용이거나 데이터가 없습니다.")

        # 정렬: 상태(비즈니스 순서) → 주문일시(최신우선)
        def _parse_dt_safe(s: str):
            s = (s or "").strip()
            if not s:
                return None
            try:
                if s.endswith("Z"):
                    return datetime.fromisoformat(s.replace("Z", "+00:00"))
                return datetime.fromisoformat(s)
            except Exception:
                m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
                if m:
                    try:
                        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    except Exception:
                        return None
                return None

        def _sort_key(row: dict):
            st_txt = str(row.get("상태", ""))
            st_rank = STATUS_ORDER.get(st_txt, 999)
            dt = _parse_dt_safe(row.get("주문일시"))
            ts = -(dt.timestamp()) if dt else float("inf")
            return (st_rank, ts)

        all_rows.sort(key=_sort_key)

        self._log(f"📦 쿠팡 주문 수집 완료: {len(all_rows)}건")
        return all_rows

    def _write_coupang_orders_to_sheet(self, rows: list[dict]):
        if self.sheets.ws is None:
            self._log("⚠️ Sheets 연결이 필요합니다. 먼저 [Sheets 연결] 버튼을 눌러주세요.")
            return
        try:
            ws = self.sheets.gc.open_by_key(SHEET_ID).worksheet(COUPANG_WS_NAME)
        except gspread.WorksheetNotFound:
            ws = self.sheets.gc.open_by_key(SHEET_ID).add_worksheet(title=COUPANG_WS_NAME, rows=4000, cols=40)

        if not rows:
            # 등록상품명 3분할 + 최종 수익 헤더
            headers = [
                "주문일시","상태","주문번호","주문아이템ID",
                "등록상품명","등록상품명-1","등록상품명-2","등록상품명-3",
                "수량","결제금액","최종 수익","수취인","연락처","주소","송장번호","택배사","셀러상품ID"
            ]
            ws.clear()
            # 컬럼 수 17개 → A~Q 이지만, 헤더는 17개라 실제로는 A~Q 중 1칸은 비게 됩니다.
            # 크게 문제는 없으니 그대로 두셔도 되고, 엄밀하게 맞추려면 A1:Q1 → A1:Q1 그대로 둬도 무방합니다.
            ws.update(values=[headers], range_name="A1:Q1")
            self._log("ℹ️ 쿠팡 주문 데이터가 없어 헤더만 갱신했습니다.")
            return



        headers = list(rows[0].keys())
        values = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]

        ws.clear()
        end_col_letter = _a1_col(len(headers))
        rng = f"A1:{end_col_letter}{len(values)}"
        ws.update(values=values, range_name=rng, value_input_option="USER_ENTERED")
        self._log(f"✅ '{COUPANG_WS_NAME}' 탭에 {len(rows)}건 업데이트 완료")

    # === 금일 올린 상품 갯수 계산 ===
    def update_today_product_count(self):
        """
        B열의 날짜가 어떤 형식이든 (M/D, YYYY-MM-DD 등)
        '오늘 날짜'에 해당하는 구간의 A열 값으로 금일 올린 상품 개수를 계산해서 라벨에 표시.
        """

        # ⬇️ 내부에서 쓸 날짜 정규화 함수
        def _normalize_date_for_compare(s: str):
            """
            셀 문자열 s 를 date 객체로 변환.
            - '2025-11-29'
            - '2025/11/29'
            - '11/29/2025'
            - '11/29'
            등 여러 패턴을 허용.
            """
            s = (s or "").strip()
            if not s:
                return None

            # 1) 자주 쓰는 전체 날짜 포맷들 시도
            fmts = [
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%m/%d/%Y",
                "%m/%d/%y",
            ]
            for fmt in fmts:
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    pass

            # 2) 연도 없는 'M/D' 형식 (예: 11/29)
            m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
            if m:
                year = datetime.today().year
                try:
                    return date(year, int(m.group(1)), int(m.group(2)))
                except ValueError:
                    return None

            # 그 외는 인식 불가
            return None

        # 1) Sheets 연결 확인
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(금일 올린 상품 갯수 계산): {e}")
                self.lbl_today_count.setText("금일 상품 갯수 : 오류")
                return

        try:
            ws = self.sheets.ws

            # 2) B열 전체 값 가져오기
            col_b = ws.col_values(2)

            # 오늘 날짜 (연/월/일)
            today_date = datetime.today().date()

            # 3) 오늘 날짜에 해당하는 행 번호들 찾기 (1-based)
            today_rows = []
            for idx, v in enumerate(col_b):
                val = str(v).strip()
                if not val:
                    continue
                d = _normalize_date_for_compare(val)
                if d and d == today_date:
                    today_rows.append(idx + 1)

            # 디버깅용 로그: B열에 어떤 값들이 있었는지 보고 싶으면 주석 해제
            # self._log(f"[DEBUG] B열 샘플: {col_b[:10]}")

            if not today_rows:
                count = 0
                self._log(
                    f"📊 오늘 날짜({today_date.isoformat()})와 일치하는 B열 값이 없어 0개로 계산합니다."
                )
            else:
                first_row = today_rows[0]
                last_row = today_rows[-1]

                # 4) A열 값 읽어서 번호 기준으로 갯수 계산
                col_a = ws.col_values(1)
                a_first = str(col_a[first_row - 1]).strip() if len(col_a) >= first_row else ""
                a_last  = str(col_a[last_row  - 1]).strip() if len(col_a) >= last_row  else ""

                try:
                    n_first = int(a_first)
                    n_last  = int(a_last)
                    # 하단 A - 상단 A + 1
                    count = n_last - n_first + 1
                    if count < 0:
                        # A열이 꼬여 있으면, 그냥 행 개수로 처리
                        count = len(today_rows)
                except Exception:
                    # A열이 숫자가 아니면, 단순히 오늘 날짜 행 개수로 처리
                    count = len(today_rows)

                self._log(
                    f"📊 금일 상품 갯수 계산: B열 날짜={today_date.isoformat()} 구간 "
                    f"A({a_first})~A({a_last}) → {count}개"
                )

            # 5) 라벨 업데이트
            self.lbl_today_count.setText(f"금일 상품 갯수 : {count}")

        except Exception as e:
            self._log(f"❌ 금일 상품 갯수 계산 중 오류: {e}")
            self.lbl_today_count.setText("금일 상품 갯수 : 오류")




    # === 쿠팡 주문현황 버튼 동작 ===
    def coupang_orders(self):
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패: {e}")
                return
        try:
            rows = self._fetch_coupang_orders()
        except Exception as e:
            self._log(f"❌ 쿠팡 주문 조회 중 오류: {e}")
            return
        try:
            self._write_coupang_orders_to_sheet(rows)
        except Exception as e:
            self._log(f"❌ 쿠팡 주문 기록 중 오류: {e}")

    # === (통합) 쿠팡 키 확인 + 헬스체크 버튼 동작 ===
    def coupang_key_and_health(self):
        self.check_coupang_keys()
        self.coupang_healthcheck()

    # === 쿠팡 키 확인 ===
    def check_coupang_keys(self):
        try:
            p = Path(COUPANG_KEYS_JSON)
            if not p.exists():
                self._log(f"❌ 키 파일을 찾지 못했습니다: {COUPANG_KEYS_JSON}")
                self._log("➡ 경로/파일명을 다시 확인하거나 JSON을 생성해 주세요.")
                return
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            vendor_id = (data.get("vendor_id") or "").strip()
            access_key = (data.get("access_key") or "").strip()
            secret_key = (data.get("secret_key") or "").strip()

            self._log("✅ JSON 파일 읽기 성공")
            self._log(f"• Vendor ID: {vendor_id or '(빈 값)'}")
            self._log(f"• Access Key: {access_key or '(빈 값)'}")
            self._log(f"• Secret Key: {_mask(secret_key) if secret_key else '(빈 값)'}")

            problems = []
            if not vendor_id: problems.append("vendor_id가 비어 있습니다.")
            if not access_key: problems.append("access_key가 비어 있습니다.")
            if not secret_key: problems.append("secret_key가 비어 있습니다.")
            if problems:
                for m in problems:
                    self._log(f"⚠️ {m}")
                return

            mismatches = []
            if COUPANG_VENDOR_ID != vendor_id:
                mismatches.append("전역 Vendor ID와 JSON의 vendor_id가 다릅니다.")
            if COUPANG_ACCESS_KEY != access_key:
                mismatches.append("전역 Access Key와 JSON의 access_key가 다릅니다.")
            if COUPANG_SECRET_KEY != secret_key:
                mismatches.append("전역 Secret Key와 JSON의 secret_key가 다릅니다.")
            if mismatches:
                self._log("⚠️ 전역 설정과 JSON 파일의 값이 일치하지 않습니다:")
                for m in mismatches:
                    self._log(f"   - {m}")
                self._log("➡ JSON을 수정했으면 프로그램을 재시작하거나, 상단 상수 경로/로딩 부분을 확인하세요.")

            # 간단 HMAC 메시지 생성 테스트
            try:
                test_path = f"/v2/providers/openapi/apis/api/v4/vendors/{vendor_id}/ordersheets"
                test_query = urlencode({"status": "ACCEPT", "maxPerPage": 50}, doseq=True)
                signed_date = datetime.now(timezone.utc).strftime("%y%m%dT%H%M%SZ")
                msg = f"{signed_date}{'GET'}{test_path}{test_query}"
                signature = hmac.new(secret_key.encode("utf-8"), msg.encode("utf-8"), hashlib.sha256).hexdigest()
                auth_head = (
                    f"CEA algorithm=HmacSHA256, access-key={access_key}, "
                    f"signed-date={signed_date}, signature={signature}"
                )
                self._log("🔐 HMAC 서명 생성 테스트 성공")
                self._log(f"• Authorization 헤더 앞부분: {auth_head[:60]}...")
            except Exception as e:
                self._log(f"❌ HMAC 서명 생성 실패: {e}")

            self._log("🟢 키 확인 완료")

        except json.JSONDecodeError as e:
            self._log(f"❌ JSON 파싱 실패: {e}")
            self._log("➡ 파일 내용이 유효한 JSON 형식인지 확인하세요.")
        except Exception as e:
            self._log(f"❌ 키 확인 중 오류: {e}")

    # === 쿠팡 API 헬스체크 ===
    def coupang_healthcheck(self):
        self._log("🩺 쿠팡 API 헬스체크 시작")
        if not (COUPANG_VENDOR_ID and COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY):
            self._log("❌ 쿠팡 키/벤더ID가 비어 있습니다. coupang_keys.json 확인")
            return
        try:
            to_dt = datetime.now(timezone.utc)
            from_dt = to_dt - timedelta(days=1)  # 헬스체크는 간단히 최근 1일로 확인
            path = f"/v2/providers/openapi/apis/api/v4/vendors/{COUPANG_VENDOR_ID}/ordersheets"
            param_variants = _build_ordersheets_params(from_dt, to_dt, status="ACCEPT", max_per_page=1)
            data = _try_ordersheets_with_variants(path, param_variants)
            code = str(data.get("code", "")).upper()
            self._log(f"✅ 헬스체크 성공: path='{path}', params={param_variants[0]} (code={code or 'N/A'})")
            self._log("🟢 쿠팡 API 키/서명/경로 정상으로 보입니다.")
            return
        except requests.HTTPError as e:
            self._log_http_error(e, context="헬스체크(ordersheets) 실패")
        except Exception as e:
            self._log(f"❌ 헬스체크(ordersheets) 중 예외: {repr(e)}")
        self._log("❌ 헬스체크가 실패했습니다. 다음을 점검해 주세요:\n"
                  "  1) 판매자센터(Wing) OpenAPI 키 여부 (파트너스 키 아님)\n"
                  "  2) 시스템연동 > Open API 사용 활성 및 권한 승인\n"
                  "  3) 허용 IP에 현재 PC 공인 IP 등록\n"
                  "  4) PC 시간 자동 동기화(UTC, 수초 이하 오차)\n")
    
    # === 쿠팡주문현황 '최종 수익' 채우기 (주문정리) ===
    def settle_orders(self):
        """쿠팡주문현황 시트에서 등록상품명-1/결제금액으로 소싱상품목록의 O열 값을 찾아와 K열(최종 수익)에 채워 넣는다."""
        # 1) Sheets 연결 확인
        if self.sheets.gc is None or self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(주문정리): {e}")
                return

        try:
            sh = self.sheets.gc.open_by_key(SHEET_ID)
            ws_orders = sh.worksheet(COUPANG_WS_NAME)
        except Exception as e:
            self._log(f"❌ '쿠팡주문현황' 시트를 찾지 못했습니다: {e}")
            return

        try:
            # 소싱상품목록 시트
            try:
                ws_source = sh.worksheet(WORKSHEET_NAME)
            except Exception:
                # 이미 self.sheets.ws 가 소싱상품목록이면 그걸 사용
                ws_source = self.sheets.ws

            # 2) 쿠팡주문현황 전체 값 가져오기
            orders_values = ws_orders.get_all_values()
            if not orders_values or len(orders_values) < 2:
                self._log("ℹ️ '쿠팡주문현황' 시트에 데이터가 없습니다.")
                return

            header = orders_values[0]

            # 필요한 컬럼 인덱스 찾기
            try:
                idx_reg1 = header.index("등록상품명-1")
                idx_paid = header.index("결제금액")
            except ValueError:
                self._log("❌ '등록상품명-1' 또는 '결제금액' 헤더를 찾지 못했습니다. 헤더명을 확인해 주세요.")
                return

            # '최종 수익' 컬럼 인덱스 확보 (없으면 결제금액 바로 오른쪽에 새로 추가)
            if "최종 수익" in header:
                idx_profit = header.index("최종 수익")
            else:
                idx_profit = idx_paid + 1
                # 모든 행에 대해 '최종 수익' 컬럼을 삽입 (초기값 "")
                for r in range(len(orders_values)):
                    row = orders_values[r]
                    # 결제금액 위치까지는 최소 길이 확보
                    while len(row) <= idx_paid:
                        row.append("")
                    row.insert(idx_profit, "")

                # 헤더 다시 갱신
                header = orders_values[0]
                self._log(f"🆕 '최종 수익' 컬럼을 추가했습니다. (열 인덱스: {idx_profit+1})")

            # 3) 소싱상품목록에서 (A,Q)->O 매핑 생성
            source_values = ws_source.get_all_values()
            if not source_values or len(source_values) < 2:
                self._log("ℹ️ '소싱상품목록' 시트에 데이터가 없어 주문정리를 건너뜁니다.")
                return

            profit_map = {}  # key: (A값, Q값) → O값
            # 0-based: A=0, O=14, Q=16
            for srow in source_values[1:]:
                if len(srow) < 17:
                    continue
                a_val = digits_only(srow[0])
                q_val = digits_only(srow[16])
                if not a_val or not q_val:
                    continue
                key = (a_val, q_val)
                # 같은 키가 여러 번 나올 수 있지만, 첫 번째 값만 사용
                if key not in profit_map:
                    o_val = srow[14] if len(srow) > 14 else ""
                    profit_map[key] = o_val

            self._log(f"📚 소싱상품목록 매핑 생성 완료: {len(profit_map)}개 키")

            # 4) 쿠팡주문현황 각 행에 대해 '최종 수익' 채우기
            updated_count = 0
            max_idx = max(idx_reg1, idx_paid, idx_profit)

            for i in range(1, len(orders_values)):  # 2행부터
                row = orders_values[i]
                # 최소 길이 확보
                if len(row) <= max_idx:
                    row.extend([""] * (max_idx + 1 - len(row)))

                reg_code = digits_only(row[idx_reg1])
                paid_val = digits_only(row[idx_paid])

                if not reg_code or not paid_val:
                    continue

                key = (reg_code, paid_val)
                profit_val = profit_map.get(key, "")

                if profit_val:
                    row[idx_profit] = profit_val
                    updated_count += 1

            # 5) 시트에 다시 반영
            end_col_letter = _a1_col(max(len(r) for r in orders_values))
            end_row = len(orders_values)
            rng = f"A1:{end_col_letter}{end_row}"
            ws_orders.update(rng, orders_values, value_input_option="USER_ENTERED")

            self._log(f"✅ 주문정리 완료: {updated_count}건에 '최종 수익'(열 K)을 반영했습니다.")

        except Exception as e:
            self._log(f"❌ 주문정리 처리 중 오류: {e}")

    
     
    # === 구글시트 A열 첫 빈 행 상단 테두리 (구글 밑줄) ===
    def google_underline(self):
        """A열의 비어있는 첫 번째 셀을 포함한 행 전체에 '윗부분'만 테두리를 긋기."""
        # Sheets 연결 확인
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(구글 밑줄): {e}")
                return

        try:
            ws = self.sheets.ws

            # 1) A열에서 비어있는 첫 번째 행 찾기
            target_row = self.sheets.find_first_empty_row_in_col_a_from_top()
            self._log(f"🔎 A열 기준 첫 빈 행: {target_row}행")

            # 2) sheetId 가져오기
            try:
                sheet_id = ws.id  # gspread 최신 버전 속성
            except AttributeError:
                sheet_id = ws._properties.get("sheetId")

            if sheet_id is None:
                self._log("❌ sheetId를 가져오지 못했습니다. (ws.id / ws._properties['sheetId'] 확인 필요)")
                return

            # 3) 몇 번째 컬럼까지 테두리를 칠지 결정
            #    - 헤더(row 1)의 실제 사용 컬럼 수를 기준으로, 최소 10컬럼 이상은 잡도록 처리
            try:
                header_values = ws.row_values(1)
                used_cols = max(len(header_values), 10)
            except Exception:
                used_cols = 10

            # Google Sheets API index는 0-based 이므로 변환
            start_row_index = target_row - 1
            end_row_index = target_row
            start_col_index = 0          # A열
            end_col_index = used_cols    # A ~ (used_cols)열

            body = {
                "requests": [
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": start_row_index,
                                "endRowIndex": end_row_index,
                                "startColumnIndex": start_col_index,
                                "endColumnIndex": end_col_index,
                            },
                            # 윗부분 테두리만 적용
                            "top": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                }
                            }
                            # left/right/bottom은 지정하지 않아 기존 스타일 유지
                        }
                    }
                ]
            }

            # 4) batch_update 실행
            ws.spreadsheet.batch_update(body)
            self._log(
                f"✅ 구글 밑줄 적용 완료: 행 {target_row} (A{target_row} ~ { _a1_col(end_col_index) }{target_row}) 상단 테두리"
            )

        except Exception as e:
            self._log(f"❌ 구글 밑줄 처리 중 오류: {e}")
            
    def _capture_costco_spec(self, row_idx: int):
        """
        코스트코 상품 페이지의 '스펙' 패널을 열고
        패널 전체(mat-expansion-panel#product_specs)만 그대로 캡처한다.
        - 파일명: {row_idx}_spec.png
        - 경로: sellertool_upload.xlsm 이 있는 폴더 아래 /YYYYMMDD/
        """
        try:
            driver = self._attach_driver()
        except Exception as e:
            self._log(f"❌ 코스트코 스펙 캡처: 드라이버 연결 실패: {e}")
            return

        try:
            # 1) '스펙' 패널 열고, 패널 요소와 body 요소를 얻음
            spec_panel, spec_body = self._open_costco_spec_section()

            # 2) 저장 폴더 준비 (날짜별)
            base_dir = Path(SELLERTOOL_XLSM_PATH).parent
            date_folder = datetime.now().strftime("%Y%m%d")
            save_dir = base_dir / date_folder
            save_dir.mkdir(parents=True, exist_ok=True)

            save_path = save_dir / f"{row_idx}_spec.png"

            # 3) 캡처 대상: 패널 전체(mat-expansion-panel#product_specs)
            target_el = spec_panel

            # 화면 가운데로 스크롤 + 레이아웃 안정화 기다리기
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", target_el
                )
            except Exception:
                pass

            try:
                self._wait_until_element_stable(driver, target_el)
            except Exception:
                pass

            # 4) 요소 스크린샷
            target_el.screenshot(str(save_path))
            self._log(f"✅ 스펙 캡처 완료(패널 전체): {save_path}")

        except TimeoutException:
            self._log("❌ '스펙' 패널 또는 내용 영역을 찾지 못했습니다. 코스트코 페이지 구조를 다시 한 번 확인해 주세요.")
        except Exception as e:
            self._log(f"❌ 코스트코 스펙 캡처 중 오류: {e}")
 
    def _open_costco_spec_section(self):
        """
        코스트코 상품 페이지에서
        - 헤더 텍스트에 '스펙' 이 들어가는 아코디언 패널을 찾고
        - 접혀 있으면 클릭해서 열고
        - 그 패널 요소(= mat-expansion-panel#product_specs 전체)와
          패널 안의 내용 영역(body)을 함께 리턴한다.
        """
        driver = self._attach_driver()
        wait = WebDriverWait(driver, 10)

        # 1) '스펙' 이라는 텍스트를 가진 아코디언 헤더 찾기
        header_xpath = (
            "//*[contains(@class,'mat-expansion-panel-header') and "
            "     .//*[contains(normalize-space(),'스펙')]]"
        )

        spec_header = wait.until(
            EC.element_to_be_clickable((By.XPATH, header_xpath))
        )

        # 2) 이 헤더가 속한 패널(<mat-expansion-panel> 또는 div.mat-expansion-panel)을 찾기
        parent_panel = spec_header.find_element(
            By.XPATH, "ancestor::*[contains(@class,'mat-expansion-panel')][1]"
        )

        # 만약 id='product_specs' 가 붙어 있으면, 나중에 디버깅이 더 쉽도록 로그
        panel_id = parent_panel.get_attribute("id") or ""
        if panel_id:
            self._log(f"🧩 스펙 패널 id={panel_id}")

        # 3) 이미 펼쳐져 있는지 확인 (aria-expanded 또는 클래스에 mat-expanded 여부)
        expanded_attr = spec_header.get_attribute("aria-expanded") or ""
        if not expanded_attr:
            expanded_attr = parent_panel.get_attribute("class") or ""

        if ("true" not in expanded_attr.lower()
                and "mat-expanded" not in expanded_attr):
            spec_header.click()
            # 펼쳐질 때까지 잠깐 대기
            try:
                wait.until(
                    lambda d: (
                        "mat-expanded" in (parent_panel.get_attribute("class") or "")
                        or spec_header.get_attribute("aria-expanded") == "true"
                    )
                )
            except Exception:
                pass  # 너무 빡빡하게 볼 필요는 없어서 실패해도 그냥 진행

        # 4) 이 패널 안의 내용 영역(파란 영역 내부 body) 찾기
        body_xpath = (
            ".//*[contains(@class,'mat-expansion-panel-content') "
            "   or contains(@class,'mat-expansion-panel-body')]"
        )
        try:
            spec_body = parent_panel.find_element(By.XPATH, body_xpath)
        except Exception:
            spec_body = parent_panel  # 못 찾으면 패널 전체로 fallback

        # 패널을 화면 중앙으로
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", parent_panel
            )
        except Exception:
            pass
        time.sleep(0.3)

        self._log("🟢 '스펙' 패널 열기 및 panel/body 요소 찾기 완료")
        # 👉 패널 전체, 내부 body 둘 다 반환
        return parent_panel, spec_body


    def _wait_until_element_stable(
        self,
        driver,
        element,
        check_interval: float = 0.2,
        stable_checks: int = 3,
        max_wait: float = 10.0,
    ) -> bool:
        """
        요소의 getBoundingClientRect() 값이 일정 시간 동안 변하지 않을 때까지 기다린다.
        - 슬라이드/애니메이션이 끝나고 '화면에서 고정된 상태'가 되었을 때 True 반환.
        - max_wait 동안 안정되지 않으면 False 반환.
        """
        import time

        last_rect = None
        stable_count = 0
        end_time = time.time() + max_wait

        while time.time() < end_time:
            try:
                rect = driver.execute_script(
                    """
                    const r = arguments[0].getBoundingClientRect();
                    return [r.x, r.y, r.width, r.height];
                    """,
                    element,
                )
            except Exception:
                # 요소가 더 이상 없으면 안정화 의미가 없으니 종료
                break

            if rect == last_rect:
                stable_count += 1
            else:
                stable_count = 0
            last_rect = rect

            if stable_count >= stable_checks:
                return True

            time.sleep(check_interval)

        return False

    def _save_image_from_browser(self, driver, img_element, save_path):
        # 브라우저에서 이미 로드된 이미지를 base64로 추출
        img_data = driver.execute_script("""
            var img = arguments[0];
            var canvas = document.createElement('canvas');
            canvas.width = img.naturalWidth;
            canvas.height = img.naturalHeight;
            var ctx = canvas.getContext('2d');
            ctx.drawImage(img, 0, 0);
            return canvas.toDataURL('image/png').split(',')[1];
        """, img_element)
        img_bytes = base64.b64decode(img_data)
        img = Image.open(io.BytesIO(img_bytes))
        img.save(save_path, format="PNG")
        
    def _fetch_image_via_browser(self, driver, url: str, timeout: float = 15.0) -> bytes | None:
        """
        브라우저(JS fetch)를 이용해 image URL을 가져온 뒤,
        base64 문자열로 Python에 전달해서 bytes 로 반환한다.
        - Chrome 세션 쿠키/헤더/연결을 그대로 활용할 수 있음.
        """
        if not url:
            return None

        script = """
        const url = arguments[0];
        const callback = arguments[arguments.length - 1];

        fetch(url, { credentials: 'include' })
          .then(resp => {
            if (!resp.ok) {
              throw new Error('HTTP ' + resp.status);
            }
            return resp.arrayBuffer();
          })
          .then(buf => {
            const bytes = new Uint8Array(buf);
            let binary = '';
            const len = bytes.byteLength;
            for (let i = 0; i < len; i++) {
              binary += String.fromCharCode(bytes[i]);
            }
            // base64 로 인코딩해서 콜백으로 넘김
            callback(btoa(binary));
          })
          .catch(err => {
            callback(null);
          });
        """

        try:
            # Selenium의 async script 사용 (마지막 인수가 callback)
            driver.set_script_timeout(timeout)
            b64_data = driver.execute_async_script(script, url)
        except Exception as e:
            self._log(f"⚠️ 브라우저 fetch 실행 중 오류: {e}")
            return None

        if not b64_data:
            return None

        try:
            return base64.b64decode(b64_data)
        except Exception as e:
            self._log(f"⚠️ 브라우저 fetch base64 디코딩 실패: {e}")
            return None
