# ================== 크롬 크롤링 + 구글시트 + 쿠팡 OpenAPI (UI로 조회기간/헬스체크/주문현황) ==================
import sys
import os
import re
import time
import json
import platform
import socket
import subprocess
from pathlib import Path
from urllib.parse import urlparse, urlencode, quote  # canonical query 생성을 위해 quote 사용
from datetime import datetime, timedelta, timezone

# ==== PyQt6 ====
from PyQt6.QtGui import QKeySequence, QShortcut
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QTextEdit, QHBoxLayout, QSpinBox
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal

# ==== UI/OS/입력 ====
import pygetwindow as gw
import pyautogui
import pyperclip
from pynput import mouse
from pynput.mouse import Listener as MouseListener

# ==== Selenium ====
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ==== Google Sheets ====
import gspread
from google.oauth2.service_account import Credentials
from google.auth.exceptions import TransportError

# ==== HTTP/HMAC ====
import requests
import hmac, hashlib

# ==== Excel ====
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ==== costco 크롤링 관련 ====
from PIL import Image
import io
import base64


# =========================
# 설정값 (튜닝 포인트)
# =========================
# --- Google Sheets ---
SERVICE_ACCOUNT_JSON = "/Users/jeehoonkim/Desktop/api/google_api/service_account.json"  # 서비스계정 키 경로
SHEET_ID = "1OEg01RdJyesSy7iQSEyQHdYpCX5MSsNUfD0lkUYq8CM"  # 스프레드시트 ID
WORKSHEET_NAME = "소싱상품목록"  # 시트 탭 이름

# --- 크롬 디버그 포트/경로 ---
DEBUGGER_ADDR = "127.0.0.1:9222"
DEBUGGER_PORT = 9222
CHROME_PATHS = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    "/Applications/Google Chrome Beta.app/Contents/MacOS/Google Chrome Beta",
    "/Applications/Google Chrome Canary.app/Contents/MacOS/Google Chrome Canary",
]
USER_DATA_DIR = str(Path("/Users/Shared/chrome_dev"))

# --- 지연/타임아웃 ---
CLICK_TIMEOUT_MS_SELECT = 5000   # 대상윈도우 선택(본문 클릭) 대기 타임아웃
CLICK_TIMEOUT_MS_RECORD = 10000  # 시트 클릭 대기 타임아웃
KEY_DELAY_SEC = 0.01
CLICK_STABILIZE_SEC = 0.01
NAV_DELAY_SEC = 0.005

DATE_FORMAT = "M/D"        # 날짜 포맷
FIXED_CONST_FEE = "3000"   # I열 고정 수수료

# --- URL→라벨 매핑(C열) ---
DOMAIN_LABELS = {
    "domeme.domeggook.com": "도매매",
    "naver.com": "네이버",
    "costco.co.kr": "코코",
    "ownerclan.com": "오너",
}

# --- 크롤링용 기본/사이트별 셀렉터 ---
SITE_SELECTORS = {
    "domeme.domeggook.com": ["#lInfoItemTitle", "h1#lInfoItemTitle", "h1"],
    "costco.co.kr": [".product-detail__name", "h1.product-detail__name", "h1"]
}

SITE_PRICE_SELECTORS = {
    "domeme.domeggook.com": ["#lItemPrice", ".lItemPrice", "#lItemPriceText"]
}
DEFAULT_SELECTORS = [
    '#lInfoItemTitle', 'h1.l.infoItemTitle',
    'h1#l\\.infoItemTitle', 'h1',
    '[role="heading"][aria-level="1"]'
]

# ✅ 여기: 코스트코 패턴 추가
URL_PATTERNS = [
    "domeme.domeggook.com/s/",
    "domeme.domeggook.com",
    "costco.co.kr"           # << 이 줄 추가
]

def today_iso() -> str:
    """YYYY-MM-DD 형식 오늘 날짜"""
    return datetime.now().strftime("%Y-%m-%d")


# --- Coupang Open API (Wing) ---
COUPANG_BASE_URL = "https://api-gateway.coupang.com"
COUPANG_KEYS_JSON = str(Path("/Users/jeehoonkim/Desktop/api/coupang_api/coupang_keys.json"))
try:
    with open(COUPANG_KEYS_JSON, "r", encoding="utf-8") as f:
        coupang_keys = json.load(f)
        COUPANG_VENDOR_ID = (coupang_keys.get("vendor_id") or "").strip()
        COUPANG_ACCESS_KEY = (coupang_keys.get("access_key") or "").strip()
        COUPANG_SECRET_KEY = (coupang_keys.get("secret_key") or "").strip()
except Exception as e:
    print(f"❌ 쿠팡 키 파일을 불러오지 못했습니다: {e}")
    COUPANG_VENDOR_ID = COUPANG_ACCESS_KEY = COUPANG_SECRET_KEY = None

COUPANG_WS_NAME = "쿠팡주문현황"

# 👇 하드코딩된 조회일수 제거 → UI의 SpinBox로 제어 (초기값만 준다)
DEFAULT_LOOKBACK_DAYS = 7

# ---- 조회/표시할 상태: 결제완료 → 상품준비중 → 배송지시 → 배송중 → 배송완료
CP_QUERY_STATUSES = ["ACCEPT", "INSTRUCT", "DEPARTURE", "DELIVERING", "DELIVERED"]


# ---- 시트에 적을 한글 상태 라벨
CP_STATUS_MAP = {
    "ACCEPT":     "결제완료",     # 쿠팡 헬프/연동 문서에서 ACCEPT를 결제완료로 표현
    "INSTRUCT":   "상품준비중",
    "DEPARTURE":  "배송지시",     # ★ 핵심: 배송지시 = DEPARTURE
    "DELIVERING": "배송중",
    "DELIVERED":  "배송완료",
}

# ---- API별 상태 이름이 다른 경우를 흡수 (우선순위 순)
ORDER_STATUS_ALIASES = {
    # 결제완료
    "ACCEPT":     ["ACCEPT", "PAID", "PAYMENT_COMPLETED", "ORDER_COMPLETE"],
    # 상품준비중
    "INSTRUCT":   ["INSTRUCT", "READY", "READY_FOR_DELIVERY", "PREPARE_SHIPMENT"],
    # 배송지시 (핵심)
    "DEPARTURE":  ["DEPARTURE", "DELIVERY_REQUESTED", "SHIPPING_READY"],
    # 배송중
    "DELIVERING": ["DELIVERING"],
    # 배송완료 (계정/버전별 상이)
    "DELIVERED":  ["DELIVERED", "DELIVERY_COMPLETED", "DONE", "FINAL_DELIVERY"],
}

STATUS_ORDER = {
    "결제완료": 0,
    "상품준비중": 1,
    "배송지시": 2,   # ★ 추가
    "배송중":   3,
    "배송완료": 4,
}

# --- 코스트코 → 쿠팡 대량등록 엑셀 (sellertool_upload.xlsm) ---
SELLERTOOL_XLSM_PATH = "/Users/jeehoonkim/Desktop/Python_Project/crawling_temp/sellertool_upload.xlsm"  # <-- 경로 직접 수정
SELLERTOOL_SHEET_NAME = "data"  # 실제 시트 이름으로 바꿔 주세요 (예: '상품등록', 'Sheet1' 등)



# =========================
# 유틸 함수
# =========================
def is_macos() -> bool:
    return platform.system().lower() == "darwin"

def safe_str(v) -> str:
    try:
        if callable(v): v = v()
    except Exception:
        pass
    try:
        return "" if v is None else str(v)
    except Exception:
        return ""

def digits_only(s: str) -> str:
    return re.sub(r"[^0-9]", "", safe_str(s))

def is_int_string(s: str) -> bool:
    return re.fullmatch(r"\s*[+-]?\d+\s*", safe_str(s)) is not None

def today_fmt() -> str:
    now = datetime.now()
    return f"{now.month}/{now.day}" if DATE_FORMAT == "M/D" else f"{now.month:02d}/{now.day:02d}"

def is_port_open(host: str, port: int, timeout=0.3) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except OSError:
        return False

def selectors_for_url(url: str):
    host = urlparse(url).netloc if url else ""
    site_specific = []
    for key, sels in SITE_SELECTORS.items():
        if key in host:
            site_specific += sels
    seen, ordered = set(), []
    for sel in site_specific + DEFAULT_SELECTORS:
        if sel not in seen:
            seen.add(sel); ordered.append(sel)
    return ordered

def price_selectors_for_url(url: str):
    host = urlparse(url).netloc if url else ""
    site_specific = []
    for key, sels in SITE_PRICE_SELECTORS.items():
        if key in host:
            site_specific += sels
    general = ["#lItemPrice", ".lItemPrice", ".price .num", ".price-value", ".final_price",
               ".sale_price", ".price", "[data-testid='price']"]
    seen, ordered = set(), []
    for sel in site_specific + general:
        if sel not in seen:
            seen.add(sel); ordered.append(sel)
    return ordered

def label_for_domain(url: str) -> str:
    host = urlparse(url or "").netloc.lower()
    for dom, lab in DOMAIN_LABELS.items():
        if dom in host:
            return lab
    return ""

def is_costco_url(url: str) -> bool:
    host = urlparse(url or "").netloc.lower()
    return "costco.co.kr" in host

def is_domeme_url(url: str) -> bool:
    host = urlparse(url or "").netloc.lower()
    return "domeme.domeggook.com" in host

def _mask(s: str, left: int = 4, right: int = 3) -> str:
    """키 마스킹: 앞/뒤 일부만 보이고 나머지는 * 처리"""
    s = str(s or "")
    if len(s) <= left + right:
        return "*" * len(s)
    return s[:left] + "*" * (len(s) - left - right) + s[-right:]

def _a1_col(index: int) -> str:
    """1-based column index -> A1 column letters (1->A, 26->Z, 27->AA ...)"""
    if index <= 0:
        raise ValueError("index must be >= 1")
    s = ""
    while index > 0:
        index, r = divmod(index - 1, 26)
        s = chr(65 + r) + s
    return s

# =========================
# 쿠팡 OpenAPI: “성공 예제” 규격으로 HMAC 구현
# =========================
#  - 메시지: signed-date + METHOD + PATH + QUERY   (구분자/개행/물음표 없음)
#  - 서명  : HMAC-SHA256(hex)
#  - 날짜  : YYMMDDTHHMMSSZ  (예: 251111T110106Z)
#  - 쿼리  : urllib.parse.urlencode 기본값(공백→+), URL과 서명에서 “동일 문자열” 사용
def _cp_build_query(params: dict | None) -> str:
    if not params:
        return ""
    return urlencode(params, doseq=True)  # quote_plus 방식 (공백→+)

def _cp_signed_headers_v2(method: str, path: str, sign_query: str,
                          access_key: str, secret_key: str,
                          *, signed_date: str | None = None, vendor_id: str | None = None) -> dict:
    if signed_date is None:
        signed_date = datetime.now(timezone.utc).strftime("%y%m%dT%H%M%SZ")  # YYMMDDTHHMMSSZ
    message = f"{signed_date}{method.upper()}{path}{sign_query}"
    signature = hmac.new(
        secret_key.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    authorization = (
        f"CEA algorithm=HmacSHA256, access-key={access_key}, "
        f"signed-date={signed_date}, signature={signature}"
    )
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Authorization": authorization,
    }
    if vendor_id:
        headers["X-Requested-By"] = vendor_id
    return headers

def _cp_request(method: str, path: str, params: dict | None) -> dict:
    if not (COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY):
        raise RuntimeError("쿠팡 API 키가 설정되지 않았습니다.")
    url_query = _cp_build_query(params)
    url = f"{COUPANG_BASE_URL}{path}" + (f"?{url_query}" if url_query else "")
    try:
        headers = _cp_signed_headers_v2(
            method, path, url_query, COUPANG_ACCESS_KEY, COUPANG_SECRET_KEY,
            vendor_id=COUPANG_VENDOR_ID
        )
        resp = requests.request(method=method, url=url, headers=headers, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.HTTPError as e:
        body = ""
        try:
            body = resp.text[:1000]
        except Exception:
            pass
        msg = f"{resp.status_code} {resp.reason}\nurl={url}\nresp_body={body}"
        raise requests.HTTPError(msg, response=resp, request=resp.request) from e

# === ordersheets 파라미터 빌더 (yyyy-MM-dd) + 폴백 ===
def _build_ordersheets_params(date_from_utc: datetime, date_to_utc: datetime, status: str, max_per_page: int = 50):
    d_from = date_from_utc.strftime("%Y-%m-%d")
    d_to   = date_to_utc.strftime("%Y-%m-%d")
    primary = {
        "createdAtFrom": d_from,
        "createdAtTo": d_to,
        "status": status,
        "maxPerPage": max_per_page,
    }
    fallback = {
        "startTime": d_from,
        "endTime": d_to,
        "status": status,
        "maxPerPage": max_per_page,
    }
    return [primary, fallback]

def _try_ordersheets_with_variants(path: str, param_variants: list[dict]) -> dict:
    last_err = None
    for params in param_variants:
        try:
            return _cp_request("GET", path, params)
        except requests.HTTPError as e:
            resp = getattr(e, "response", None)
            status = getattr(resp, "status_code", None)
            body = ""
            try:
                body = (resp.text or "")[:500]
            except Exception:
                pass
            if status == 400 and "yyyy-MM-dd" in body:
                last_err = e
                continue
            raise
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise RuntimeError("ordersheets 호출 시도 실패: 유효한 파라미터 조합이 없습니다.")

# =========================
# 결제 금액 파서
# =========================
def extract_money_amount(m: dict | None) -> int:
    """
    Coupang 금액 오브젝트({currencyCode, units, nanos})에서
    '원' 기준 정수 금액을 추출.
    """
    if not isinstance(m, dict):
        return 0
    units = m.get("units", 0)
    nanos = m.get("nanos", 0)

    try:
        units = int(units)
    except Exception:
        units = 0

    try:
        nanos = int(nanos)
    except Exception:
        nanos = 0

    # KRW는 보통 소수점이 없으므로 nanos는 이론상 0일 것.
    # 혹시 모를 값이 있어도 반올림해서 원 단위로 맞춤.
    if nanos:
        return units + round(nanos / 1_000_000_000)

    return units


# =========================
# 결제 금액용 계산 함수
# =========================
def extract_paid_price_from_item(it: dict) -> int:
    """
    쿠팡 ordersheets 응답의 orderItems 항목(it)에서
    '결제금액'을 최대한 안정적으로 계산한다.
    우선순위:
      1) orderPrice money-object (currencyCode/units/nanos)
      2) orderPrice가 숫자/문자열이면 그대로
      3) salesPrice money-object × 수량(shippingCount/quantity)
      4) 그 외 후보 필드(paidPrice, paymentAmount 등)를 숫자로 파싱
    """
    if not isinstance(it, dict):
        return 0

    # --- 1차: orderPrice money-object (공식 스펙)
    op = it.get("orderPrice")
    if isinstance(op, dict):
        v = extract_money_amount(op)
        if v:
            return v

    # --- 2차: orderPrice가 그냥 숫자/문자열인 경우
    if op is not None and not isinstance(op, dict):
        s = digits_only(op)
        if s:
            try:
                return int(s)
            except Exception:
                pass

    # --- 3차: salesPrice × 수량(shippingCount/quantity)
    sales = it.get("salesPrice")
    sales_val = 0
    if isinstance(sales, dict):
        sales_val = extract_money_amount(sales)
    elif sales is not None:
        s = digits_only(sales)
        if s:
            try:
                sales_val = int(s)
            except Exception:
                sales_val = 0

    qty = it.get("shippingCount") or it.get("quantity") or 1
    try:
        qty = int(qty)
    except Exception:
        qty = 1

    if sales_val and qty:
        return sales_val * qty

    # --- 4차: 기타 필드들에서 숫자만 뽑아보기 (혹시 계정마다 다르게 내려오는 경우 대비)
    for key in ("paidPrice", "paymentAmount", "price"):
        if key in it and it[key] is not None:
            s = digits_only(it[key])
            if s:
                try:
                    return int(s)
                except Exception:
                    pass

    # 모두 실패하면 0
    return 0


# =========================
# Google Sheets 래퍼
# =========================
class SheetsClient:
    def __init__(self, json_path: str, sheet_id: str, worksheet_name: str, logger):
        self.json_path = json_path
        self.sheet_id = sheet_id
        self.worksheet_name = worksheet_name
        self.logger = logger
        self.gc = None
        self.ws = None
        self.CREATE_WORKSHEET_IF_MISSING = False

    def connect(self):
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_file(self.json_path, scopes=scopes)
        self.gc = gspread.authorize(creds)
        sh = self.gc.open_by_key(self.sheet_id)
        try:
            self.ws = sh.worksheet(self.worksheet_name)
            self.logger(f"✅ Google Sheets 연결 완료 (워크시트: {self.worksheet_name})")
        except gspread.WorksheetNotFound:
            titles = [w.title for w in sh.worksheets()]
            self.logger(f"⚠️ 워크시트 '{self.worksheet_name}'를 찾지 못함. 현재 탭들: {titles}")
            if self.CREATE_WORKSHEET_IF_MISSING:
                self.ws = sh.add_worksheet(title=self.worksheet_name, rows=1000, cols=30)
                self.logger(f"🆕 워크시트 생성: {self.worksheet_name}")
            else:
                raise

    def get_next_index(self) -> int:
        try:
            col_values = self.ws.col_values(1)
            last = None
            for v in reversed(col_values):
                if v.strip():
                    last = v
                    break
            if last is None:
                return 1
            return int(last) + 1 if is_int_string(last) else 1
        except Exception as e:
            self.logger(f"⚠️ A열 인덱스 계산 실패, 1로 시작: {e}")
            return 1

    def find_first_empty_row_in_col_a_from_top(self) -> int:
        values = self.ws.col_values(1)
        if not values:
            return 1
        for i, v in enumerate(values, start=1):
            if not str(v).strip():
                return i
        return len(values) + 1

    def append_row_with_retry(self, row_values, max_tries=5, base_sleep=0.6):
        attempt = 0
        while True:
            try:
                self.ws.append_row(row_values, value_input_option="USER_ENTERED")
                return True
            except gspread.exceptions.APIError as e:
                attempt += 1
                try:
                    resp = getattr(e, "response", None)
                    status = getattr(resp, "status_code", None)
                    text = getattr(resp, "text", "")
                    self.logger(f"❌ APIError(status={status}): {text[:500]}")
                except Exception:
                    self.logger(f"❌ APIError: {e}")
                if attempt >= max_tries:
                    return False
                sleep_s = base_sleep * (2 ** (attempt - 1))
                self.logger(f"⏳ 재시도 {attempt}/{max_tries} ... {sleep_s:.1f}s")
                time.sleep(sleep_s)
            except (TransportError, Exception) as e:
                attempt += 1
                self.logger(f"❌ 전송/기타 오류: {repr(e)}")
                if attempt >= max_tries:
                    return False
                sleep_s = base_sleep * (2 ** (attempt - 1))
                self.logger(f"⏳ 재시도 {attempt}/{max_tries} ... {sleep_s:.1f}s")
                time.sleep(sleep_s)

# =========================
# 메인 앱
# =========================
class ChromeCrawler(QWidget):
    clickDetected = pyqtSignal(int, int)

    # ✅ 테스트용 플래그: True로 두면 무조건 "다운로드 건너뛰고 캡처" 경로로 테스트
    FORCE_CAPTURE_TEST = False # 테스트 미진행. 다운로드 우선순위 진행
    #FORCE_CAPTURE_TEST = True  # 테스트용 플래그: True로 두면 무조건 "다운로드 건너뛰고 캡처" 경로로 테스트
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("크롬 크롤링 도구 (gspread + Coupang OpenAPI)")
        self.setGeometry(0, 0, 460, 580)

        # 등록상품명 캐시 (sellerProductId -> 등록상품명)
        self._cp_seller_name_cache: dict[str, str] = {}

        # 상태값
        self.target_title = None
        self.target_window = None
        self.driver = None
        self._listener = None
        self._waiting_click = False
        self._sheet_click_wait = False
        self._click_timer = None

        # 크롤 결과
        self.crawled_title = ""
        self.crawled_price = ""
        self.crawled_url = ""

        # Google Sheets
        self.sheets = SheetsClient(SERVICE_ACCOUNT_JSON, SHEET_ID, WORKSHEET_NAME, self._log)
        self.row_index_cache = None

        # =========================
        # UI
        # =========================
        layout = QVBoxLayout()
        layout.setSpacing(6)
        layout.setContentsMargins(8, 8, 8, 8)

        self.label = QLabel("🖱 대상 윈도우: 없음")
        layout.addWidget(self.label)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        # 1) clear + Sheets 연결
        row_a = QHBoxLayout()
        self.btn_clear = QPushButton("Txt clear")
        self.btn_clear.clicked.connect(self.log.clear)
        row_a.addWidget(self.btn_clear)

        self.btn_sheets = new_btn = QPushButton("Sheets 연결")
        new_btn.clicked.connect(self.connect_sheets)
        row_a.addWidget(new_btn)
        layout.addLayout(row_a)

        # 2) 크롬(디버그) + 기존 창 연결 테스트
        row_b = QHBoxLayout()
        self.btn_launch = QPushButton("크롬(디버그) 실행")
        self.btn_launch.clicked.connect(self.launch_debug_chrome)
        row_b.addWidget(self.btn_launch)

        self.btn_test = QPushButton("기존 창 연결 테스트")
        self.btn_test.clicked.connect(self.test_attach_existing)
        row_b.addWidget(self.btn_test)
        layout.addLayout(row_b)

        # 3) 대상윈도우 + 기록
        row_c = QHBoxLayout()
        self.btn_select = QPushButton("대상윈도우 (Shift+Z)")
        self.btn_select.clicked.connect(self.select_target_window)
        row_c.addWidget(self.btn_select)

        self.btn_record = QPushButton("기록 (Shift+X)")
        self.btn_record.clicked.connect(self.record_data)
        row_c.addWidget(self.btn_record)
        layout.addLayout(row_c)

        # 4) STOP + 네이버(최저가)
        row_d = QHBoxLayout()
        self.btn_stop = QPushButton("STOP (프로그램 off)")
        self.btn_stop.clicked.connect(self.close)
        row_d.addWidget(self.btn_stop)

        self.btn_health = QPushButton("네이버 (최저가))")
        self.btn_health.clicked.connect(self.naver_check)
        row_d.addWidget(self.btn_health)
        layout.addLayout(row_d)

        # 5) 금일 상품 갯수 + (통합) 쿠팡 키 확인 + 헬스체크
        row_e = QHBoxLayout()

        # ← 금일 상품 갯수 라벨
        self.lbl_today_count = QLabel("금일 상품 갯수 : 0")
        row_e.addWidget(self.lbl_today_count)

        # 가운데: 상품개수계산 버튼
        self.btn_calc_today = QPushButton("상품개수계산")
        self.btn_calc_today.clicked.connect(self.update_today_product_count)
        row_e.addWidget(self.btn_calc_today)

        # 오른쪽: (통합) 쿠팡 키+헬스체크
        self.btn_cp_keyhealth = QPushButton("쿠팡 키+헬스체크")
        self.btn_cp_keyhealth.clicked.connect(self.coupang_key_and_health)
        row_e.addWidget(self.btn_cp_keyhealth)

        layout.addLayout(row_e)


        # 6) 하단: 확인기간 + 스핀박스 + (우측) 쿠팡 주문현황
        row_z = QHBoxLayout()
        self.lbl_days = QLabel("확인기간 :")
        row_z.addWidget(self.lbl_days)

        self.spin_days = QSpinBox()
        self.spin_days.setRange(1, 365)       # 1~365일 범위 허용
        self.spin_days.setValue(DEFAULT_LOOKBACK_DAYS)  # 초기값 7일
        self.spin_days.setSuffix(" 일")
        self.spin_days.setSingleStep(1)
        row_z.addWidget(self.spin_days)

        row_z.addStretch(1)  # 왼쪽 요소들 뒤로 공간 확보 (우측 버튼 정렬)

        self.btn_coupang = QPushButton("쿠팡 주문현황")
        self.btn_coupang.clicked.connect(self.coupang_orders)
        row_z.addWidget(self.btn_coupang)

        # 👉 새로 추가: 주문정리 버튼
        self.btn_order_settle = QPushButton("주문정리")
        self.btn_order_settle.clicked.connect(self.settle_orders)
        row_z.addWidget(self.btn_order_settle)

        # 👉 기존: 구글시트 밑줄 버튼
        self.btn_google_underline = QPushButton("구글시트 밑줄")
        self.btn_google_underline.clicked.connect(self.google_underline)
        row_z.addWidget(self.btn_google_underline)

        layout.addLayout(row_z)


        # 버튼 높이/패딩
        for btn in (
            self.btn_clear, self.btn_sheets, self.btn_launch, self.btn_test,
            self.btn_select, self.btn_record, self.btn_stop, self.btn_health,
            self.btn_cp_keyhealth, self.btn_coupang, self.btn_order_settle, self.btn_google_underline
        ):

            btn.setMinimumHeight(28)
            btn.setStyleSheet("QPushButton { padding: 4px 8px; }")


        # 안내
        self._log(
            "ℹ️ 사용법:\n"
            "1) [Sheets 연결] → [크롬(디버그) 실행] 후 대상 페이지를 엽니다.\n"
            "2) [대상윈도우] 클릭 → 안내에 따라 '본문'을 클릭(5초 내).\n"
            "3) 하단 [확인기간] 일수를 설정 후 [쿠팡 주문현황]으로 조회합니다.\n"
            "4) [쿠팡 키+헬스체크] 버튼으로 키/서명/경로 정상 여부를 점검합니다.\n"
        )

        self.setLayout(layout)

        # 단축키
        QShortcut(QKeySequence("Shift+Z"), self, activated=self.select_target_window)
        QShortcut(QKeySequence("Shift+X"), self, activated=self.record_data)

        # 전역 클릭 시그널
        self.clickDetected.connect(self._handle_click_on_main)

        # 자동 초기화
        QTimer.singleShot(300, self._startup_sequence)

    # ---------- 로깅 ----------
    def _log(self, msg: str):
        self.log.append(msg)
        print(msg)

    # ---------- 공통 HTTP 에러 로깅 ----------
    def _log_http_error(self, e: Exception, context: str = ""):
        if isinstance(e, requests.HTTPError):
            resp = getattr(e, "response", None)
            req = getattr(e, "request", None)
            status = getattr(resp, "status_code", None)
            reason = getattr(resp, "reason", "")
            url = getattr(req, "url", "(unknown)")
            try:
                body = resp.text if resp is not None else str(e)
            except Exception:
                body = str(e)
            if context:
                self._log(f"❌ {context}: {status or 'N/A'} {reason or e.__class__.__name__}")
            else:
                self._log(f"❌ 요청 실패: {status or 'N/A'} {reason or e.__class__.__name__}")
            self._log(f"url={url}")
            self._log(f"resp_body={(body or '')[:1000]}")
        else:
            if context:
                self._log(f"❌ {context} 중 예외: {repr(e)}")
            else:
                self._log(f"❌ 예외: {repr(e)}")

    # ---------- 자동 시작 시퀀스 ----------
    def _startup_sequence(self):
        self._log("🚀 시작: 자동 초기화 시퀀스 실행")
        try:
            self.connect_sheets()
        except Exception as e:
            self._log(f"⚠️ 자동 시트 연결 실패: {e}")

        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결 → '기존 창 연결 테스트' 수행")
            ok = self._attach_existing_ok()
            if ok:
                self.test_attach_existing()
            else:
                self._log("ℹ️ 기존 창 연결 실패 → '크롬(디버그) 실행' 수행")
                self.launch_debug_chrome()
        else:
            self._log("✅ Sheets 연결 완료(자동)")

    def _attach_existing_ok(self) -> bool:
        try:
            if not is_port_open("127.0.0.1", DEBUGGER_PORT):
                self._log("ℹ️ 디버그 포트가 열려 있지 않음")
                return False
            driver = self._attach_driver()
            _ = driver.window_handles
            self._log("✅ 기존 창 연결 OK")
            return True
        except Exception as e:
            self._log(f"ℹ️ 기존 창 연결 실패: {e}")
            return False

    # 네이버 쇼핑 열기
    def _open_naver_shopping_with_title(self, sort_low_price: bool = True):
        try:
            title = (self.crawled_title or "").strip()
            if not title:
                self._log("ℹ️ 제목이 없어 네이버 쇼핑 검색을 생략합니다.")
                return
            driver = self._attach_driver()
            from urllib.parse import quote_plus
            base_url = "https://search.shopping.naver.com/search/all"
            q = f"query={quote_plus(title)}"
            sort = "sort=price_asc" if sort_low_price else "sort=rel"
            search_url = f"{base_url}?{q}&{sort}"
            driver.execute_script("window.open(arguments[0], '_blank');", search_url)
            driver.switch_to.window(driver.window_handles[-1])
            self._log(f"🟢 네이버 쇼핑 검색 탭 오픈(낮은가격순 시도): {search_url}")
            if not sort_low_price:
                return
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
            except Exception:
                pass
            if "sort=price_asc" in (driver.current_url or ""):
                return
            click_js = r"""
            const keywords = ['낮은가격순', '가격낮은순'];
            function clickByText(nodes) {
              for (const el of nodes) {
                try {
                  const t = (el.innerText || el.textContent || '').trim();
                  if (!t) continue;
                  for (const k of keywords) {
                    if (t.includes(k)) { el.click(); return true; }
                  }
                } catch (e) {}
              }
              return false;
            }
            const order = ['button','a','span','div','li'];
            for (const tag of order) {
              const list = document.querySelectorAll(tag);
              if (clickByText(list)) return true;
            }
            return false;
            """
            clicked = driver.execute_script(click_js)
            if clicked:
                self._log("✅ 정렬 UI 클릭으로 '낮은 가격순' 적용 시도")
                try:
                    WebDriverWait(driver, 5).until(lambda d: "price_asc" in (d.current_url or ""))
                except Exception:
                    pass
            else:
                self._log("⚠️ 정렬 UI 요소를 찾지 못했습니다. (페이지 UI 변경 가능)")
        except Exception as e:
            self._log(f"⚠️ 네이버 쇼핑 검색/정렬 처리 실패: {e}")

    # ---------- Sheets ----------
    def connect_sheets(self):
        try:
            self.sheets.connect()
        except Exception as e:
            self._log(f"❌ Sheets 연결 실패: {e}")
            raise

    def naver_check(self):
        self._open_naver_shopping_with_title(sort_low_price=True)

    # ---------- Chrome ----------
    def launch_debug_chrome(self):
        try:
            if is_port_open("127.0.0.1", DEBUGGER_PORT):
                self._log(f"ℹ️ 디버그 포트 {DEBUGGER_PORT} 이미 열림. 기존 창에 연결하세요.")
                return
            chrome_bin = None
            for p in CHROME_PATHS:
                if os.path.exists(p):
                    chrome_bin = p; break
            if chrome_bin is None:
                self._log("⚠️ Chrome 실행 파일을 찾지 못했습니다.")
                return
            Path(USER_DATA_DIR).mkdir(parents=True, exist_ok=True)
            cmd = [
                chrome_bin,
                f"--remote-debugging-port={DEBUGGER_PORT}",
                f"--user-data-dir={USER_DATA_DIR}",
                "--no-first-run", "--no-default-browser-check"
            ]
            subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, start_new_session=True)
            for _ in range(25):
                if is_port_open("127.0.0.1", DEBUGGER_PORT):
                    self._log(f"✅ 디버깅 모드 Chrome 실행됨 (포트 {DEBUGGER_PORT}).")
                    return
                time.sleep(0.2)
            self._log("⚠️ 디버그 포트 연결 확인 실패")
        except Exception as e:
            self._log(f"[오류] 크롬(디버그) 실행 실패: {e}")

    def _attach_driver(self):
        if not is_port_open("127.0.0.1", DEBUGGER_PORT):
            raise RuntimeError("디버그 포트가 열려 있지 않습니다. 먼저 '크롬(디버그) 실행'을 눌러주세요.")
        if self.driver:
            return self.driver
        options = webdriver.ChromeOptions()
        options.debugger_address = f"127.0.0.1:{DEBUGGER_PORT}"
        self.driver = webdriver.Chrome(options=options)
        return self.driver

    def test_attach_existing(self):
        try:
            driver = self._attach_driver()
            tabs_info = []
            for h in driver.window_handles:
                driver.switch_to.window(h)
                tabs_info.append(f"- {safe_str(driver.title).strip()} | {safe_str(driver.current_url).strip()}")
            msg = "🔗 디버그 세션 탭 목록:\n" + ("\n".join(tabs_info) if tabs_info else "(없음)")
            self._log(msg)
        except Exception as e:
            self._log(f"[오류] 기존 창 연결 테스트 실패: {e}")

    # 시트 창 활성화
    def _bring_sheet_to_front(self):
        try:
            sheet_url_prefix = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}"
            if is_macos():
                osa = f'''
                tell application "Google Chrome"
                    set thePrefix to "{sheet_url_prefix}"
                    set foundWin to missing value
                    set foundIdx to -1
                    repeat with w in windows
                        set i to 0
                        repeat with t in tabs of w
                            set i to i + 1
                            if (URL of t) starts with thePrefix then
                                set foundWin to w
                                set active tab index of w to i
                                set index of w to 1
                                activate
                                return
                            end if
                        end repeat
                    end repeat
                    open location thePrefix & "/edit"
                    activate
                end tell
                '''
                subprocess.run(["osascript", "-e", osa], check=False)
            else:
                titles = []
                try:
                    titles = gw.getAllTitles()
                except Exception:
                    pass
                cand = [t for t in titles if isinstance(t, str) and ("Google Sheets" in t or "스프레드시트" in t)]
                if cand:
                    wlist = gw.getWindowsWithTitle(cand[0])
                    if wlist:
                        try:
                            wlist[0].activate()
                        except Exception:
                            pass
                try:
                    import webbrowser
                    webbrowser.open(sheet_url_prefix + "/edit", new=0, autoraise=True)
                except Exception:
                    pass
        except Exception as e:
            self._log(f"⚠️ 시트 창 활성화 실패: {e}")

    # ---------- 대상 선택 & 크롤 ----------
    def select_target_window(self):
        # 대상윈도우 버튼을 누를 때 금일 상품 갯수 자동 계산
        self.update_today_product_count()
        
        self._log("🖱 **크롤링할 크롬 탭의 본문**을 클릭해 주세요. (5초 내)")
        self.label.setText("🔍 본문을 클릭하세요 (주소창 X). 5초 내 미클릭 시 경고.")

        self.showMinimized()
        self._waiting_click = True
        self._sheet_click_wait = False

        if self._click_timer is None:
            self._click_timer = QTimer(self)
            self._click_timer.setSingleShot(True)
            self._click_timer.timeout.connect(self._on_click_timeout_select)
        self._click_timer.start(CLICK_TIMEOUT_MS_SELECT)

        def on_click(x, y, button, pressed):
            if pressed and self._waiting_click:
                self.clickDetected.emit(int(x), int(y))
        self._listener = MouseListener(on_click=on_click)
        self._listener.start()

    def _on_click_timeout_select(self):
        if not self._waiting_click:
            return
        self._waiting_click = False
        try:
            if self._listener: self._listener.stop()
        except Exception:
            pass
        finally:
            self._listener = None
        self._log("⏰ 5초 내 클릭이 감지되지 않았습니다. 다시 [대상윈도우]를 눌러 본문을 클릭하세요.")

    def _handle_click_on_main(self, x: int, y: int):
        if not self._waiting_click:
            return
        self._waiting_click = False
        if self._click_timer and self._click_timer.isActive():
            self._click_timer.stop()
        try:
            if self._listener: self._listener.stop()
        except Exception:
            pass
        finally:
            self._listener = None

        wins_at = self._gw_get_windows_at(x, y)
        win = wins_at[0] if wins_at else None
        picked_title = safe_str(getattr(win, "title", "")) if win else ""
        if not picked_title:
            self._log("❌ 클릭 지점에서 활성 창 제목을 찾지 못했습니다. 본문 클릭/권한 확인.")
            return

        self.target_window = win
        self.target_title = picked_title
        self.label.setText(f"🎯 대상 윈도우: {self.target_title}")

        self.showNormal(); self.raise_(); self.activateWindow()
        self.crawl_data()

    def _gw_get_windows_at(self, x: int, y: int):
        try:
            fn = getattr(gw, "getWindowsAt", None)
            if callable(fn):
                return fn(x, y)
        except Exception:
            pass
        try:
            active = getattr(gw, "getActiveWindow", lambda: None)()
            return [active] if active else []
        except Exception:
            return []

    def crawl_data(self):
        if not self.target_title:
            self._log("⚠️ 대상 탭이 선택되지 않았습니다.")
            return
        try:
            if self.target_window:
                try:
                    self.target_window.activate(); time.sleep(0.2)
                except Exception:
                    pass

            driver = self._attach_driver()

            self._log("🧭 탭 매칭: URL패턴 → 제목 포함")
            end_time = time.time() + 5.0
            target_handle = None

            if URL_PATTERNS:
                while time.time() < end_time and not target_handle:
                    for h in driver.window_handles:
                        driver.switch_to.window(h)
                        if any(p in (driver.current_url or "") for p in URL_PATTERNS):
                            target_handle = h; break
                    if not target_handle:
                        time.sleep(0.2)

            if not target_handle:
                end_time2 = time.time() + 5.0
                raw_want = safe_str(self.target_title).strip()

                # 윈도우 제목에서 ' - ' 뒤에 붙는 브라우저 이름 제거 (예: " - Google Chrome")
                want_base = raw_want.split(" - ")[0].strip() if raw_want else ""

                while time.time() < end_time2 and not target_handle:
                    for h in driver.window_handles:
                        driver.switch_to.window(h)
                        page_title = safe_str(driver.title).strip()
                        page_base = page_title.split(" - ")[0].strip() if page_title else ""

                        # 1) 전체 제목 포함 여부
                        if raw_want and raw_want in page_title:
                            target_handle = h
                            break

                        # 2) '앞부분만' 비교 (상품명 부분만 비교)
                        if want_base and want_base in page_base:
                            target_handle = h
                            break

                    if not target_handle:
                        time.sleep(0.2)


            if not target_handle:
                self._log("❌ 5초 내 '대상 탭'을 찾지 못했습니다.")
                return

            driver.switch_to.window(target_handle)

            current_url = safe_str(driver.current_url).strip()
            self.crawled_url = current_url
            self._log(f"🔗 URL: {current_url}")

            blocked = ("chrome://", "chrome-extension://", "edge://", "about:", "data:")
            if any(current_url.startswith(s) for s in blocked) or current_url.lower().endswith(".pdf"):
                self._log("❌ 이 페이지는 DOM 접근이 제한됩니다.")
                return

            try:
                WebDriverWait(driver, 3).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
            except Exception:
                pass

            title_value = ""
            wait = WebDriverWait(driver, 5)
            for sel in selectors_for_url(current_url):
                try:
                    el = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, sel)))
                    title_value = (el.text or "").strip()
                    if title_value:
                        break
                except Exception:
                    continue
            if not title_value:
                try:
                    el = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.TAG_NAME, "h1")))
                    title_value = (el.text or "").strip()
                except Exception:
                    title_value = ""
            self.crawled_title = title_value
            self._log(f"🟢 제목: {self.crawled_title or '(없음)'}")

            price_digits = ""
            wait_p = WebDriverWait(driver, 3)
            for sel in price_selectors_for_url(current_url):
                try:
                    el = wait_p.until(EC.visibility_of_element_located((By.CSS_SELECTOR, sel)))
                    txt = (el.text or "").strip()
                    if not txt:
                        txt = (driver.execute_script(
                            "const e=document.querySelector(arguments[0]); return e?(e.innerText||e.textContent||''):'';", sel
                        ) or "").strip()
                    if txt:
                        price_digits = re.sub(r"[^0-9]", "", txt)
                        if price_digits:
                            break
                except Exception:
                    continue
            if not price_digits:
                try:
                    body = (driver.execute_script(
                        "return (document.body && document.body.innerText) ? document.body.innerText : '';"
                    ) or "")
                    m = re.search(r'([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)\s*원', body)
                    if not m:
                        m = re.search(r'₩\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)', body)
                    if m:
                        price_digits = re.sub(r"[^0-9]", "", m.group(1))
                except Exception:
                    pass
            self.crawled_price = price_digits
            self._log(f"💰 가격(숫자만): {self.crawled_price or '(없음)'}")

            self._log("—" * 40)
            self._log(f"제목: {self.crawled_title or '(없음)'}")
            self._log(f"가격(숫자만): {self.crawled_price or '(없음)'}")
            self._log(f"URL: {self.crawled_url or '(없음)'}")
            self._log("—" * 40)

            self._log("📝 크롤 완료: 시트에 바로 기록합니다.")
            self.record_data()

        except Exception as e:
            self._log(f"[오류] 크롤링 실패: {e}")

    # ---------- 시트 기록(핵심) ----------
    def _write_row_to_first_empty_a(self):
        if self.sheets.ws is None:
            self._log("⚠️ 먼저 [Sheets 연결]을 눌러 구글시트에 연결해 주세요.")
            return

        target_row = self.sheets.find_first_empty_row_in_col_a_from_top()

        COLS = {c:i for i,c in enumerate(
            ["A","B","C","D","E","F","G","H","I","J",
             "K","L","M","N","O","P","Q","R","S","T",
             "U","V","W","X","Y"], start=1)}

        a_index = str(self.sheets.get_next_index())

        row_buffer = [""] * 25
        row_buffer[COLS["A"]-1] = a_index
        row_buffer[COLS["B"]-1] = today_fmt()
        row_buffer[COLS["C"]-1] = label_for_domain(self.crawled_url)
        row_buffer[COLS["F"]-1] = self.crawled_title or ""
        row_buffer[COLS["H"]-1] = self.crawled_price or ""
        row_buffer[COLS["I"]-1] = FIXED_CONST_FEE
        row_buffer[COLS["J"]-1] = f"=H{target_row}+I{target_row}"
        row_buffer[COLS["K"]-1] = "10.8%"
        row_buffer[COLS["M"]-1] = f"=J{target_row}+(R{target_row}*(K{target_row}*1.1))"
        row_buffer[COLS["N"]-1] = f"=O{target_row}/R{target_row}"
        row_buffer[COLS["O"]-1] = f"=R{target_row}-M{target_row}+K{target_row}-P{target_row}+L{target_row}"
        row_buffer[COLS["R"]-1] = f"=Q{target_row}"
        row_buffer[COLS["S"]-1] = f"=R{target_row}/1.1"
        row_buffer[COLS["T"]-1] = f"=S{target_row}*1.1-S{target_row}"
        row_buffer[COLS["V"]-1] = self.crawled_url or ""

        rng = f"A{target_row}:Y{target_row}"
        self.sheets.ws.update(values=[row_buffer], range_name=rng, value_input_option="USER_ENTERED")
        self._log(f"✅ 행 {target_row} (A..Y)에 기록 완료")

        try:
            if self.crawled_url:
                pyperclip.copy(self.crawled_url)
                self._log("📋 현재 상품 URL을 클립보드에 복사했습니다.")
        except Exception as e:
            self._log(f"⚠️ 클립보드 복사 실패: {e}")

        self._bring_sheet_to_front()

    # ---------- 코스트코 → sellertool_upload.xlsm 기록 ----------
    def _write_costco_to_seller_excel(self):
        """
        코스트코 상품(현재 self.crawled_title / self.crawled_price / self.crawled_url)을
        sellertool_upload.xlsm 에 다음 규칙으로 기록한다.

        - A열 : 카테고리 명 (지금은 공란)
        - B열 : 등록상품명 (크롤링한 상품명)
        - C열 : 오늘 날짜 'YYYY-MM-DD'
        - D열 : 공란
        - E열 : '새상품'
        - F열 : 공란
        - G열 : 등록상품명의 첫 단어
        - H열 : 등록상품명의 첫 단어
        - I열 ~ Z열 : 공란

        - BJ : 5만원 이하 -> 코스트코 가격 * 1.3
               5만원 초과 ~ 10만원 이하 -> 코스트코 가격 * 1.2
               10만원 초과 -> 원가 그대로
        - BL : BJ * 1.05
        - BM : 999
        - BN : 2
        - BX : '상세정보별도표기'
        - CK : '기타재화'
        - CZ : 행번호.png (예: 5행이면 '5.png')
        """

        if not self.crawled_title:
            self._log("⚠️ 코스트코 엑셀 기록: 상품명이 없습니다.")
            return None

        if not os.path.exists(SELLERTOOL_XLSM_PATH):
            self._log(f"❌ 코스트코 엑셀 기록: 파일을 찾지 못했습니다 → {SELLERTOOL_XLSM_PATH}")
            return None

        try:
            self._log(f"📂 엑셀 열기: {SELLERTOOL_XLSM_PATH}")
            wb = load_workbook(SELLERTOOL_XLSM_PATH, keep_vba=True)
        except Exception as e:
            self._log(f"❌ 엑셀 로드 실패: {e}")
            return None

        # 시트 선택
        try:
            if SELLERTOOL_SHEET_NAME in wb.sheetnames:
                ws = wb[SELLERTOOL_SHEET_NAME]
            else:
                ws = wb[wb.sheetnames[0]]
                self._log(f"⚠️ 시트 '{SELLERTOOL_SHEET_NAME}'를 찾지 못해 첫 번째 시트('{ws.title}')를 사용합니다.")
        except Exception as e:
            self._log(f"❌ 시트 선택 실패: {e}")
            return None

        # ==== 1) 입력할 행 찾기 (3행부터) ====
        start_row = 3
        row_idx = start_row
        while True:
            cell_val = ws.cell(row=row_idx, column=2).value  # B열
            if cell_val is None or str(cell_val).strip() == "":
                break
            row_idx += 1

        # ==== 2) 공통 데이터 준비 ====
        full_name = self.crawled_title.strip()
        words = full_name.split()
        first_word = words[0] if words else ""

        # 가격(숫자)
        try:
            base_price = int(digits_only(self.crawled_price))
        except Exception:
            base_price = 0

        # BJ 계산
        bj_price = 0
        if base_price > 0:
            if base_price <= 50000:
                bj_price = int(round(base_price * 1.3))
            elif base_price <= 100000:
                bj_price = int(round(base_price * 1.2))
            else:
                bj_price = base_price

        # BL = BJ * 1.05
        bl_price = int(round(bj_price * 1.05)) if bj_price > 0 else 0

        today_str = today_iso()

        # ==== 3) A~Z 채우기 ====
        ws.cell(row=row_idx, column=1).value  = ""         # A
        ws.cell(row=row_idx, column=2).value  = full_name  # B
        ws.cell(row=row_idx, column=3).value  = today_str  # C
        ws.cell(row=row_idx, column=4).value  = ""         # D
        ws.cell(row=row_idx, column=5).value  = "새상품"   # E
        ws.cell(row=row_idx, column=6).value  = ""         # F
        ws.cell(row=row_idx, column=7).value  = first_word # G
        ws.cell(row=row_idx, column=8).value  = first_word # H
        ws.cell(row=row_idx, column=9).value  = ""         # I

        for col in range(10, 27):                          # J~Z
            ws.cell(row=row_idx, column=col).value = ""

        # ==== 4) 확장 열 채우기 ====
        col_BJ = column_index_from_string("BJ")
        col_BL = column_index_from_string("BL")
        col_BM = column_index_from_string("BM")
        col_BN = column_index_from_string("BN")
        col_BX = column_index_from_string("BX")
        col_CK = column_index_from_string("CK")
        col_CZ = column_index_from_string("CZ")

        ws.cell(row=row_idx, column=col_BJ).value = bj_price
        ws.cell(row=row_idx, column=col_BL).value = bl_price
        ws.cell(row=row_idx, column=col_BM).value = 999
        ws.cell(row=row_idx, column=col_BN).value = 2
        ws.cell(row=row_idx, column=col_BX).value = "상세정보별도표기"
        ws.cell(row=row_idx, column=col_CK).value = "기타재화"
        ws.cell(row=row_idx, column=col_CZ).value = f"{row_idx}.png"

        try:
            wb.save(SELLERTOOL_XLSM_PATH)
            self._log(f"✅ 코스트코 상품 기록 완료 → 행 {row_idx}")
        except Exception as e:
            self._log(f"❌ 엑셀 저장 실패: {e}")
            return None

        # URL 클립보드 (선택)
        try:
            if self.crawled_url:
                pyperclip.copy(self.crawled_url)
                self._log("📋 코스트코 상품 URL을 클립보드에 복사했습니다.")
        except Exception as e:
            self._log(f"⚠️ 클립보드 복사 실패: {e}")

        return row_idx

    def _capture_costco_image(self, row_idx: int):
        """
        코스트코 상품 이미지 여러 장 저장 (다운로드 우선, 실패 시 캡처 백업)
        - 1순위: img.src / srcset 에서 직접 다운로드
        - 2순위: 다운로드 불가 시 → 애니메이션이 멈출 때까지 기다렸다가 요소 캡처

        파일명 규칙:
          - 첫 번째 이미지는  row_idx.png       (예: 5.png)
          - 두 번째부터는   row_idx-1.png, row_idx-2.png ...
        저장 위치:
          - SELLERTOOL_XLSM_PATH 폴더 기준 /YYYYMMDD/
        """
        try:
            driver = self._attach_driver()
        except Exception as e:
            self._log(f"❌ 코스트코 이미지 처리: 드라이버 연결 실패: {e}")
            return

        try:
            # 코스트코 상품 영역의 이미지들
            img_elements = driver.find_elements(By.CSS_SELECTOR, "picture img")
        except Exception as e:
            self._log(f"❌ 이미지 요소 검색 실패: {e}")
            return

        if not img_elements:
            self._log("⚠️ 처리할 picture img 요소를 찾지 못했습니다. 셀렉터를 점검해 주세요.")
            return

        # 날짜별 폴더 생성 (YYYYMMDD)
        base_dir = Path(SELLERTOOL_XLSM_PATH).parent
        date_folder = datetime.now().strftime("%Y%m%d")
        save_dir = base_dir / date_folder
        save_dir.mkdir(parents=True, exist_ok=True)

        seen_urls = set()
        saved_count = 0

        def _pick_image_url(el):
            """
            img 태그에서 src / srcset 중 가장 적절한 URL 하나를 골라 반환.
            """
            src = (el.get_attribute("src") or "").strip()
            srcset = (el.get_attribute("srcset") or "").strip()

            # srcset 이 있으면 가장 해상도 높은(마지막) 항목을 선택하는 간단 로직
            if srcset:
                try:
                    # "url1 300w, url2 600w, url3 1200w" 형태에서 마지막 url3 추출
                    parts = [p.strip() for p in srcset.split(",") if p.strip()]
                    last = parts[-1]
                    url_part = last.split()[0]
                    if url_part.startswith("http"):
                        return url_part
                except Exception:
                    pass

            if src.startswith("http"):
                return src

            return ""

        for idx, el in enumerate(img_elements):
            # 1) 이미지 URL 선택
            image_url = _pick_image_url(el)

            # 동일 URL 중복 제거
            if image_url and image_url in seen_urls:
                continue
            if image_url:
                seen_urls.add(image_url)

            # 파일명 구성
            if saved_count == 0:
                final_name = f"{row_idx}.png"
            else:
                final_name = f"{row_idx}-{saved_count}.png"

            temp_path = save_dir / f"{row_idx}_raw_{saved_count}.png"
            final_path = save_dir / final_name

            downloaded = False

            # ✅ [추가] 테스트 플래그가 켜져 있으면 다운로드 로직을 강제로 건너뛰기
            if not self.FORCE_CAPTURE_TEST:
                # 2) 먼저 브라우저(fetch + base64)로 다운로드 시도
                if image_url:
                    try:
                        self._log(f"🌐 [브라우저 fetch] 이미지 다운로드 시도: {image_url}")
                        start = time.time()

                        # ✅ 브라우저 세션을 그대로 활용해서 fetch
                        img_bytes = self._fetch_image_via_browser(driver, image_url, timeout=15.0)

                        elapsed = time.time() - start
                        self._log(f"⏱ 다운로드 소요시간(fetch+base64): {elapsed:.2f}초 | {image_url}")

                        if img_bytes:
                            # temp 파일로 한 번 저장 후, 기존 후처리 로직 재사용
                            with open(temp_path, "wb") as f:
                                f.write(img_bytes)

                            self._process_and_save_image_1000x1000(
                                temp_path, final_path
                            )

                            try:
                                temp_path.unlink()
                            except Exception:
                                pass

                            self._log(f"📥 브라우저 fetch 다운로드 성공 → {final_path.name}")
                            saved_count += 1
                            downloaded = True
                        else:
                            self._log("⚠️ 브라우저 fetch로 이미지를 가져오지 못했습니다.")
                    except Exception as e:
                        self._log(f"⚠️ 브라우저 fetch 이미지 다운로드 중 예외 발생: {e}")

            else:
                self._log("🧪 [TEST] FORCE_CAPTURE_TEST=True → 다운로드 단계 건너뛰고 캡처 경로로 이동")

            if saved_count == 0:
                # 다운로드 실패 시 캡처 시도
                try:
                    self._log(f"🌐 브라우저에서 이미지 직접 저장 시도")
                    self._save_image_from_browser(driver, el, final_path)
                    self._log(f"📥 브라우저 저장 성공 → {final_path.name}")
                    saved_count += 1
                    downloaded = True
                except Exception as e:
                    self._log(f"⚠️ 브라우저 이미지 저장 실패: {e}")


        if saved_count == 0:
            self._log("⚠️ 어떤 이미지도 저장하지 못했습니다.")
        else:
            self._log(f"✅ 총 {saved_count}장의 코스트코 이미지를 저장했습니다.")





    def _process_and_save_image_1000x1000(self, src_path: Path, dst_path: Path):
        """
        - 배경 제거(흰색을 투명으로 만드는 작업)를 하지 않는다.
        - 원본 비율을 유지하면서 긴 변 기준 1000 이하로 축소하고
        - 1000x1000 흰색 배경 캔버스에 중앙 정렬해서 저장한다.
        """
        try:
            img = Image.open(src_path).convert("RGB")

            # 긴 변 기준 1000으로 축소
            img.thumbnail((1000, 1000), Image.Resampling.LANCZOS)

            # 1000x1000 흰색 캔버스 위에 중앙 배치
            canvas = Image.new("RGB", (1000, 1000), (255, 255, 255))
            x = (1000 - img.width) // 2
            y = (1000 - img.height) // 2
            canvas.paste(img, (x, y))

            canvas.save(dst_path, format="PNG")
        except Exception as e:
            self._log(f"❌ 이미지 후처리 실패: {e}")




    def record_data(self):
        if not self.crawled_url:
            self._log("⚠️ 먼저 [대상윈도우]로 제목/가격/URL을 크롤링해 주세요.")
            return

        # === 1) 도메인에 따라 분기 ===
        host = urlparse(self.crawled_url or "").netloc.lower()

        # 코스트코: sellertool_upload.xlsm 에 기록 + 이미지 캡처
        if "costco.co.kr" in host:
            self._log("🧾 코스트코 상품으로 인식 → 엑셀 기록 + 이미지/스펙 캡처")
            row_idx = None
            try:
                row_idx = self._write_costco_to_seller_excel()
            except Exception as e:
                self._log(f"[오류] 코스트코 엑셀 기록 실패: {e}")

            if row_idx:
                # 상품 이미지 여러 장
                try:
                    self._capture_costco_image(row_idx)
                except Exception as e:
                    self._log(f"[오류] 코스트코 이미지 캡처 실패: {e}")

                # 스펙 영역 캡처
                try:
                    self._capture_costco_spec(row_idx)
                except Exception as e:
                    self._log(f"[오류] 코스트코 스펙 캡처 실패: {e}")

            return




        # (선택) 도매매 전용 분기도 가능하지만,
        # 현재는 '도매매 외 사이트도 구글시트 소싱목록에 기록' 구조이므로 그대로 둠.
        # if "domeme.domeggook.com" in host:
        #     ...  # 필요시 별도 처리

        # === 2) 나머지(도매매/네이버/기타)는 기존 로직 그대로 ===
        try:
            # ✅ 1) 먼저 구글 밑줄 실행 (에러가 나도 기록은 계속 진행)
            try:
                self.google_underline()
            except Exception as e:
                self._log(f"⚠️ 구글 밑줄 실행 중 오류(기록은 계속 진행): {e}")

            # ✅ 2) 그 다음 실제 데이터 기록 (소싱상품목록 시트)
            self._write_row_to_first_empty_a()

        except Exception as e:
            self._log(f"[오류] 시트 기록 실패: {e}")



    # ---------- 시트 클릭 대기 → 기록 ----------
    def _wait_for_sheet_click_then_write(self):
        if self._sheet_click_wait:
            return
        self._sheet_click_wait = True
        start_ts = time.time()
        self._log("⌛ 시트 클릭 대기 시작 (10초)")

        def wait_click():
            nonlocal start_ts
            with mouse.Events() as events:
                for event in events:
                    if (time.time() - start_ts) * 1000 >= CLICK_TIMEOUT_MS_RECORD:
                        break
                    if isinstance(event, mouse.Events.Click) and event.pressed:
                        self._sheet_click_wait = False
                        self.record_data()
                        return
            self._sheet_click_wait = False
            self._log("⏰ 10초 내 시트 클릭이 감지되지 않았습니다. [기록] 버튼으로 입력하세요.")

        import threading
        t = threading.Thread(target=wait_click, daemon=True)
        t.start()

    # ---------- 등록상품명 URL + 숫자 분리 ----------
    def _split_registered_name(self, text: str) -> tuple[str, str, str, str]:
        """
        '등록상품명' 문자열을 다음 4개로 분리:
          - orig : 원문 전체
          - num_part : 맨 앞의 숫자 (숫자 + 공백 패턴, 없으면 "")
          - mid_part : 숫자와 URL 사이의 중간 텍스트
          - url_part : https:// 로 시작하는 URL (여러 개면 첫 번째만)
        예)
          '10 샴푸 세트 https://example.com/abc'
            → num_part='10', mid_part='샴푸 세트', url_part='https://example.com/abc'
          '샴푸 세트 https://example.com/abc'
            → num_part='', mid_part='샴푸 세트', url_part='https://example.com/abc'
          '10 샴푸 세트'
            → num_part='10', mid_part='샴푸 세트', url_part=''
        """
        t = (text or "").strip()
        if not t:
            return "", "", "", ""

        # 1) URL 먼저 분리
        m_url = re.search(r'(https?://\S+)', t)
        url_part = ""
        before = t
        if m_url:
            url_part = m_url.group(1).rstrip('),].;\'"')  # 흔한 꼬리표 제거
            before = t[:m_url.start()].strip()
        else:
            before = t

        # 2) 맨 앞 숫자 분리 (숫자 + 공백 + 나머지)
        num_part = ""
        mid_part = ""
        m_num = re.match(r'^\s*(\d+)\s+(.*)$', before)
        if m_num:
            num_part = m_num.group(1)
            mid_part = m_num.group(2).strip()
        else:
            # 맨 앞에 숫자가 없으면, 전체를 중간 텍스트로 사용
            mid_part = before

        return t, num_part, mid_part, url_part


    # ==== 등록상품명(셀러상품 상세) 조회 유틸 ====
    def _cp_get_registered_product_name(self, seller_product_id: str) -> str | None:
        if not seller_product_id:
            return None
        if seller_product_id in self._cp_seller_name_cache:
            return self._cp_seller_name_cache[seller_product_id]
        paths = [
            f"/v2/providers/openapi/apis/api/v1/marketplace/seller-products/{seller_product_id}",
            f"/v2/providers/openapi/apis/api/v2/vendors/{COUPANG_VENDOR_ID}/seller-products/{seller_product_id}",
        ]
        for path in paths:
            try:
                data = _cp_request("GET", path, None)
                info = (data or {}).get("data") or {}
                name = info.get("sellerProductName") or info.get("name")
                if name:
                    self._cp_seller_name_cache[seller_product_id] = name
                    return name
            except Exception:
                continue
        return None

    # ==== 쿠팡 주문조회 + 시트기록 ====
    def _fetch_coupang_orders(self) -> list[dict]:
        if not (COUPANG_VENDOR_ID and COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY):
            self._log("❌ 쿠팡 API 키/벤더ID 설정이 비어 있습니다.")
            return []

        days = int(self.spin_days.value() if hasattr(self, "spin_days") else DEFAULT_LOOKBACK_DAYS)
        to_dt = datetime.now(timezone.utc)
        from_dt = to_dt - timedelta(days=days)
        created_from = from_dt.strftime("%Y-%m-%d")
        created_to = to_dt.strftime("%Y-%m-%d")
        self._log(f"🔍 조회기간: 최근 {days}일 (UTC {created_from} ~ {created_to})")

        path = f"/v2/providers/openapi/apis/api/v4/vendors/{COUPANG_VENDOR_ID}/ordersheets"
        all_rows: list[dict] = []

        for st in CP_QUERY_STATUSES:
            api_status_candidates = ORDER_STATUS_ALIASES.get(st, [st])
            status_succeeded = False

            for api_status in api_status_candidates:
                next_token = None
                while True:
                    params = {
                        "createdAtFrom": created_from,
                        "createdAtTo": created_to,
                        "status": api_status,
                        "maxPerPage": 50,
                    }
                    if next_token:
                        params["nextToken"] = next_token

                    try:
                        data = _cp_request("GET", path, params)
                    except requests.HTTPError as e:
                        resp = getattr(e, "response", None)
                        body = ""
                        try:
                            body = resp.text or ""
                        except Exception:
                            pass
                        if getattr(resp, "status_code", None) == 400 and "Invalid Status" in body:
                            self._log(f"ℹ️ 상태 '{api_status}' 미허용 → 다음 후보로 폴백 시도")
                            break
                        self._log_http_error(e, context=f"쿠팡 API 호출 실패(status={st}, api_status={api_status})")
                        break
                    except Exception as e:
                        self._log(f"⚠️ 쿠팡 API 호출 실패(status={st}, api_status={api_status}): {repr(e)}")
                        break

                    result_code = str(data.get("code", "")).upper()
                    if result_code and result_code not in ("SUCCESS", "OK", "200"):
                        msg = safe_str(data.get("message"))
                        if "Invalid Status" in msg:
                            self._log(f"ℹ️ 상태 '{api_status}' 미허용(code={result_code}) → 다음 후보로 폴백")
                            break
                        self._log(f"⚠️ 응답 코드 이상(status={st}, api_status={api_status}): {msg}")
                        break

                    datas = data.get("data")
                    if isinstance(datas, list):
                        sheets = datas
                    elif isinstance(datas, dict):
                        sheets = (
                            datas.get("orderSheets")
                            or datas.get("shipmentBoxInfos")
                            or datas.get("items")
                            or []
                        )
                    else:
                        sheets = []
                    if isinstance(sheets, dict):
                        sheets = [sheets]
                    if not isinstance(sheets, list):
                        sheets = []

                    for sheet in sheets:
                        if not isinstance(sheet, dict):
                            continue
                        order_id = (sheet.get("orderId") or sheet.get("orderIdMask") or sheet.get("orderNo") or "")
                        order_date = (sheet.get("orderedAt") or sheet.get("orderDate") or sheet.get("orderTime") or "")
                        receiver = sheet.get("receiver") or {}
                        if not isinstance(receiver, dict):
                            receiver = {}
                        recv_name = (receiver.get("name") or receiver.get("receiverName") or "")
                        recv_addr = (receiver.get("addr1") or receiver.get("address1") or receiver.get("address") or "")
                        recv_phone = (receiver.get("contact1") or receiver.get("contact2") or receiver.get("phone") or "")
                        items = (sheet.get("orderItems") or sheet.get("orderSheetItems") or sheet.get("items") or [])
                        if isinstance(items, dict):
                            items = [items]
                        if not isinstance(items, list):
                            items = []

                        for it in items:
                            if not isinstance(it, dict):
                                continue
                            item_name = (
                                it.get("sellerProductName")
                                or it.get("vendorItemName")
                                or it.get("productName")
                                or ""
                            )
                            order_item_id = it.get("orderItemId") or it.get("vendorItemId") or ""
                            
                            # 수량: ordersheets 문서 기준 shippingCount 가 정식 필드
                            qty = it.get("shippingCount") or it.get("quantity") or 1

                            # 결제금액: orderPrice / salesPrice × 수량 등 복합 로직
                            paid_price = extract_paid_price_from_item(it)

                            
                            tracking_no = it.get("invoiceNumber") or it.get("trackingNumber") or ""
                            carrier = it.get("deliveryCompanyName") or it.get("deliveryCompanyCode") or ""
                            status_text = CP_STATUS_MAP.get(st, st)

                            seller_product_id = (
                                it.get("sellerProductId")
                                or sheet.get("sellerProductId")
                                or ""
                            )
                            registered_name = (
                                it.get("sellerProductName")
                                or (self._cp_get_registered_product_name(str(seller_product_id)) if seller_product_id else None)
                                or ""
                            )

                            orig_reg, reg_num, reg_mid, reg_url = self._split_registered_name(registered_name)


                            all_rows.append({
                                "주문일시": order_date,
                                "상태": status_text,
                                "주문번호": order_id,
                                "주문아이템ID": order_item_id,

                                # 등록상품명 관련 4분할
                                "등록상품명":   orig_reg,   # 원문 전체
                                "등록상품명-1": reg_num,    # 맨 앞 숫자 (없으면 "")
                                "등록상품명-2": reg_mid,    # 숫자와 URL 사이 텍스트
                                "등록상품명-3": reg_url,    # URL

                                "수량": qty,
                                "결제금액": paid_price,

                                # ★ 새 컬럼: 최종 수익 (초기에는 빈 값, 나중에 '주문정리' 버튼에서 채움)
                                "최종 수익": "",

                                "수취인": recv_name,
                                "연락처": recv_phone,
                                "주소": recv_addr,
                                "송장번호": tracking_no,
                                "택배사": carrier,

                                "셀러상품ID": str(seller_product_id or ""),
                            })



                    next_token = None
                    # 1) data["data"] 는 리스트이므로, 여기서 nextToken 을 찾지 말고
                    # 2) 응답 최상위에서 nextToken 을 읽어야 함
                    if isinstance(data, dict):
                        nt = data.get("nextToken")
                        if nt:
                            next_token = nt
                    if not next_token:
                        status_succeeded = True
                        break  # while
                # 다음 상태 별칭으로 폴백
            if not status_succeeded:
                self._log(f"ℹ️ 상태 '{st}'는 제공 계정/엔드포인트 조합에서 미허용이거나 데이터가 없습니다.")

        # 정렬: 상태(비즈니스 순서) → 주문일시(최신우선)
        def _parse_dt_safe(s: str):
            s = (s or "").strip()
            if not s:
                return None
            try:
                if s.endswith("Z"):
                    return datetime.fromisoformat(s.replace("Z", "+00:00"))
                return datetime.fromisoformat(s)
            except Exception:
                m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
                if m:
                    try:
                        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    except Exception:
                        return None
                return None

        def _sort_key(row: dict):
            st_txt = str(row.get("상태", ""))
            st_rank = STATUS_ORDER.get(st_txt, 999)
            dt = _parse_dt_safe(row.get("주문일시"))
            ts = -(dt.timestamp()) if dt else float("inf")
            return (st_rank, ts)

        all_rows.sort(key=_sort_key)

        self._log(f"📦 쿠팡 주문 수집 완료: {len(all_rows)}건")
        return all_rows

    def _write_coupang_orders_to_sheet(self, rows: list[dict]):
        if self.sheets.ws is None:
            self._log("⚠️ Sheets 연결이 필요합니다. 먼저 [Sheets 연결] 버튼을 눌러주세요.")
            return
        try:
            ws = self.sheets.gc.open_by_key(SHEET_ID).worksheet(COUPANG_WS_NAME)
        except gspread.WorksheetNotFound:
            ws = self.sheets.gc.open_by_key(SHEET_ID).add_worksheet(title=COUPANG_WS_NAME, rows=4000, cols=40)

        if not rows:
            # 등록상품명 3분할 + 최종 수익 헤더
            headers = [
                "주문일시","상태","주문번호","주문아이템ID",
                "등록상품명","등록상품명-1","등록상품명-2","등록상품명-3",
                "수량","결제금액","최종 수익","수취인","연락처","주소","송장번호","택배사","셀러상품ID"
            ]
            ws.clear()
            # 컬럼 수 17개 → A~Q 이지만, 헤더는 17개라 실제로는 A~Q 중 1칸은 비게 됩니다.
            # 크게 문제는 없으니 그대로 두셔도 되고, 엄밀하게 맞추려면 A1:Q1 → A1:Q1 그대로 둬도 무방합니다.
            ws.update(values=[headers], range_name="A1:Q1")
            self._log("ℹ️ 쿠팡 주문 데이터가 없어 헤더만 갱신했습니다.")
            return



        headers = list(rows[0].keys())
        values = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]

        ws.clear()
        end_col_letter = _a1_col(len(headers))
        rng = f"A1:{end_col_letter}{len(values)}"
        ws.update(values=values, range_name=rng, value_input_option="USER_ENTERED")
        self._log(f"✅ '{COUPANG_WS_NAME}' 탭에 {len(rows)}건 업데이트 완료")

    # === 금일 올린 상품 갯수 계산 ===
    def update_today_product_count(self):
        """B열의 '오늘 날짜' 구간에 해당하는 A열 값으로 금일 올린 상품 개수를 계산해서 라벨에 표시."""
        # Sheets 연결 확인
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(금일 올린 상품 갯수 계산): {e}")
                self.lbl_today_count.setText("금일 올린 상품 갯수 : 오류")
                return

        try:
            ws = self.sheets.ws

            # B열 전체 값 가져오기
            col_b = ws.col_values(2)
            today = today_fmt()

            # 오늘 날짜가 들어있는 행 번호들(1-based)
            today_rows = [
                idx + 1
                for idx, v in enumerate(col_b)
                if str(v).strip() == today
            ]

            if not today_rows:
                count = 0
                self._log(f"📊 오늘 날짜({today}) 데이터가 B열에 없어 0개로 계산합니다.")
            else:
                first_row = today_rows[0]
                last_row = today_rows[-1]

                # A열 값 읽어서 번호 기준으로 계산
                col_a = ws.col_values(1)
                a_first = str(col_a[first_row - 1]).strip() if len(col_a) >= first_row else ""
                a_last = str(col_a[last_row - 1]).strip() if len(col_a) >= last_row else ""

                try:
                    n_first = int(a_first)
                    n_last = int(a_last)
                    # 하단 A - 상단 A + 1
                    count = n_last - n_first + 1
                    if count < 0:
                        # 혹시라도 값이 꼬여 있으면 fallback
                        count = len(today_rows)
                except Exception:
                    # A열이 숫자가 아니면, 단순히 오늘 날짜가 들어간 행 개수로 계산
                    count = len(today_rows)

                self._log(
                    f"📊 금일 상품 갯수 계산: A({a_first})~A({a_last}) → {count}개"
                )

            # 라벨 업데이트
            self.lbl_today_count.setText(f"금일 상품 갯수 : {count}")

        except Exception as e:
            self._log(f"❌ 금일 상품 갯수 계산 중 오류: {e}")
            self.lbl_today_count.setText("금일 상품 갯수 : 오류")



    # === 쿠팡 주문현황 버튼 동작 ===
    def coupang_orders(self):
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패: {e}")
                return
        try:
            rows = self._fetch_coupang_orders()
        except Exception as e:
            self._log(f"❌ 쿠팡 주문 조회 중 오류: {e}")
            return
        try:
            self._write_coupang_orders_to_sheet(rows)
        except Exception as e:
            self._log(f"❌ 쿠팡 주문 기록 중 오류: {e}")

    # === (통합) 쿠팡 키 확인 + 헬스체크 버튼 동작 ===
    def coupang_key_and_health(self):
        self.check_coupang_keys()
        self.coupang_healthcheck()

    # === 쿠팡 키 확인 ===
    def check_coupang_keys(self):
        try:
            p = Path(COUPANG_KEYS_JSON)
            if not p.exists():
                self._log(f"❌ 키 파일을 찾지 못했습니다: {COUPANG_KEYS_JSON}")
                self._log("➡ 경로/파일명을 다시 확인하거나 JSON을 생성해 주세요.")
                return
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            vendor_id = (data.get("vendor_id") or "").strip()
            access_key = (data.get("access_key") or "").strip()
            secret_key = (data.get("secret_key") or "").strip()

            self._log("✅ JSON 파일 읽기 성공")
            self._log(f"• Vendor ID: {vendor_id or '(빈 값)'}")
            self._log(f"• Access Key: {access_key or '(빈 값)'}")
            self._log(f"• Secret Key: {_mask(secret_key) if secret_key else '(빈 값)'}")

            problems = []
            if not vendor_id: problems.append("vendor_id가 비어 있습니다.")
            if not access_key: problems.append("access_key가 비어 있습니다.")
            if not secret_key: problems.append("secret_key가 비어 있습니다.")
            if problems:
                for m in problems:
                    self._log(f"⚠️ {m}")
                return

            mismatches = []
            if COUPANG_VENDOR_ID != vendor_id:
                mismatches.append("전역 Vendor ID와 JSON의 vendor_id가 다릅니다.")
            if COUPANG_ACCESS_KEY != access_key:
                mismatches.append("전역 Access Key와 JSON의 access_key가 다릅니다.")
            if COUPANG_SECRET_KEY != secret_key:
                mismatches.append("전역 Secret Key와 JSON의 secret_key가 다릅니다.")
            if mismatches:
                self._log("⚠️ 전역 설정과 JSON 파일의 값이 일치하지 않습니다:")
                for m in mismatches:
                    self._log(f"   - {m}")
                self._log("➡ JSON을 수정했으면 프로그램을 재시작하거나, 상단 상수 경로/로딩 부분을 확인하세요.")

            # 간단 HMAC 메시지 생성 테스트
            try:
                test_path = f"/v2/providers/openapi/apis/api/v4/vendors/{vendor_id}/ordersheets"
                test_query = urlencode({"status": "ACCEPT", "maxPerPage": 50}, doseq=True)
                signed_date = datetime.now(timezone.utc).strftime("%y%m%dT%H%M%SZ")
                msg = f"{signed_date}{'GET'}{test_path}{test_query}"
                signature = hmac.new(secret_key.encode("utf-8"), msg.encode("utf-8"), hashlib.sha256).hexdigest()
                auth_head = (
                    f"CEA algorithm=HmacSHA256, access-key={access_key}, "
                    f"signed-date={signed_date}, signature={signature}"
                )
                self._log("🔐 HMAC 서명 생성 테스트 성공")
                self._log(f"• Authorization 헤더 앞부분: {auth_head[:60]}...")
            except Exception as e:
                self._log(f"❌ HMAC 서명 생성 실패: {e}")

            self._log("🟢 키 확인 완료")

        except json.JSONDecodeError as e:
            self._log(f"❌ JSON 파싱 실패: {e}")
            self._log("➡ 파일 내용이 유효한 JSON 형식인지 확인하세요.")
        except Exception as e:
            self._log(f"❌ 키 확인 중 오류: {e}")

    # === 쿠팡 API 헬스체크 ===
    def coupang_healthcheck(self):
        self._log("🩺 쿠팡 API 헬스체크 시작")
        if not (COUPANG_VENDOR_ID and COUPANG_ACCESS_KEY and COUPANG_SECRET_KEY):
            self._log("❌ 쿠팡 키/벤더ID가 비어 있습니다. coupang_keys.json 확인")
            return
        try:
            to_dt = datetime.now(timezone.utc)
            from_dt = to_dt - timedelta(days=1)  # 헬스체크는 간단히 최근 1일로 확인
            path = f"/v2/providers/openapi/apis/api/v4/vendors/{COUPANG_VENDOR_ID}/ordersheets"
            param_variants = _build_ordersheets_params(from_dt, to_dt, status="ACCEPT", max_per_page=1)
            data = _try_ordersheets_with_variants(path, param_variants)
            code = str(data.get("code", "")).upper()
            self._log(f"✅ 헬스체크 성공: path='{path}', params={param_variants[0]} (code={code or 'N/A'})")
            self._log("🟢 쿠팡 API 키/서명/경로 정상으로 보입니다.")
            return
        except requests.HTTPError as e:
            self._log_http_error(e, context="헬스체크(ordersheets) 실패")
        except Exception as e:
            self._log(f"❌ 헬스체크(ordersheets) 중 예외: {repr(e)}")
        self._log("❌ 헬스체크가 실패했습니다. 다음을 점검해 주세요:\n"
                  "  1) 판매자센터(Wing) OpenAPI 키 여부 (파트너스 키 아님)\n"
                  "  2) 시스템연동 > Open API 사용 활성 및 권한 승인\n"
                  "  3) 허용 IP에 현재 PC 공인 IP 등록\n"
                  "  4) PC 시간 자동 동기화(UTC, 수초 이하 오차)\n")
    
    # === 쿠팡주문현황 '최종 수익' 채우기 (주문정리) ===
    def settle_orders(self):
        """쿠팡주문현황 시트에서 등록상품명-1/결제금액으로 소싱상품목록의 O열 값을 찾아와 K열(최종 수익)에 채워 넣는다."""
        # 1) Sheets 연결 확인
        if self.sheets.gc is None or self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(주문정리): {e}")
                return

        try:
            sh = self.sheets.gc.open_by_key(SHEET_ID)
            ws_orders = sh.worksheet(COUPANG_WS_NAME)
        except Exception as e:
            self._log(f"❌ '쿠팡주문현황' 시트를 찾지 못했습니다: {e}")
            return

        try:
            # 소싱상품목록 시트
            try:
                ws_source = sh.worksheet(WORKSHEET_NAME)
            except Exception:
                # 이미 self.sheets.ws 가 소싱상품목록이면 그걸 사용
                ws_source = self.sheets.ws

            # 2) 쿠팡주문현황 전체 값 가져오기
            orders_values = ws_orders.get_all_values()
            if not orders_values or len(orders_values) < 2:
                self._log("ℹ️ '쿠팡주문현황' 시트에 데이터가 없습니다.")
                return

            header = orders_values[0]

            # 필요한 컬럼 인덱스 찾기
            try:
                idx_reg1 = header.index("등록상품명-1")
                idx_paid = header.index("결제금액")
            except ValueError:
                self._log("❌ '등록상품명-1' 또는 '결제금액' 헤더를 찾지 못했습니다. 헤더명을 확인해 주세요.")
                return

            # '최종 수익' 컬럼 인덱스 확보 (없으면 결제금액 바로 오른쪽에 새로 추가)
            if "최종 수익" in header:
                idx_profit = header.index("최종 수익")
            else:
                idx_profit = idx_paid + 1
                # 모든 행에 대해 '최종 수익' 컬럼을 삽입 (초기값 "")
                for r in range(len(orders_values)):
                    row = orders_values[r]
                    # 결제금액 위치까지는 최소 길이 확보
                    while len(row) <= idx_paid:
                        row.append("")
                    row.insert(idx_profit, "")

                # 헤더 다시 갱신
                header = orders_values[0]
                self._log(f"🆕 '최종 수익' 컬럼을 추가했습니다. (열 인덱스: {idx_profit+1})")

            # 3) 소싱상품목록에서 (A,Q)->O 매핑 생성
            source_values = ws_source.get_all_values()
            if not source_values or len(source_values) < 2:
                self._log("ℹ️ '소싱상품목록' 시트에 데이터가 없어 주문정리를 건너뜁니다.")
                return

            profit_map = {}  # key: (A값, Q값) → O값
            # 0-based: A=0, O=14, Q=16
            for srow in source_values[1:]:
                if len(srow) < 17:
                    continue
                a_val = digits_only(srow[0])
                q_val = digits_only(srow[16])
                if not a_val or not q_val:
                    continue
                key = (a_val, q_val)
                # 같은 키가 여러 번 나올 수 있지만, 첫 번째 값만 사용
                if key not in profit_map:
                    o_val = srow[14] if len(srow) > 14 else ""
                    profit_map[key] = o_val

            self._log(f"📚 소싱상품목록 매핑 생성 완료: {len(profit_map)}개 키")

            # 4) 쿠팡주문현황 각 행에 대해 '최종 수익' 채우기
            updated_count = 0
            max_idx = max(idx_reg1, idx_paid, idx_profit)

            for i in range(1, len(orders_values)):  # 2행부터
                row = orders_values[i]
                # 최소 길이 확보
                if len(row) <= max_idx:
                    row.extend([""] * (max_idx + 1 - len(row)))

                reg_code = digits_only(row[idx_reg1])
                paid_val = digits_only(row[idx_paid])

                if not reg_code or not paid_val:
                    continue

                key = (reg_code, paid_val)
                profit_val = profit_map.get(key, "")

                if profit_val:
                    row[idx_profit] = profit_val
                    updated_count += 1

            # 5) 시트에 다시 반영
            end_col_letter = _a1_col(max(len(r) for r in orders_values))
            end_row = len(orders_values)
            rng = f"A1:{end_col_letter}{end_row}"
            ws_orders.update(rng, orders_values, value_input_option="USER_ENTERED")

            self._log(f"✅ 주문정리 완료: {updated_count}건에 '최종 수익'(열 K)을 반영했습니다.")

        except Exception as e:
            self._log(f"❌ 주문정리 처리 중 오류: {e}")

    
     
    # === 구글시트 A열 첫 빈 행 상단 테두리 (구글 밑줄) ===
    def google_underline(self):
        """A열의 비어있는 첫 번째 셀을 포함한 행 전체에 '윗부분'만 테두리를 긋기."""
        # Sheets 연결 확인
        if self.sheets.ws is None:
            self._log("ℹ️ Sheets 미연결: 자동으로 연결 시도합니다.")
            try:
                self.connect_sheets()
            except Exception as e:
                self._log(f"❌ Sheets 연결 실패(구글 밑줄): {e}")
                return

        try:
            ws = self.sheets.ws

            # 1) A열에서 비어있는 첫 번째 행 찾기
            target_row = self.sheets.find_first_empty_row_in_col_a_from_top()
            self._log(f"🔎 A열 기준 첫 빈 행: {target_row}행")

            # 2) sheetId 가져오기
            try:
                sheet_id = ws.id  # gspread 최신 버전 속성
            except AttributeError:
                sheet_id = ws._properties.get("sheetId")

            if sheet_id is None:
                self._log("❌ sheetId를 가져오지 못했습니다. (ws.id / ws._properties['sheetId'] 확인 필요)")
                return

            # 3) 몇 번째 컬럼까지 테두리를 칠지 결정
            #    - 헤더(row 1)의 실제 사용 컬럼 수를 기준으로, 최소 10컬럼 이상은 잡도록 처리
            try:
                header_values = ws.row_values(1)
                used_cols = max(len(header_values), 10)
            except Exception:
                used_cols = 10

            # Google Sheets API index는 0-based 이므로 변환
            start_row_index = target_row - 1
            end_row_index = target_row
            start_col_index = 0          # A열
            end_col_index = used_cols    # A ~ (used_cols)열

            body = {
                "requests": [
                    {
                        "updateBorders": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": start_row_index,
                                "endRowIndex": end_row_index,
                                "startColumnIndex": start_col_index,
                                "endColumnIndex": end_col_index,
                            },
                            # 윗부분 테두리만 적용
                            "top": {
                                "style": "SOLID",
                                "width": 1,
                                "color": {
                                    "red": 0,
                                    "green": 0,
                                    "blue": 0,
                                }
                            }
                            # left/right/bottom은 지정하지 않아 기존 스타일 유지
                        }
                    }
                ]
            }

            # 4) batch_update 실행
            ws.spreadsheet.batch_update(body)
            self._log(
                f"✅ 구글 밑줄 적용 완료: 행 {target_row} (A{target_row} ~ { _a1_col(end_col_index) }{target_row}) 상단 테두리"
            )

        except Exception as e:
            self._log(f"❌ 구글 밑줄 처리 중 오류: {e}")
            
    def _capture_costco_spec(self, row_idx: int):
        """
        코스트코 상품 페이지의 '스펙' 패널을 열고
        파란색 스펙 영역(div.mat-expansion-panel-body)만 그대로 캡처해서 저장한다.
        - 파일명: {row_idx}_spec.png
        - 경로: sellertool_upload.xlsm 이 있는 폴더 아래 /YYYYMMDD/
        """
        try:
            driver = self._attach_driver()
        except Exception as e:
            self._log(f"❌ 코스트코 스펙 캡처: 드라이버 연결 실패: {e}")
            return

        try:
            # 1) '스펙' 패널 열고 body 요소 받기
            spec_body = self._open_costco_spec_section()

            # 2) 저장 폴더 준비 (날짜별)
            base_dir = Path(SELLERTOOL_XLSM_PATH).parent
            date_folder = datetime.now().strftime("%Y%m%d")
            save_dir = base_dir / date_folder
            save_dir.mkdir(parents=True, exist_ok=True)

            save_path = save_dir / f"{row_idx}_spec.png"

            # 3) 요소 스크린샷: 스펙 내용 길이에 맞게 자동으로 캡처
            spec_body.screenshot(str(save_path))

            self._log(f"✅ 스펙 캡처 완료: {save_path}")

        except TimeoutException:
            self._log("❌ '스펙' 패널 또는 내용 영역을 찾지 못했습니다. 코스트코 페이지 구조를 다시 한 번 확인해 주세요.")
        except Exception as e:
            self._log(f"❌ 코스트코 스펙 캡처 중 오류: {e}")

    def _open_costco_spec_section(self):
        """
        코스트코 상품 페이지에서
        - 헤더 텍스트에 '스펙' 이 들어가는 아코디언 패널을 찾고
        - 접혀 있으면 클릭해서 열고
        - 그 패널 안의 내용 영역(파란 영역: mat-expansion-panel-content/body)을 리턴한다.
        """
        driver = self._attach_driver()
        wait = WebDriverWait(driver, 10)

        # 1) '스펙' 이라는 텍스트를 가진 아코디언 헤더 찾기
        #    (태그가 div 가 아니라 <mat-expansion-panel-header> 이라서 * 로 잡습니다)
        header_xpath = (
            "//*[contains(@class,'mat-expansion-panel-header') and "
            " .//*[contains(normalize-space(),'스펙')]]"
        )

        spec_header = wait.until(
            EC.element_to_be_clickable((By.XPATH, header_xpath))
        )

        # 화면 가운데로 스크롤
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", spec_header
        )
        time.sleep(0.3)

        # 2) 이 헤더가 속한 패널(<mat-expansion-panel> 또는 div.mat-expansion-panel)을 찾기
        parent_panel = spec_header.find_element(
            By.XPATH, "ancestor::*[contains(@class,'mat-expansion-panel')][1]"
        )

        # 3) 이미 펼쳐져 있는지 확인 (aria-expanded 또는 클래스에 mat-expanded 여부)
        expanded_attr = spec_header.get_attribute("aria-expanded") or ""
        if not expanded_attr:
            expanded_attr = parent_panel.get_attribute("class") or ""

        if ("true" not in expanded_attr.lower()
                and "mat-expanded" not in expanded_attr):
            spec_header.click()
            # 펼쳐질 때까지 잠깐 대기
            try:
                wait.until(
                    lambda d: (
                        "mat-expanded" in (parent_panel.get_attribute("class") or "")
                        or spec_header.get_attribute("aria-expanded") == "true"
                    )
                )
            except Exception:
                pass  # 너무 빡빡하게 볼 필요는 없어서 실패해도 그냥 진행

        # 4) 이 패널 안의 내용 영역(파란 영역) 찾기
        #    - 실제 페이지는 mat-expansion-panel-content 가 상위 컨테이너
        #    - 혹시 버전에 따라 body 클래스를 쓰면 그것도 같이 허용
        body_xpath = (
            ".//*[contains(@class,'mat-expansion-panel-content') "
            "   or contains(@class,'mat-expansion-panel-body')]"
        )

        spec_body = parent_panel.find_element(By.XPATH, body_xpath)

        # 스펙 내용이 길어도 보이도록 다시 가운데 스크롤
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", spec_body
        )
        time.sleep(0.3)

        self._log("🟢 '스펙' 패널 열기 및 body 요소 찾기 완료")
        return spec_body

    def _wait_until_element_stable(
        self,
        driver,
        element,
        check_interval: float = 0.2,
        stable_checks: int = 3,
        max_wait: float = 10.0,
    ) -> bool:
        """
        요소의 getBoundingClientRect() 값이 일정 시간 동안 변하지 않을 때까지 기다린다.
        - 슬라이드/애니메이션이 끝나고 '화면에서 고정된 상태'가 되었을 때 True 반환.
        - max_wait 동안 안정되지 않으면 False 반환.
        """
        import time

        last_rect = None
        stable_count = 0
        end_time = time.time() + max_wait

        while time.time() < end_time:
            try:
                rect = driver.execute_script(
                    """
                    const r = arguments[0].getBoundingClientRect();
                    return [r.x, r.y, r.width, r.height];
                    """,
                    element,
                )
            except Exception:
                # 요소가 더 이상 없으면 안정화 의미가 없으니 종료
                break

            if rect == last_rect:
                stable_count += 1
            else:
                stable_count = 0
            last_rect = rect

            if stable_count >= stable_checks:
                return True

            time.sleep(check_interval)

        return False

    def _save_image_from_browser(self, driver, img_element, save_path):
        # 브라우저에서 이미 로드된 이미지를 base64로 추출
        img_data = driver.execute_script("""
            var img = arguments[0];
            var canvas = document.createElement('canvas');
            canvas.width = img.naturalWidth;
            canvas.height = img.naturalHeight;
            var ctx = canvas.getContext('2d');
            ctx.drawImage(img, 0, 0);
            return canvas.toDataURL('image/png').split(',')[1];
        """, img_element)
        img_bytes = base64.b64decode(img_data)
        img = Image.open(io.BytesIO(img_bytes))
        img.save(save_path, format="PNG")
        
    def _fetch_image_via_browser(self, driver, url: str, timeout: float = 15.0) -> bytes | None:
        """
        브라우저(JS fetch)를 이용해 image URL을 가져온 뒤,
        base64 문자열로 Python에 전달해서 bytes 로 반환한다.
        - Chrome 세션 쿠키/헤더/연결을 그대로 활용할 수 있음.
        """
        if not url:
            return None

        script = """
        const url = arguments[0];
        const callback = arguments[arguments.length - 1];

        fetch(url, { credentials: 'include' })
          .then(resp => {
            if (!resp.ok) {
              throw new Error('HTTP ' + resp.status);
            }
            return resp.arrayBuffer();
          })
          .then(buf => {
            const bytes = new Uint8Array(buf);
            let binary = '';
            const len = bytes.byteLength;
            for (let i = 0; i < len; i++) {
              binary += String.fromCharCode(bytes[i]);
            }
            // base64 로 인코딩해서 콜백으로 넘김
            callback(btoa(binary));
          })
          .catch(err => {
            callback(null);
          });
        """

        try:
            # Selenium의 async script 사용 (마지막 인수가 callback)
            driver.set_script_timeout(timeout)
            b64_data = driver.execute_async_script(script, url)
        except Exception as e:
            self._log(f"⚠️ 브라우저 fetch 실행 중 오류: {e}")
            return None

        if not b64_data:
            return None

        try:
            return base64.b64decode(b64_data)
        except Exception as e:
            self._log(f"⚠️ 브라우저 fetch base64 디코딩 실패: {e}")
            return None

        
        

# =========================
# 엔트리 포인트
# =========================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = ChromeCrawler()
    win.show()
    sys.exit(app.exec())
