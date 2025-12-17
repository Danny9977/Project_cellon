import json
import hmac
import hashlib
import requests
import time
from datetime import datetime, timezone
from urllib.parse import urlencode
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import digits_only, is_int_string
from typing import Mapping, Any

from cellon.config import (
    COUPANG_KEYS_JSON, COUPANG_BASE_URL,
    SERVICE_ACCOUNT_JSON, SHEET_ID, WORKSHEET_NAME,
    COUPANG_WS_NAME,
    FIXED_CONST_FEE, DEFAULT_LOOKBACK_DAYS
)

# APIs에서 필요한 함수 import
from .apis.coupang_client import (
    load_coupang_keys,
    cp_request,
    build_ordersheets_params,
    try_ordersheets_with_variants,
)


def extract_money_amount(m: dict | None) -> int:
    if not isinstance(m, dict):
        return 0
    units = m.get("units", 0)
    nanos = m.get("nanos", 0)
    try:
        units = int(units)
    except Exception:
        units = 0
    try:
        nanos = int(nanos)
    except Exception:
        nanos = 0
    if nanos:
        return units + round(nanos / 1_000_000_000)
    return units

def extract_paid_price_from_item(it: dict) -> int:
    if not isinstance(it, dict):
        return 0
    op = it.get("orderPrice")
    if isinstance(op, dict):
        v = extract_money_amount(op)
        if v:
            return v
    if op is not None and not isinstance(op, dict):
        s = digits_only(op)
        if s:
            try:
                return int(s)
            except Exception:
                pass
    sales = it.get("salesPrice")
    sales_val = 0
    if isinstance(sales, dict):
        sales_val = extract_money_amount(sales)
    elif sales is not None:
        s = digits_only(sales)
        if s:
            try:
                sales_val = int(s)
            except Exception:
                sales_val = 0
    qty = it.get("shippingCount") or it.get("quantity") or 1
    try:
        qty = int(qty)
    except Exception:
        qty = 1
    if sales_val and qty:
        return sales_val * qty
    for key in ("paidPrice", "paymentAmount", "price"):
        if key in it and it[key] is not None:
            s = digits_only(it[key])
            if s:
                try:
                    return int(s)
                except Exception:
                    pass
    return 0



class SheetsClient:
    def __init__(self, json_path: str, sheet_id: str, worksheet_name: str, logger):
        self.json_path = json_path
        self.sheet_id = sheet_id
        self.worksheet_name = worksheet_name
        self.logger = logger
        self.gc = None
        self.ws = None
        self.CREATE_WORKSHEET_IF_MISSING = False

    def connect(self):
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_file(self.json_path, scopes=scopes)
        self.gc = gspread.authorize(creds)
        sh = self.gc.open_by_key(self.sheet_id)
        try:
            self.ws = sh.worksheet(self.worksheet_name)
            self.logger(f"✅ Google Sheets 연결 완료 (워크시트: {self.worksheet_name})")
        except gspread.WorksheetNotFound:
            titles = [w.title for w in sh.worksheets()]
            self.logger(f"⚠️ 워크시트 '{self.worksheet_name}'를 찾지 못함. 현재 탭들: {titles}")
            if self.CREATE_WORKSHEET_IF_MISSING:
                self.ws = sh.add_worksheet(title=self.worksheet_name, rows=1000, cols=30)
                self.logger(f"🆕 워크시트 생성: {self.worksheet_name}")
            else:
                raise

    def get_next_index(self) -> int:
        try:
            col_values = self.ws.col_values(1)
            last = None
            for v in reversed(col_values):
                if v.strip():
                    last = v
                    break
            if last is None:
                return 1
            return int(last) + 1 if is_int_string(last) else 1
        except Exception as e:
            self.logger(f"⚠️ A열 인덱스 계산 실패, 1로 시작: {e}")
            return 1

    def find_first_empty_row_in_col_a_from_top(self) -> int:
        values = self.ws.col_values(1)
        if not values:
            return 1
        for i, v in enumerate(values, start=1):
            if not str(v).strip():
                return i
        return len(values) + 1

    def append_row_with_retry(self, row_values, max_tries=5, base_sleep=0.6):
        attempt = 0
        while True:
            try:
                self.ws.append_row(row_values, value_input_option="USER_ENTERED")
                return True
            except gspread.exceptions.APIError as e:
                attempt += 1
                try:
                    resp = getattr(e, "response", None)
                    status = getattr(resp, "status_code", None)
                    text = getattr(resp, "text", "")
                    self.logger(f"❌ APIError(status={status}): {text[:500]}")
                except Exception:
                    self.logger(f"❌ APIError: {e}")
                if attempt >= max_tries:
                    return False
                sleep_s = base_sleep * (2 ** (attempt - 1))
                self.logger(f"⏳ 재시도 {attempt}/{max_tries} ... {sleep_s:.1f}s")
                time.sleep(sleep_s)
            except Exception as e:
                attempt += 1
                self.logger(f"❌ 전송/기타 오류: {repr(e)}")
                if attempt >= max_tries:
                    return False
                sleep_s = base_sleep * (2 ** (attempt - 1))
                self.logger(f"⏳ 재시도 {attempt}/{max_tries} ... {sleep_s:.1f}s")
                time.sleep(sleep_s)
                
    # ====== 로컬 엑셀(.xlsx)에 값 쓰기 헬퍼들 ======
    def write_cells_to_excel(xlsx_path: str,
                            sheet_name: str,
                            cell_value_map: dict[str, object]) -> None:
        """
        단순 A1, B5 처럼 "셀 주소 → 값" 형태로 여러 셀에 값을 쓰는 헬퍼.
        예:
            write_cells_to_excel(
                "쿠팡업로드.xlsx",
                "data",
                {
                    "A10": "상품명",
                    "B10": "옵션명",
                    "C10": "카테고리ID",
                }
            )
        """
        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 를 엑셀에서 찾을 수 없습니다.")

        ws = wb[sheet_name]

        for addr, value in cell_value_map.items():
            ws[addr] = value

        wb.save(xlsx_path)

    # ====== 카테고리 정보 엑셀 쓰기 헬퍼 ======
    def write_category_info_to_excel_row(
        xlsx_path: str,
        sheet_name: str,
        row: int,
        category_info: Mapping[str, Any],
        column_mapping: Mapping[str, str],
    ) -> None:
        """
        category_loader.get_category_info() 의 dict 를
        엑셀 특정 행(row)에 써 넣는 헬퍼.

        예시:
            info = get_category_info(80289)
            write_category_info_to_excel_row(
                "쿠팡업로드.xlsx",
                "data",
                row=10,
                category_info=info,
                column_mapping={
                    "category_id": "C",        # C10 셀에 category_id
                    "category_path": "D",      # D10 셀에 category_path
                    "level1": "E",             # E10 에 level1
                    "level2": "F",
                    "level3": "G",
                    "level4": "H",
                    # 필요하면 col_c ~ col_j 도 매핑 가능
                    "col_c": "J",
                    "col_d": "K",
                    # ...
                },
            )
        """
        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 를 엑셀에서 찾을 수 없습니다.")
        ws = wb[sheet_name]

        for key, col_letter in column_mapping.items():
            if key not in category_info:
                continue
            col_index = column_index_from_string(col_letter)
            ws.cell(row=row, column=col_index, value=category_info[key])

        wb.save(xlsx_path)


# digits_only, is_int_string 등 유틸 함수는 config.py에서 import 하세요.


