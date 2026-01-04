# cellon/category_ai/category_worker.py
import os
import shutil
import openpyxl
from copy import copy
# 기존 import 문 아래에 추가

from pathlib import Path
from typing import Optional

import pandas as pd
from PyQt6.QtCore import QThread, pyqtSignal

from .category_loader import get_category_master
from .category_rules_builder import build_coupang_rules_for_all_groups

from .category_ai_meta_rules_builder import build_meta_rules


class CategoryBuildWorker(QThread):
    """
    카테고리 엑셀들을 읽어서 마스터 테이블을 만드는 워커(QThread).
    - 진행 상황은 progress(int, str) 시그널로 UI에 전달
    - 완료 시 finished(DataFrame) 시그널로 결과 전달
    - 오류 시 error(str) 시그널로 메시지 전달
    """
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(object)  # pd.DataFrame
    error = pyqtSignal(str)

    def __init__(self, category_dir: Path, parent=None):
        super().__init__(parent)
        self.category_dir = Path(category_dir)

    def run(self):
        try:
            def _cb(p: int, msg: str):
                # category_loader 에서 progress_cb 로 호출되면
                # 여기서 시그널로 UI에 넘김
                self.progress.emit(int(p), str(msg))

            # 1) category_excels → category_master DataFrame 생성
            df: pd.DataFrame = get_category_master(
                category_dir=self.category_dir,
                progress_cb=_cb
            )

            # 1) category_excels → category_master DataFrame 생성
            df: pd.DataFrame = get_category_master(
                category_dir=self.category_dir,
                progress_cb=_cb
            )

            # ✅ [PATCH] 1차 meta 룰(엑셀/index 기반) 자동 보강: meta/coupang_{group}.json
            try:
                self.progress.emit(90, "meta 룰 JSON(엑셀/index 기반) 보강 중...")
                project_root = Path(__file__).resolve().parents[3]  # .../src/cellon/category_ai/.. -> repo root
                build_meta_rules(
                    index_path=project_root / "assets/cache/coupang_upload_index.json",
                    meta_dir=project_root / "src/cellon/rules/meta",
                    group=None,          # kitchen/food 모두(감지된 그룹)
                    dry_run=False,
                    verbose=False,
                )
                self.progress.emit(92, "meta 룰 JSON 보강 완료")
            except Exception as e:
                self.error.emit(f"[meta_rules_builder] meta 룰 보강 실패: {e}")


            # 2) 방금 만든 df 를 기반으로 coupang 룰 JSON 자동 생성
            try:
                self.progress.emit(95, "coupang 룰 JSON 자동 생성 중...")
                build_coupang_rules_for_all_groups(df)
                self.progress.emit(100, "coupang 룰 JSON 생성 완료")
            except Exception as e:
                # 룰 생성이 실패해도 전체 워커가 죽지 않게 로그만 남김
                self.error.emit(f"[rules_builder] coupang 룰 생성 실패: {e}")

            # 3) UI 쪽으로 최종 df 전달
            self.finished.emit(df)

        except Exception as e:
            self.error.emit(str(e))
            
            
    # CategoryWorker 클래스 내부 혹은 유틸리티 섹션에 삽입
    def save_to_excel_template(self, product, matched_category):
        """
        설명하신 로직: 템플릿 복사 -> 행 탐색 -> 데이터 삽입
        """
        base_dir = "assets/crawling_temp"
        # 1. 원본 파일 식별 (카테고리에 따라 폴더/파일명 결정)
        # 현재 구조상 '주방용품' 등의 대분류를 폴더명으로 활용
        main_cat = matched_category.split('>')[0].strip() 
        source_folder = os.path.join(base_dir, "coupang_upload_form", main_cat)
        
        # 예시로 '주방용품>조리용품' 파일을 타겟으로 설정 (로직에 따라 동적 변경 가능)
        source_file = "sellertool_upload_주방용품>조리용품.xlsm" 
        source_path = os.path.join(source_folder, source_file)
        
        # 2. 결과 저장 경로 설정
        target_dir = os.path.join(base_dir, "upload_ready")
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
        
        target_path = os.path.join(target_dir, f"ready_{product.title[:10]}.xlsm")
        
        # 3. 파일 복사 및 엑셀 작업
        shutil.copy2(source_path, target_path)
        wb = openpyxl.load_workbook(target_path, keep_vba=True)
        ws = wb.active
        
        # 4. 행 찾기 (A열: 카테고리, CK열: 기타제화)
        target_row = None
        for r in range(1, 2000):
            if str(ws.cell(row=r, column=1).value) == matched_category:
                ck_val = str(ws.cell(row=r, column=89).value) # CK열
                if "기타제화" in ck_val or "기타 제화" in ck_val:
                    target_row = r
                    break
        
        if target_row:
            # 5. 최상단 삽입 및 데이터 기록
            ws.insert_rows(1, amount=2)
            ws.cell(row=1, column=1).value = "--- New Product Data ---"
            
            # 데이터 매핑 (Product 객체의 속성 활용)
            src_idx = target_row + 2
            ws.cell(row=2, column=1).value = matched_category
            ws.cell(row=2, column=2).value = product.title
            ws.cell(row=2, column=5).value = product.price
            # ... (이전 코드의 매핑 로직 적용)
            
        wb.save(target_path)

