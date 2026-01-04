# cellon/sellertool_excel.py
from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from copy import copy
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    COUPANG_UPLOAD_FORM_DIR,
    UPLOAD_READY_DIR,
    SELLERTOOL_SHEET_NAME,
    COUPANG_UPLOAD_INDEX_JSON,
)
from .core.product import Product

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from typing import Dict, Tuple

from copy import copy
from openpyxl.worksheet.worksheet import Worksheet

import zipfile
from typing import Iterable, Optional

# ===============================
# 템플릿 파일명 접두어(prefix) 추출 + 이미지명 생성
# ===============================

import re
from pathlib import Path

_PREFIX_RE = re.compile(r"^sellertool_upload_(?P<prefix>\d{1,3}-\d{1,3})_", re.IGNORECASE)



_WB_CACHE: Dict[str, Tuple[Workbook, float]] = {}
# key: str(dest_path), value: (workbook, last_mtime)

# ===============================
# 입력 행 판별 (A/B/C 기준)
# ===============================

def _cell_str(ws, row: int, col_letter: str) -> str:
    col = column_index_from_string(col_letter)
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def is_empty_row_abc(ws, row: int) -> bool:
    """
    A/B/C 중 2개 이상 비어있으면 입력 가능한 빈 행
    """
    vals = [
        _cell_str(ws, row, "A"),
        _cell_str(ws, row, "B"),
        _cell_str(ws, row, "C"),
    ]
    empty_cnt = sum(1 for x in vals if x == "")
    return empty_cnt >= 2


def find_next_input_row(ws, start_row: int, max_scan: int = 5000) -> int:
    for r in range(start_row, start_row + max_scan):
        if is_empty_row_abc(ws, r):
            return r
    raise RuntimeError("ABC 기준 입력 가능한 빈 행을 찾지 못했습니다.")

# --- prefix 로 upload ready 에 들어갈 xlsm file 의 번호 관련 코드 ----
def extract_template_prefix_from_filename(xlsm_path: Path) -> str | None:
    """
    예:
      sellertool_upload_14-10_주방용품>취사도구.xlsm -> '14-10'
    """
    name = xlsm_path.name
    m = _PREFIX_RE.match(name)
    if not m:
        return None
    return m.group("prefix")

def build_prefixed_image_names(prefix: str, row_idx: int) -> tuple[str, str]:
    """
    예:
      prefix='14-10', row_idx=125
    결과:
      14-10_125.png
      14-10_125_spec.png
    """
    base = f"{prefix}_{row_idx}"
    return f"{base}.png", f"{base}_spec.png"


# ===============================
# Template source / 구분자
# ===============================

def find_separator_row(ws, keyword="여기서부터", max_scan: int = 5000) -> int:
    """
    ✅ 구분자 행을 찾는다.
    - A열에서 keyword 포함 문구를 찾으면 그 행 번호 리턴
    - 없으면: 템플릿 마지막 행 바로 아래에 구분자 행을 자동 삽입하고 그 행 번호 리턴

    이유:
    - 템플릿마다 구분자 존재 여부가 다름
    - 최초 기록 시 구분자가 없는 경우가 흔함
    """
    # 1) 기존 구분자 탐색 (A열)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and keyword in v:
            return r

    # 2) 없으면 자동 삽입: "템플릿 마지막(=A열 값이 있는 마지막 행)" 아래에 삽입
    last_template_row = 0
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            last_template_row = r
            break

    # 템플릿이 비어있는 예외 케이스
    if last_template_row <= 0:
        last_template_row = 1

    sep_row = last_template_row + 1

    sep_text = "------------------ 여기서부터 크롤링 데이터 등록 ------------------"
    ws.cell(row=sep_row, column=1).value = sep_text

    # (선택) 구분자 강조(노란색)
    try:
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=sep_row, column=1).fill = yellow_fill
    except Exception:
        pass

    return sep_row



def find_template_source_row(
    ws,
    *,
    coupang_category_id: str | None = None,
    coupang_category_path: str | None = None,
    ck_candidates=("기타 재화", "기타재화"),
    template_source_max_row: int | None = None,
    max_scan: int = 1000,
) -> int:
    """
    Template source 행 선택 정책(개선):
    - 스캔 상한(upper)은 'template source 영역'까지만
      * template_source_max_row가 주어지면 그걸 사용
      * 없으면 get_template_source_max_row(ws)로 계산(구분자/없음 모두 대응)

    선택 우선순위:
    1) category_id 일치 행들을 모두 모음
       - 그 중 CK가 '기타 재화/기타재화'면 우선 선택(뒤에서부터 탐색)
       - 없으면 "마지막 id 행" 선택 (식품에서 농수산물=앞, 기타류=뒤 패턴 대응)
    2) (옵션) category_path 포함 행들 중 "마지막 매칭 행" 선택
    3) fallback: CK가 '기타 재화/기타재화'인 행들 중 "마지막" 선택
    """
    if template_source_max_row is None:
        template_source_max_row = get_template_source_max_row(ws)

    # 안전장치: template_source_max_row가 비정상인 경우만 백업 제한
    upper = template_source_max_row if template_source_max_row and template_source_max_row > 0 else max_scan
    if template_source_max_row is None:
        upper = min(upper, max_scan)

    upper = max(1, upper)

    # 1) category_id 우선 (요청 정책 반영)
    if coupang_category_id:
        token = f"[{coupang_category_id}]"
        id_rows: list[int] = []
        for r in range(1, upper + 1):
            a_val = ws[f"A{r}"].value
            if isinstance(a_val, str) and token in a_val:
                id_rows.append(r)

        if id_rows:
            # CK='기타 재화'가 있으면 우선(보통 뒤쪽에 있어서 reversed 탐색)
            for r in reversed(id_rows):
                ck_val = ws[f"CK{r}"].value
                if isinstance(ck_val, str) and ck_val.strip() in ck_candidates:
                    return r

            # 없으면 "마지막 id 행"
            return id_rows[-1]

    # 2) category_path (보조): 여러 개면 "마지막 매칭 행"
    if coupang_category_path:
        path_rows: list[int] = []
        for r in range(1, upper + 1):
            a_val = ws[f"A{r}"].value
            if isinstance(a_val, str) and coupang_category_path in a_val:
                path_rows.append(r)
        if path_rows:
            return path_rows[-1]

    # 3) fallback: CK 기준으로 마지막 후보
    ck_rows: list[int] = []
    for r in range(1, upper + 1):
        ck_val = ws[f"CK{r}"].value
        if isinstance(ck_val, str) and ck_val.strip() in ck_candidates:
            ck_rows.append(r)
    if ck_rows:
        return ck_rows[-1]

    # 진짜 최후: 1행(또는 2행)을 반환하기보다는 upper의 마지막으로(오탐 최소화)
    return upper



_ID_TOKEN_RE = re.compile(r"\[(\d+)\]")  # A열의 "[73134]" 같은 토큰 감지


def detect_separator_row(ws, keyword: str = "여기서부터", scan_limit: int = 5000) -> Optional[int]:
    """
    구분자 행을 '찾기만' 합니다. (삽입 X)
    - A열에서 keyword 포함 문구를 찾으면 그 행 번호 반환
    - 없으면 None
    """
    upper = min(getattr(ws, "max_row", scan_limit), scan_limit)
    for r in range(1, upper + 1):
        v = ws[f"A{r}"].value
        if isinstance(v, str) and keyword in v:
            return r
    return None


def infer_template_source_max_row(ws, scan_limit: int = 50000, blank_run_stop: int = 200) -> int:
    """
    구분자가 없는 '템플릿 소스만 있는 파일'에서,
    템플릿 소스 영역의 마지막 행을 추정합니다.

    전략:
    - A열에서 "[숫자]" 토큰이 등장하는 행들을 찾고, 마지막 발견 행을 max_row로 사용
    - 템플릿이 시작된 이후 A열이 연속으로 blank_run_stop만큼 비면 종료(시트 끝까지 스캔 방지)
    """
    max_row = getattr(ws, "max_row", 0) or 0
    upper = min(max_row if max_row > 0 else scan_limit, scan_limit)

    last_id_row = 0
    seen_any = False
    blank_run = 0

    for r in range(1, upper + 1):
        v = ws[f"A{r}"].value
        s = v.strip() if isinstance(v, str) else ""

        if _ID_TOKEN_RE.search(s):
            last_id_row = r
            seen_any = True
            blank_run = 0
            continue

        # 템플릿 시작 이후 공백이 오래 지속되면 종료
        if seen_any and (s == ""):
            blank_run += 1
            if blank_run >= blank_run_stop:
                break
        elif seen_any:
            # ID 토큰은 없지만 뭔가 적혀있으면(설명행 등) blank_run 리셋
            blank_run = 0

    # 마지막 id 행이 없으면(정말 비정상 템플릿) 안전하게 1000행 정도로 제한
    return last_id_row if last_id_row > 0 else min(upper, 1000)


def get_template_source_max_row(ws, *, keyword: str = "여기서부터") -> int:
    """
    ✅ '끝까지 허용'의 끝 = template source 영역의 끝
    - 구분자 있으면: sep_row - 1
    - 구분자 없으면: infer_template_source_max_row()로 마지막 [id] 행까지
    """
    sep = detect_separator_row(ws, keyword=keyword)
    if sep is not None and sep > 1:
        return sep - 1
    return infer_template_source_max_row(ws)

# ===============================
# Template source 보호 write
# ===============================

def safe_set_cell(ws, row: int, col: str, value, template_source_max_row: int):
    if row <= template_source_max_row:
        raise RuntimeError(
            f"Template source 영역({row})에 write 시도 차단: {col}{row}"
        )
    ws[f"{col}{row}"].value = value



# =========================
# 유틸
# =========================

def _safe_str(v) -> str:
    if v is None:
        return ""
    try:
        return str(v)
    except Exception:
        return ""


def _normalize_category_text(text: str) -> str:
    """
    템플릿 파일명과 category_path 를 최대한 유연하게 매칭하기 위한 정규화.

    - 공백 제거
    - 한글, 숫자, 영문만 남김
    - '/', ':', '>' 등을 모두 '>' 로 통합
    - 유니코드 정규화 적용
    """
    if not text:
        return ""

    import unicodedata
    t = unicodedata.normalize("NFKC", text)

    # 구분자 통합
    t = t.replace("/", ">").replace(":", ">")

    # 소문자
    t = t.lower()

    # 공백 제거
    t = t.replace(" ", "")

    # 한글/영문/숫자/구분자만 남김
    allowed = []
    for ch in t:
        if ch.isalnum() or ch == ">":
            allowed.append(ch)
    t = "".join(allowed)

    return t

# ========================
# 가격 계산 정책
# ========================

# 10의 자리 절삭 (백원 단위로 내림)
def _floor_to_100(x: int) -> int:
    # 1의 자리 절삭 (백원 단위로 내림)
    return (x // 100) * 100


def calculate_pricing_from_base(base_price: int) -> tuple[int, int, int, int]:
    """
    ui_main.py 에 있던 기존 가격 정책을 그대로 사용

    반환:
      (bj_price, bl_price, stock_qty, lead_time)
    """
    if base_price <= 0:
        return 0, 0, 0, 0

    # BJ: 판매가
    if base_price <= 10000:
        bj_price = int(round(base_price * 1.8))
    elif base_price <= 30000:
        bj_price = int(round(base_price * 1.3))
    elif base_price <= 50000:
        bj_price = int(round(base_price * 1.2))
    else:
        bj_price = int(round(base_price * 1.15))

    # 10의 자리 절삭
    bj_price = _floor_to_100(bj_price)
    
    # BL: 할인 기준가 (기존 로직: BJ * 1.05)
    bl_price = int(round(bj_price * 1.05)) if bj_price > 0 else 0
    # 10의 자리 절삭
    bl_price = _floor_to_100(bl_price)

    # BM / BN: 기존 고정 정책
    stock_qty = 999
    lead_time = 2 # costco 기준. 2~3일 소요

    return bj_price, bl_price, stock_qty, lead_time


# =========================
# 1) 템플릿 인덱스 (파일명 → 절대경로 폴더 내 재귀검색)
# =========================

@lru_cache(maxsize=1)
def _build_template_index() -> dict[str, Path]:
    """
    쿠팡 업로드 템플릿 인덱스를 생성한다.

    기본 정책:
    - (B) JSON 인덱스가 있으면: JSON 을 신뢰하고 빠르게 로드한다. (추천)
    - (A) JSON 인덱스가 없거나/깨졌거나/비어 있으면: rglob() 재귀 탐색으로 백업한다.

    - JSON 포맷:
      {
        "templates": [
          { "key": "주방용품>취사도구", "relative_path": "주방용품/..." },
          ...
        ]
      }
    """
    root = COUPANG_UPLOAD_FORM_DIR

    # -------------------------
    # (B) JSON 인덱스 우선 로드
    # -------------------------
    if COUPANG_UPLOAD_INDEX_JSON.exists():
        try:
            with COUPANG_UPLOAD_INDEX_JSON.open("r", encoding="utf-8") as f:
                data = json.load(f)

            templates = data.get("templates", [])
            if not templates:
                # JSON은 있으나 내용이 비어있음 → 기존처럼 강하게 안내 + (A) 백업 시도
                print(
                    f"[WARN] 쿠팡 업로드 템플릿 인덱스가 비어 있습니다: {COUPANG_UPLOAD_INDEX_JSON}\n"
                    "카테고리 분석을 다시 실행하는 것을 권장합니다. (백업 탐색을 시도합니다)"
                )
            else:
                index: dict[str, Path] = {}
                missing_count = 0

                for item in templates:
                    key = item.get("key")
                    rel_raw = item.get("relative_path")

                    if not key or not rel_raw:
                        # 포맷 이상 항목은 건너뛰되 경고만 남김
                        print(f"[WARN] 잘못된 템플릿 인덱스 항목을 건너뜁니다: {item}")
                        continue

                    rel = Path(rel_raw)
                    abs_path = (root / rel).resolve()
                    index[key] = abs_path

                    # 파일이 실제로 없으면 카운트만 하고 계속 (나중에 백업 여부 결정)
                    if not abs_path.exists():
                        missing_count += 1

                # JSON 기반 인덱스가 실질적으로 유효하면 그대로 사용
                if index and missing_count == 0:
                    return index

                # 일부/전체 경로가 깨진 경우 → 경고 후 (A) 백업 탐색
                if index and missing_count > 0:
                    print(
                        "[WARN] 쿠팡 템플릿 인덱스 JSON은 있으나, 실제 파일 경로가 누락된 항목이 있습니다.\n"
                        f"- 누락 항목 수: {missing_count}\n"
                        f"- JSON 경로: {COUPANG_UPLOAD_INDEX_JSON}\n"
                        "해결:\n"
                        "1) 템플릿 파일 이동/삭제 여부를 확인하거나\n"
                        "2) build_coupang_upload_index.py 를 다시 실행해 인덱스를 재생성하세요.\n"
                        "우선 백업 탐색(rglob)을 시도합니다."
                    )

                # JSON이 있긴 하나 결과가 비어 있거나 경로가 깨짐 → 백업으로 진행
        except Exception as e:
            # JSON 파손/인코딩 문제 등 → 기존처럼 강하게 안내 + (A) 백업 시도
            print(
                f"[WARN] 쿠팡 업로드 템플릿 인덱스 JSON 로드에 실패했습니다: {COUPANG_UPLOAD_INDEX_JSON}\n"
                f"- 원인: {repr(e)}\n"
                "해결:\n"
                "1) 카테고리 분석(또는 build_coupang_upload_index.py)을 다시 실행하거나\n"
                "2) JSON 파일이 정상인지 확인하세요.\n"
                "우선 백업 탐색(rglob)을 시도합니다."
            )

    else:
        # 기존 안전장치 메시지 톤을 유지하되, 즉시 종료하지 않고 백업을 시도
        print(
            f"[WARN] 쿠팡 업로드 템플릿 인덱스 JSON이 없습니다: {COUPANG_UPLOAD_INDEX_JSON}\n"
            "먼저 '카테고리 분석' 또는 build_coupang_upload_index.py 를 실행해서 "
            "인덱스를 생성하는 것을 권장합니다.\n"
            "우선 백업 탐색(rglob)을 시도합니다."
        )

    # -------------------------
    # (A) 백업: rglob 재귀 탐색
    # -------------------------
    index: dict[str, Path] = {}
    for path in root.rglob("sellertool_upload_*.xlsm"):
        key = path.stem.replace("sellertool_upload_", "")
        index[key] = path.resolve()

    if index:
        return index

    # -------------------------
    # 최종 실패: 기존과 유사한 강한 에러
    # -------------------------
    raise RuntimeError(
        f"쿠팡 업로드 폼 템플릿을 찾지 못했습니다: {root}\n"
        "확인:\n"
        "1) coupang_upload_form 내의 쿠팡 템플릿 구조 아래에 sellertool_upload_*.xlsm 이 존재하는지\n"
        f"2) 인덱스 JSON({COUPANG_UPLOAD_INDEX_JSON})이 정상인지\n"
        "해결:\n"
        "1) '카테고리 분석' 또는 build_coupang_upload_index.py 로 인덱스를 생성하고\n"
        "2) 템플릿 파일들이 올바른 위치에 있는지 확인해 주세요."
    )




def find_template_for_category_path(category_path: str) -> Path:
    """
    쿠팡 카테고리 경로 (예: '주방용품>취사도구>냄비>양수냄비')
    에 가장 잘 매칭되는 템플릿 파일을 찾는다.

    전략:
      1) category_path 를 '>' 기준으로 자른 뒤
         가장 긴 prefix 부터 줄여가며(4뎁스 → 3뎁스 → 2뎁스 → 1뎁스)
         템플릿 key 와 매칭을 시도한다.
      2) prefix 와 템플릿 key 를 공백 제거한 문자열로 비교:
         - key_norm == prefix_norm
         - 또는 key_norm 이 prefix_norm 안에 포함
         - 또는 prefix_norm 이 key_norm 안에 포함
      3) 그래도 없으면, 전체 category_path 기준으로
         "가장 긴 부분 문자열로 겹치는" 기존 방식으로 한 번 더 시도.
      4) 그래도 없으면 KeyError 발생.
    -------------------------------------------------------
    ✅ 이 함수는 “쿠팡 카테고리 길(경로)”을 보고,
       그 카테고리에 맞는 “엑셀 템플릿 파일”을 고르는 역할을 합니다.

    예)
      category_path = '주방용품>취사도구>냄비>양수냄비'
      → 이 카테고리에 맞는 sellertool_upload_...xlsm 파일을 찾습니다.

    ⭐ 아주 중요한 규칙
      - 뎁스(단계)는 오직 '>' 기호로만 나뉩니다.
      - '/' 는 단계가 아닙니다. (그냥 글자일 뿐)
        예: '냄비/냄비세트' 는 “하나의 이름”입니다.
    """

    # 0) 템플릿 목록(인덱스)을 준비한다.
    #    index 예시:
    #      {
    #        "10-3_식품>건강식품>건강식품/영양식": Path("...xlsm"),
    #        "10-4_식품>건강식품>전통건강식품/헬스/다이어트": Path("...xlsm"),
    #        ...
    #      }
    index = _build_template_index()

    # 1) 입력이 비어있으면 바로 에러
    if not category_path:
        raise KeyError("카테고리 경로가 비어 있습니다.")

    # 2) '>' 기준으로 자른다. 이것이 곧 “뎁스(단계)”가 된다.
    #    예: '식품>건강식품>전통건강식품/헬스/다이어트'
    #       parts = ['식품', '건강식품', '전통건강식품/헬스/다이어트']
    parts = [p.strip() for p in category_path.split(">") if p.strip()]
    if not parts:
        raise KeyError(f"파싱할 수 없는 카테고리 경로입니다: {category_path}")

    # 3) 템플릿 key들도 “비교하기 쉽게” 미리 정리해 둔다(정규화).
    #    _normalize_category_text()는 보통
    #      - 공백 제거
    #      - 특수문자 처리
    #    같은 것을 해서 비교를 쉽게 만듭니다.
    norm_index: dict[str, str] = {
        key: _normalize_category_text(key) for key in index.keys()
    }

    # ------------------------------------------------------------
    # 4) 여기부터가 핵심:
    #    “긴 경로부터” 먼저 맞춰보고, 안 맞으면 “조금씩 짧게” 줄여서 다시 맞춰본다.
    #
    #    예: parts = ['식품', '건강식품', '전통건강식품/헬스/다이어트']
    #      depth=3 : '식품>건강식품>전통건강식품/헬스/다이어트'  (가장 자세함)
    #      depth=2 : '식품>건강식품'                           (덜 자세함)
    #      depth=1 : '식품'                                    (너무 넓음)
    #
    #    ✅ 원칙: “가능하면 3뎁스 이상에서 딱 맞추는 게 가장 안전”합니다.
    # ------------------------------------------------------------
    for depth in range(len(parts), 0, -1):

        prefix = ">".join(parts[:depth])
        prefix_norm = _normalize_category_text(prefix)

        # --------------------------------------------------------
        # 4-1) 1순위: exact match (정확히 똑같이 맞는지)
        # --------------------------------------------------------
        exact_matches: list[str] = [
            key for key, key_norm in norm_index.items()
            if key_norm == prefix_norm
        ]
        if exact_matches:
            # 여러 개면 가장 “긴 key” 선택
            # (보통 더 자세한 템플릿이 길어서 이게 유리함)
            best_key = max(exact_matches, key=lambda k: len(norm_index[k]))
            return index[best_key]

        # --------------------------------------------------------
        # 4-2) 2순위: 포함관계 match (서로 포함되어 있으면 후보)
        # --------------------------------------------------------
        candidates: list[str] = [
            key for key, key_norm in norm_index.items()
            if key_norm in prefix_norm or prefix_norm in key_norm
        ]
        if candidates:
            # 후보가 여러 개면 가장 “긴 key” 선택
            best_key = max(candidates, key=lambda k: len(norm_index[k]))
            return index[best_key]

        # --------------------------------------------------------
        # ⭐⭐⭐ 여기서 2뎁스 위험성이 발생할 수 있는 이유 ⭐⭐⭐
        #
        # depth가 2가 되었을 때 prefix는 보통 "대분류>2뎁스" 수준입니다.
        # 예: '식품>건강식품'
        #
        # 그런데 템플릿 파일이 이런 식으로 여러 개 있을 수 있습니다:
        #   - '10-3_식품>건강식품>건강식품/영양식'
        #   - '10-4_식품>건강식품>전통건강식품/헬스/다이어트'
        #
        # 이 둘은 둘 다 prefix('식품>건강식품')를 포함합니다.
        # 그러면 candidates가 2개가 되고,
        # 코드가 "그냥 길이가 더 긴 것"을 골라 버립니다.
        #
        # ✅ 만약 우리가 원하는 건 '영양식' 템플릿인데
        #    우연히 '전통건강식품...' 템플릿이 더 길면
        #    잘못된 엑셀을 선택할 수도 있습니다.
        #
        # 👉 그래서 “3뎁스 이상에서 매칭이 실패해서 2뎁스까지 내려오는 순간”
        #    오선택 위험이 생깁니다.
        #
        # (즉, 2뎁스가 중복된 폴더 구조 자체가 문제라기보다,
        #  2뎁스까지만 정보가 남아 매칭될 때가 위험한 순간입니다.)
        # --------------------------------------------------------

    # ------------------------------------------------------------
    # 5) 그래도 못 찾으면, 마지막 “응급처치 fallback”:
    #    전체 category_path와 템플릿 key가
    #    서로 포함관계인지 보면서 “가장 긴 것”을 고른다.
    #
    #    이건 정확도가 떨어질 수 있어서
    #    가능하면 위에서(3뎁스 이상) 잡히는 게 좋습니다.
    # ------------------------------------------------------------
    category_norm = _normalize_category_text(category_path)
    best_key = None
    best_len = -1

    for key, key_norm in norm_index.items():
        if key_norm and (key_norm in category_norm or category_norm in key_norm):
            if len(key_norm) > best_len:
                best_key = key
                best_len = len(key_norm)

    if best_key is not None:
        return index[best_key]

    # 6) 완전히 실패하면: 어떤 템플릿이 있는지 보여주고 에러
    available = ", ".join(sorted(index.keys()))
    raise KeyError(
        f"카테고리 경로에 맞는 템플릿을 찾지 못했습니다: {category_path} "
        f"(사용 가능한 템플릿 key: {available})"
    )

# =========================
# 2) data 시트 헬퍼
# =========================

def _get_header_col(ws: Worksheet, header_text: str) -> Optional[int]:
    """2행(헤더행)에서 header_text 와 정확히 일치하는 컬럼 index 를 찾는다."""
    for cell in ws[2]:
        if _safe_str(cell.value) == header_text:
            return cell.column
    return None


def _find_column_contains(ws: Worksheet, keyword: str) -> Optional[int]:
    """2행(헤더행)에서 keyword 를 포함하는 컬럼 index 를 찾는다."""
    for cell in ws[2]:
        val = _safe_str(cell.value)
        if keyword in val:
            return cell.column
    return None

def _get_target_insertion_row(ws: Worksheet) -> int:
    """
    [새로 추가된 함수]
    파일마다 다른 양식 구간(Template Source)의 끝을 찾아내고,
    그 아래에 구분자(Divider)를 넣은 뒤 실제 데이터가 들어갈 행 번호를 반환합니다.
    """
    # 1. 시트에서 실제로 데이터가 있는 마지막 행 찾기 (A열 기준 역순 탐색)
    # 엑셀 파일마다 10행일수도, 128행일수도, 300행일수도 있는 '양식 끝'을 동적으로 찾습니다.
    last_template_row = 0
    for r in range(ws.max_row, 0, -1):
        if _safe_str(ws.cell(row=r, column=1).value):
            last_template_row = r
            break
            
    # 2. 이미 구분자("---")가 포함된 행이 있는지 확인 (이미 상품이 하나 이상 추가된 경우)
    divider_found_row = 0
    for r in range(1, ws.max_row + 1):
        val = _safe_str(ws.cell(row=r, column=1).value)
        if "------------------" in val:
            divider_found_row = r
            break

    # 3. 상황에 따른 목적지 행(Destination Row) 결정
    if divider_found_row > 0:
        # CASE A: 이미 구분자가 있음 -> 구분자 아래에서 첫 번째 빈 행을 찾아 이어서 작성
        curr_row = divider_found_row + 1
        while True:
            if not _safe_str(ws.cell(row=curr_row, column=1).value):
                return curr_row
            curr_row += 1
    else:
        # CASE B: 구분자가 없음 (최초 작성) -> 양식 끝 바로 다음 행에 구분자 삽입
        divider_row = last_template_row + 1
        ws.cell(row=divider_row, column=1).value = "------------------ 여기서부터 크롤링 데이터 등록 ------------------"
        
        # 가독성을 위해 노란색 배경색(PatternFill) 추가 (선택사항)
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=divider_row, column=1).fill = yellow_fill
        
        # 데이터는 구분자 바로 다음 행(양식끝 + 2)부터 시작
        return divider_row + 1


#삭제필요
def _pick_template_row(
    ws: Worksheet,
    category_id: str,
    category_path: str,
) -> int:
    """
    - A열에 '[category_id]' 가 들어있는 행들을 후보로 모으고
    - 그 중 CK열('상품고시정보 카테고리') 값이
      카테고리 최상단(예: '주방용품')과 맞는 행을 우선 선택.
    """
    id_token = f"[{category_id}]"
    candidates: list[int] = []

    max_row = ws.max_row
    for r in range(5, max_row + 1):
        val = _safe_str(ws.cell(row=r, column=1).value)
        if id_token in val:
            candidates.append(r)

    if not candidates:
        raise RuntimeError(f"data 시트에서 category_id={category_id} 행을 찾지 못했습니다.")

    # CK열 (상품고시정보 카테고리) 인덱스 찾기
    ck_col = _find_column_contains(ws, "상품고시정보 카테고리")
    top_level = _safe_str(category_path).split(">")[0].strip()

    if ck_col and top_level:
        for r in candidates:
            ck_val = _safe_str(ws.cell(row=r, column=ck_col).value)
            if top_level in ck_val:
                return r

    # 적절한 CK 매칭이 없으면 그냥 첫 번째 후보 사용
    return candidates[0]


# ===== DataValidation (드롭다운) 캐시: src_row에 걸린 DV만 추출해서 재사용 =====

@dataclass(frozen=True)
class _DVSpan:
    dv: object
    min_col: int
    max_col: int

# key: id(ws) -> (src_row, spans)
_DV_TEMPLATE_CACHE: dict[int, tuple[int, list[_DVSpan]]] = {}


def _get_dv_template_for_src_row(ws: Worksheet, src_row: int) -> list[_DVSpan]:
    """
    src_row에 걸린 DV만 추려서 캐싱합니다.
    - 템플릿 DV가 수천개여도, src_row에 걸린 건 보통 10~20개대라서
      write가 반복될수록 성능이 크게 좋아집니다.
    """
    key = id(ws)
    cached = _DV_TEMPLATE_CACHE.get(key)
    if cached and cached[0] == src_row:
        return cached[1]

    spans: list[_DVSpan] = []
    dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
    for dv in dvs:
        try:
            ranges = list(dv.sqref.ranges)  # 스냅샷
        except Exception:
            continue

        for r in ranges:
            if r.min_row <= src_row <= r.max_row:
                spans.append(_DVSpan(dv=dv, min_col=r.min_col, max_col=r.max_col))

    _DV_TEMPLATE_CACHE[key] = (src_row, spans)
    return spans


def _dv_has_addr(dv, addr: str) -> bool:
    """중복 add 방지용(가벼운 체크)"""
    try:
        return addr in str(dv.sqref).split()
    except Exception:
        return False


def _copy_row_full(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 120) -> None:
    """
    src_row → dst_row 로 값, 스타일, 데이터 유효성(드롭다운)까지 복사.
    - row height 복사 (있을 때만)
    - 스타일은 _style 전체 복사로 누락 최소화
    - DV는 src_row에 걸린 것만 캐싱하여 성능 개선
    - DV add 중복 방지
    """

    # 0) row height 복사(지정된 경우만)
    try:
        h = ws.row_dimensions[src_row].height
        if h is not None:
            ws.row_dimensions[dst_row].height = h
    except Exception:
        pass

    # 1) 값 + 스타일 복사
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)

        dst_cell.value = src_cell.value

        # 스타일 전체 복사(가장 안정적)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)

        # 필요 시만 유지(대부분 템플릿에서 큰 영향은 없지만 안전)
        if src_cell.hyperlink:
            dst_cell.hyperlink = copy(src_cell.hyperlink)
        if src_cell.comment:
            dst_cell.comment = copy(src_cell.comment)

    # 2) 데이터 유효성(드롭다운) 복사: src_row에 걸린 것만 캐싱해서 적용
    spans = _get_dv_template_for_src_row(ws, src_row)
    for span in spans:
        dv = span.dv
        for col in range(span.min_col, span.max_col + 1):
            addr = f"{get_column_letter(col)}{dst_row}"
            if _dv_has_addr(dv, addr):
                continue
            try:
                dv.add(addr)
            except Exception:
                pass


def _fill_product_data(
    ws: Worksheet,
    row: int,
    *,
    product: Product,
    price: Optional[int] = None,
    search_keywords: Optional[Iterable[str]] = None,
) -> None:
    """
    템플릿 복사된 행(row)에 실제 상품 데이터를 채워 넣는다.
    - 등록상품명
    - 판매가격
    - 검색어
    등 기본적인 것만 우선 채우고, 나머지는 나중에 확장.
    """
    # 1) 헤더 컬럼 인덱스 찾기
    col_name = _get_header_col(ws, "등록상품명")
    col_price = _get_header_col(ws, "판매가격")
    col_search = _get_header_col(ws, "검색어")

    # 2) 상품명
    if col_name:
        ws.cell(row=row, column=col_name).value = product.display_name

    # 3) 판매가격
    if col_price is not None and price is not None:
        try:
            ws.cell(row=row, column=col_price).value = int(price)
        except Exception:
            ws.cell(row=row, column=col_price).value = _safe_str(price)

    # 4) 검색어
    if col_search is not None and search_keywords:
        joined = ", ".join([_safe_str(k) for k in search_keywords if _safe_str(k)])
        ws.cell(row=row, column=col_search).value = joined



def _get_cached_workbook(xlsm_path: Path) -> Workbook:
    """
    ✅ 안정성 우선 정책:
    - 크롤링 1건(또는 저장 1회)마다 workbook을 새로 열고, 저장 후 닫는다.
    - openpyxl workbook 객체 재사용(캐시)은 xlsm zip 깨짐/닫힌 핸들 이슈를 유발할 수 있어
      당분간 비활성화한다.
    """
    return load_workbook(xlsm_path, keep_vba=True)


def _validate_xlsm_zip(xlsm_path: Path) -> None:
    """
    저장 직후 xlsm(zip) 기본 구조가 유지되는지 빠르게 검증.
    - "[Content_Types].xml" 누락이면 xlsx/xlsm로서 성립 불가 → 바로 감지
    """
    with zipfile.ZipFile(xlsm_path, "r") as zf:
        names = set(zf.namelist())
        required = {"[Content_Types].xml", "_rels/.rels"}
        missing = [n for n in required if n not in names]
        if missing:
            raise RuntimeError(
                f"엑셀 파일이 손상되었습니다. 필수 엔트리 누락: {missing} (file={xlsm_path})"
            )

def _save_cached_workbook(xlsm_path: Path, wb: Workbook) -> None:
    """
    ✅ 안정성 우선 정책:
    - save → 유효성 검사 → close
    - 캐시 유지하지 않음(다음 크롤링에서 다시 open)
    """
    try:
        wb.save(xlsm_path)
        _validate_xlsm_zip(xlsm_path)
    finally:
        # openpyxl workbook 재사용을 막기 위해 항상 close
        try:
            wb.close()
        except Exception:
            pass
        _WB_CACHE.pop(str(xlsm_path), None)


def prepare_sellertool_workbook_copy(
    template_xlsm_path: Path,
    out_dir: Path,
    output_name: str | None = None,
    add_date_subdir: bool = False,
) -> Path:
    """
    ✅ output_name이 None이면 '템플릿 파일명 그대로' 복사/재사용
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if add_date_subdir:
        out_dir = out_dir / datetime.now().strftime("%Y-%m-%d")
        out_dir.mkdir(parents=True, exist_ok=True)

    if output_name:
        dest_path = out_dir / output_name
    else:
        dest_path = out_dir / Path(template_xlsm_path).name  # ✅ 파일명 유지

    # ✅ 이미 upload_ready에 있으면 복사하지 않음(재사용)
    if not dest_path.exists():
        shutil.copy2(template_xlsm_path, dest_path)

    return dest_path

# =========================
# data 시트 행 쓰기
# =========================
def write_coupang_row(
    ws,
    product_name: str,
    calculated_price: int,        # BJ (판매가)
    discount_base_price: int,     # BL (할인 기준가)
    stock_qty: int,               # BM
    lead_time: int,               # BN
    main_image_name: str,         # CZ
    spec_image_name: str,         # DF
    coupang_category_id: str | None = None,
    coupang_category_path: str | None = None,
):
    """
    - Template source 영역에는 절대 write 하지 않는다
    - 구분자 아래, ABC 기준 빈 행에만 append
    """

    # 1. 구분자 / Template source 영역
    sep_row = find_separator_row(ws)
    template_source_max_row = sep_row - 1

    # 2. Template source 행 (CK 기준)
    src_row = find_template_source_row(
        ws,
        coupang_category_id=coupang_category_id,
        coupang_category_path=coupang_category_path,
        template_source_max_row=template_source_max_row,
    )

    # 3. 입력 대상 행
    dst_row = find_next_input_row(ws, sep_row + 1)

    # 4. Template source → 입력 행 복사
    _copy_row_full(
        ws,
        src_row=src_row,
        dst_row=dst_row,
        max_col=ws.max_column,
    )

    # 5. 값 쓰기 (dst_row ONLY)
    today = datetime.now().strftime("%Y-%m-%d")

    safe_set_cell(ws, dst_row, "B", product_name, template_source_max_row)
    safe_set_cell(ws, dst_row, "C", today, template_source_max_row)

    # G/H: 등록상품명의 첫 단어(예: "AMT")
    first_word = (product_name.split()[0] if product_name and product_name.split() else "")
    safe_set_cell(ws, dst_row, "G", first_word, template_source_max_row)
    safe_set_cell(ws, dst_row, "H", first_word, template_source_max_row)

    # 가격
    safe_set_cell(ws, dst_row, "BJ", calculated_price, template_source_max_row)
    safe_set_cell(ws, dst_row, "BL", discount_base_price, template_source_max_row)

    # 사용자 지정값
    safe_set_cell(ws, dst_row, "BM", stock_qty, template_source_max_row)
    safe_set_cell(ws, dst_row, "BN", lead_time, template_source_max_row)

    # 이미지명
    safe_set_cell(ws, dst_row, "CZ", main_image_name, template_source_max_row)
    safe_set_cell(ws, dst_row, "DF", spec_image_name, template_source_max_row)

    return dst_row



def copy_row_with_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int=120):
    """
    src_row의 셀 값/스타일을 dst_row로 복사.
    - 값(value)
    - 스타일(font, fill, border, alignment, number_format, protection)
    - row 높이
    """
    # row height 복사
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for col in range(1, max_col + 1):
        c1 = ws.cell(row=src_row, column=col)
        c2 = ws.cell(row=dst_row, column=col)

        c2.value = c1.value

        if c1.has_style:
            c2.font = copy(c1.font)
            c2.fill = copy(c1.fill)
            c2.border = copy(c1.border)
            c2.alignment = copy(c1.alignment)
            c2.number_format = c1.number_format
            c2.protection = copy(c1.protection)
            c2.comment = c1.comment  # 필요하면 copy()로
    
    # 데이터 유효성(드롭다운) 복사
    for dv in ws.data_validations.dataValidation:
        for rng in dv.ranges:
            if rng.min_row <= src_row <= rng.max_row:
                for col in range(rng.min_col, rng.max_col + 1):
                    dv.add(f"{get_column_letter(col)}{dst_row}")


# =========================
# 3) 퍼블릭 API
# =========================

def prepare_and_fill_sellertool(
    *,
    product: Product,
    coupang_category_id: str,
    coupang_category_path: str,
    price: Optional[int] = None,
    search_keywords: Optional[Iterable[str]] = None,
) -> tuple[Path, int]:

    """
    1) coupang_upload_form 폴더에서 카테고리에 맞는 템플릿 엑셀을 찾고
    2) upload_ready 폴더로 '원래 파일명 그대로' 복사 (이미 있으면 재사용)
    3) data 시트에서 category_id/경로에 맞는 템플릿 행을 찾아
       첫 빈 행으로 복사하고
    4) 그 행에 product/price/search_keywords 를 채워 넣는다.

    최종적으로 수정된 upload_ready 안의 파일 Path 와,
    실제로 데이터가 기록된 행 번호(dst_row)를 함께 반환.
    """
    #for test
    
    print("[DEBUG] prepare_and_fill_sellertool called")
    
    # ---- 디버그 로그: 템플릿 인덱스 JSON 상태 확인 ----
    # 정책:
    # - JSON 이 있으면 (B) 인덱스 기반으로 빠르게 찾는다.
    # - JSON 이 없거나/깨졌으면 (A) rglob 백업 탐색으로 찾는다.
    if COUPANG_UPLOAD_INDEX_JSON.exists():
        print("[DEBUG] template index JSON exists:", COUPANG_UPLOAD_INDEX_JSON)
    else:
        print(
            "[WARN] 쿠팡 업로드 템플릿 인덱스 JSON이 없습니다.\n"
            f"- JSON 경로: {COUPANG_UPLOAD_INDEX_JSON}\n"
            "권장:\n"
            "1) '카테고리 분석' 또는 build_coupang_upload_index.py 를 실행해 "
            "인덱스를 생성해 주세요.\n"
            "우선 백업 탐색(rglob)으로 템플릿 선택을 시도합니다."
        )

    
        
    # ---- 1) 템플릿 선택 ----
    template_path = find_template_for_category_path(coupang_category_path)

    # ---- 2) upload_ready 폴더로 '원래 파일명' 그대로 복사 ----
    UPLOAD_READY_DIR.mkdir(parents=True, exist_ok=True)
    dest_name = template_path.name
    dest_path = UPLOAD_READY_DIR / dest_name

    # 같은 템플릿을 여러 번 쓰는 경우:
    # 이미 dest_path 가 있으면 복사하지 않고 기존 파일에 행만 추가
    if not dest_path.exists():
        shutil.copy2(template_path, dest_path)

    # ---- 3) 엑셀 열기 ----
    wb = _get_cached_workbook(dest_path)
    try:
        if SELLERTOOL_SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(
                f"시트 '{SELLERTOOL_SHEET_NAME}' 를 찾지 못했습니다. "
             f"파일: {dest_path}"
            )

        ws = wb[SELLERTOOL_SHEET_NAME]

        # ---- 4) 가격 정책 계산 (기존 ui_main.py 로직 재사용) ----
        base_price = int(price) if price is not None else 0

        bj_price, bl_price, stock_qty, lead_time = calculate_pricing_from_base(base_price)

        # ---- 5) 데이터 행 추가 (Template source 보호 로직 사용) ----
        dst_row = write_coupang_row(
            ws=ws,
            product_name=product.display_name,
            calculated_price=bj_price,        # BJ
            discount_base_price=bl_price,     # BL
            stock_qty=stock_qty,              # BM
            lead_time=lead_time,              # BN
            main_image_name="",               # 일단 빈 값(아래에서 채움)
            spec_image_name="",
            coupang_category_id=coupang_category_id,
            coupang_category_path=coupang_category_path,
        )
        # ✅ prefix 기반 이미지명 확정 → CZ/DF에 실제로 기록
        prefix = extract_template_prefix_from_filename(dest_path) or "no-prefix"
        main_img, spec_img = build_prefixed_image_names(prefix, dst_row)

        template_source_row = find_template_source_row(ws)
        template_source_max_row = find_template_source_row(ws)
        sep_row = find_separator_row(ws)
        template_source_max_row = sep_row - 1
        # template source 보호를 위해 구분선 기반으로 상한만 계산
        sep_row = find_separator_row(ws)
        template_source_max_row = sep_row - 1
        safe_set_cell(ws, dst_row, "CZ", main_img, template_source_max_row)
        safe_set_cell(ws, dst_row, "DF", spec_img, template_source_max_row)

    # ---- 7) 저장 ----
        _save_cached_workbook(dest_path, wb)
    finally:
        # _save_cached_workbook에서 close를 하더라도,
        # 중간 예외로 save까지 못 가는 경우를 대비한 2중 안전장치
        try:
            wb.close()
        except Exception:
            pass    
    
    print("[DEBUG] template_path =", template_path)
    print("[DEBUG] template_path.name =", template_path.name)
    print("[DEBUG] dest_path =", dest_path)
    print("[DEBUG] dest exists? =", dest_path.exists())

    return dest_path, dst_row


